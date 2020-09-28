
from django.shortcuts import render
from django.http import JsonResponse
from survey.models import *
from userinfo.models import batch
from userinfo.views import authenticate_credentials
from django.dispatch import receiver
from django.db.models.signals import post_save
from django.conf import settings
import json
from django.core.exceptions import ObjectDoesNotExist
import pandas
from .response import Response
from bson import ObjectId, json_util
from django.http import HttpResponse
#import datetime
from datetime import datetime, timedelta, timezone
from django.core.serializers import serialize
from django.core.files.storage import FileSystemStorage
import requests

from django.core.mail import EmailMultiAlternatives
from django.template.loader import get_template
from django.core import serializers
from survey.serializer import *

from itertools import groupby
from userinfo.utils.notification import Notification
import calendar


def getLaporan(request):

    if request.method == "POST":

        bulan = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
                 "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)
            month = data['month']
            year = data['year']

            pastMonth = data['month'] - 1
            pastYear = year

            if pastMonth == 0:
                pastMonth = 12
                pastYear = year-1

            ai = '5f16b4ba149882a98fc6655e'
            bts = '5f1521524f9c6764c713d73c'

            if month is None or year is None:
                return Response.badRequest(message='Need Json Body "month" & "year"')

            def pipelineDate(month, year, jenis):

                lastDay = calendar.monthrange(year=year, month=month)[1]
                return [
                    {
                        '$addFields': {
                            'last': {
                                '$arrayElemAt': [
                                    '$status', -1
                                ]
                            }
                        }
                    }, {
                        '$match': {
                            'jenissurvey': ObjectId(jenis),
                            'last.status': 'finished',
                            'last.date': {
                                '$gte': datetime(year, month, 1, 00, 00, 00, tzinfo=timezone.utc),
                                '$lte': datetime(year, month, lastDay, 23, 59, 59, tzinfo=timezone.utc)
                            }
                        }
                    }, {
                        '$count': 'count'
                    }
                ]

            def pipelineTotal(jenis, status):
                return [
                    {
                        '$match': {
                            'jenissurvey': ObjectId(jenis),
                            'status.status': status,
                        }
                    }, {
                        '$count': 'count'
                    }
                ]

            penugasanAITotal = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=ai, status='assigned')))[0]['count']
            penugasanBTSTotal = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=bts, status='assigned')))[0]['count']

            penugasanAITotalFinish = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=ai, status='finished')))[0]['count']
            penugasanBTSTotalFinish = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=bts, status='finished')))[0]['count']

            penugasanPersentaseAI = (
                penugasanAITotalFinish/penugasanAITotal) * 100
            penugasanPersentaseBTS = (
                penugasanBTSTotalFinish / penugasanBTSTotal) * 100

            penugasanAISekarangList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=ai, month=month, year=year)))
            penugasanBTSSekarangList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=bts, month=month, year=year)))

            penugasanAISebelumnyaList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=ai, month=pastMonth, year=pastYear)))
            penugasanBTSSebelumnyaList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=bts, month=pastMonth, year=pastYear)))

            penugasanAISekarang = 0 if len(
                penugasanAISekarangList) == 0 else penugasanAISekarangList[0]['count']
            penugasanBTSSekarang = 0 if len(
                penugasanBTSSekarangList) == 0 else penugasanBTSSekarangList[0]['count']

            penugasanAISebelumnya = 0 if len(
                penugasanAISebelumnyaList) == 0 else penugasanAISebelumnyaList[0]['count']
            penugasanBTSSebelumnya = 0 if len(
                penugasanBTSSebelumnyaList) == 0 else penugasanBTSSebelumnyaList[0]['count']

            persentasiKenaikanAI = ((
                abs(penugasanAISekarang - penugasanAISebelumnya)) / penugasanAISebelumnya) * 100 if penugasanAISebelumnya > 0 else 0
            persentasiKenaikanBTS = ((
                abs(penugasanBTSSekarang - penugasanBTSSebelumnya)) / penugasanBTSSebelumnya) * 100 if penugasanAISebelumnya > 0 else 0

            try:
                subject = f'Monthly Report SMASLAB {bulan[month-1]} {year}'
                text_content = ''
                htmly = get_template('email/executive/executive.html')
                d = {
                    'bulan': f'{bulan[month-1]} {year}',

                    'ai_finish': penugasanAITotalFinish,
                    'ai_total': penugasanAITotal,

                    'bts_finish': penugasanAITotalFinish,
                    'bts_total': penugasanAITotal,

                    'ai_persentase': round(penugasanPersentaseAI, 2),
                    'ai_penambahan': persentasiKenaikanAI,
                    'ai_bulan': penugasanAISekarang,

                    'bts_persentase': round(penugasanPersentaseBTS, 2),
                    'bts_penambahan': persentasiKenaikanBTS,
                    'bts_bulan': penugasanBTSSekarang,

                    'ai_perubahan': 'Kenaikan' if penugasanAISekarang > penugasanAISebelumnya else 'Penurunan' if penugasanAISekarang < penugasanAISebelumnya else 'Tidak ada Perubahan',
                    'ai_persentase_penambahan': f'{round(persentasiKenaikanAI, 2)}%' if penugasanAISekarang != penugasanAISebelumnya else '',

                    'bts_perubahan': 'Kenaikan' if penugasanBTSSekarang > penugasanBTSSebelumnya else 'Penurunan' if penugasanBTSSekarang < penugasanBTSSebelumnya else 'Tidak ada Perubahan',
                    'bts_persentase_penambahan': f'{round(persentasiKenaikanBTS, 2)}%' if penugasanBTSSekarang != penugasanBTSSebelumnya else '',

                    'media_url': settings.URL_MEDIA,
                    'survej_url': settings.URL_LOGIN,

                    'message_bottom': 'silahkan login akun SMASLAB untuk melihat detil laporan melalui aplikasi mobile ataupun website '+settings.URL_LOGIN

                }

                executive = [x.email for x in UserInfo.objects.filter(
                    role=ObjectId('5f27f7a6ea250a01d2b99d7a'))]

                html_content = htmly.render(d)
                sender = settings.EMAIL_ADMIN

                msg = EmailMultiAlternatives(
                    subject, text_content, sender, executive)
                msg.attach_alternative(html_content, "text/html")
                response = msg.send()
                # print('asdad', response)
            except:
                pass

        except Exception as e:
            print(e)
        #return HttpResponse(f' & ')
        return Response.ok(
            values=[],
            message='Berhasil'
        )
    #return HttpResponse(f'asd')
    return Response.badRequest(
        values=[],
        message='Gagal'
    )


def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])

    # haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    r = 6371  # Radius of earth in kilometers. Use 3956 for miles
    return c * r


def uploadlokasi(request):
    if request.method == 'POST':
        import openpyxl
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["List Lokasi Survey"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        lokasi_gagal = 'Lokasi gagal : '

        for row in worksheet.iter_rows():
            if str(row[1].value) == 'None':
                break
            if str(row[0].value) == 'NO':
                continue

            if str(row[1].value).upper() == 'PERMOHONAN AKSES INTERNET':
                jns = 'AI'
            else:
                jns = 'BTS'
            # try:
            data_provinsi = provinsi.objects.filter(
                name=str(row[2].value).upper()).first()
            # if len(data_provinsi) == 0:
            if data_provinsi is None:
                # continue
                data_provinsi = provinsi(
                    name=str(row[2].value).upper()
                )
                data_provinsi.save()
            if str(row[3].value)[0:3].upper() == 'KAB':
                kabupaten_ = kabupaten.objects.filter(
                    name=str(row[3].value).upper()).first()
                if kabupaten_ is None:
                    # continue
                    kabupaten_ = kabupaten(
                        name=str(row[3].value).upper(),
                        provinsi=ObjectId(data_provinsi.id)
                    )
                    kabupaten_.save()
            else:
                kota_ = kota.objects.filter(
                    name=str(row[3].value).upper()).first()
                if kota_ is None:
                    # continue
                    kota_ = kota(
                        name=str(row[3].value).upper(),
                        provinsi=ObjectId(data_provinsi.id)
                    )
                    kota_.save()
            data_kecamatan = kecamatan.objects.filter(
                name=str(row[4].value).upper()).first()
            if data_kecamatan is None:
                # continue
                try:
                    data_kecamatan = kecamatan(
                        name=str(row[4].value).upper(),
                        kabupaten=ObjectId(kabupaten_.id)
                    )
                    data_kecamatan.save()
                except:
                    data_kecamatan = kecamatan(
                        name=str(row[4].value).upper(),
                        kota=ObjectId(kota_.id)
                    )
                    data_kecamatan.save()
            data_desa = desa.objects.filter(
                name=str(row[5].value).upper()).first()

            if data_desa is None:
                # continue
                data_desa = desa(
                    name=str(row[5].value).upper(),
                    kecamatan=ObjectId(data_kecamatan.id)
                )
                data_desa.save()
            else:
                LokSurvey_ = LokasiSurvey.objects.filter(
                    desa=ObjectId(data_desa.id), jenis=jns)
                if len(LokSurvey_) > 0:
                    lokasi_gagal += str(row[5].value).upper()+', '
                    continue

            # if str(row[1].value).upper() == 'PERMOHONAN AKSES INTERNET':
            #    jns = 'AI'
            # else:
            #    jns = 'BTS'
            # try:
            # if str(row[3].value)[0:3].upper() == 'KAB':
            #    #LokasiSurvey_ = LokasiSurvey.objects.get(provinsi=ObjectId(data_provinsi["id"]), kabupaten=ObjectId(kabupaten_["id"]),
                #                                         kecamatan=ObjectId(data_kecamatan["id"]), desa=ObjectId(data_desa["id"]), jenis=jns)
            # else:
            #    LokasiSurvey_ = LokasiSurvey.objects.get(provinsi=ObjectId(data_provinsi["id"]), kota=ObjectId(kota_["id"]),
            #                                                kecamatan=ObjectId(data_kecamatan["id"]), desa=ObjectId(data_desa["id"]), jenis=jns)
                # return Response.badRequest(
                #    values='null',
                #    message='Data exists'
                # )
            # except LokasiSurvey.DoesNotExist:
            status = {'status': 'created', 'date': datetime.utcnow(
            ) + timedelta(hours=7)}
            LokasiSurvey_ = LokasiSurvey.objects.filter(latitude=str(
                row[6].value).upper(), longitude=str(row[7].value).upper())
            if len(LokasiSurvey_) == 0:
                LokasiSurvey_ = LokasiSurvey()
                #LokasiSurvey_.user = ObjectId("5f1aa3f07c33fe56ba294923")
                LokasiSurvey_.provinsi = ObjectId(data_provinsi["id"])
                try:
                    LokasiSurvey_.kabupaten = ObjectId(kabupaten_["id"])
                except:
                    LokasiSurvey_.kota = ObjectId(kota_["id"])
                LokasiSurvey_.kecamatan = ObjectId(data_kecamatan["id"])
                LokasiSurvey_.desa = ObjectId(data_desa["id"])
                LokasiSurvey_.jenis = jns
                LokasiSurvey_.latitude = str(row[6].value).upper()
                LokasiSurvey_.longitude = str(row[7].value).upper()
                LokasiSurvey_.status.append(status)

                LokasiSurvey_.save()
            else:
                lokasi_gagal += str(row[5].value).upper()+', '
                # return Response.badRequest(
                #    values=[str(row[5].value).upper()],
                #    message=lokasi_gagal
                # )

        # return Response.ok(
        #    values=[],
        #    message='Upload berhasil'
        # )
        if lokasi_gagal == 'Lokasi gagal : ':
            return Response.ok(
                values=[],
                message=''
            )
        else:
            return Response.ok(
                values=[],
                message=lokasi_gagal
            )
        #    return JsonResponse({"state": "success"})
        # return JsonResponse({"state": "success","gagal": lokasi_gagal})


def uploadlokasikodesurvey(request):
    if request.method == 'POST':
        import openpyxl
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["List Lokasi Survey"]

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row

        for row in worksheet.iter_rows():
            if str(row[1].value) == 'None':
                break
            if str(row[0].value) == 'NO':
                continue
            # try:
            data_provinsi = provinsi.objects.filter(
                name=str(row[2].value).upper()).first()
            # if len(data_provinsi) == 0:
            if data_provinsi is None:
                # continue
                data_provinsi = provinsi(
                    name=str(row[2].value).upper()
                )
                data_provinsi.save()

            if str(row[3].value)[0:3].upper() == 'KAB':
                kabupaten_ = kabupaten.objects.filter(
                    name=str(row[3].value).upper()).first()
                # if len(kabupaten_) == 0:
                if kabupaten_ is None:
                    # continue
                    kabupaten_ = kabupaten(
                        name=str(row[3].value).upper(),
                        provinsi=ObjectId(data_provinsi.id)
                    )
                    kabupaten_.save()
            else:
                kota_ = kota.objects.filter(
                    name=str(row[3].value).upper()).first()
                # if len(kota_) == 0:
                if kota_ is None:
                    # continue
                    kota_ = kota(
                        name=str(row[3].value).upper(),
                        provinsi=ObjectId(data_provinsi.id)
                    )
                    kota_.save()
            data_kecamatan = kecamatan.objects.filter(
                name=str(row[4].value).upper()).first()
            # if len(data_kecamatan) == 0:
            if data_kecamatan is None:
                # continue
                try:
                    data_kecamatan = kecamatan(
                        name=str(row[4].value).upper(),
                        kabupaten=ObjectId(kabupaten_.id)
                    )
                    data_kecamatan.save()
                except:
                    data_kecamatan = kecamatan(
                        name=str(row[4].value).upper(),
                        kota=ObjectId(kota_.id)
                    )
                    data_kecamatan.save()
            data_desa = desa.objects.filter(
                name=str(row[5].value).upper()).first()
            # if len(data_desa) == 0:
            if data_desa is None:
                # continue
                data_desa = desa(
                    name=str(row[5].value).upper(),
                    kecamatan=ObjectId(data_kecamatan.id)
                )
                data_desa.save()
            if str(row[1].value).upper() == 'PERMOHONAN AKSES INTERNET':
                jns = 'AI'
            else:
                jns = 'BTS'

            try:
                #data_jenis = JenisSurvey.objects.get(jenis=kode_jns)
                data_jenis = JenisSurvey.objects.get(jenis=jns)
            except JenisSurvey.DoesNotExist:
                return Response.badRequest(
                    values='null',
                    message='jenissurvey not found'
                )

            status = {'status': 'created', 'date': datetime.utcnow(
            ) + timedelta(hours=7)}
            status_assigned = {'status': 'assigned', 'date': datetime.utcnow(
            ) + timedelta(hours=7)}
            LokasiSurvey_ = LokasiSurvey.objects.filter(latitude=str(
                row[6].value).upper(), longitude=str(row[7].value).upper(), jenis=jns)
            if len(LokasiSurvey_) == 0:
                LokasiSurvey_ = LokasiSurvey()
                LokasiSurvey_.provinsi = ObjectId(data_provinsi["id"])
                try:
                    LokasiSurvey_.kabupaten = ObjectId(kabupaten_["id"])
                except:
                    LokasiSurvey_.kota = ObjectId(kota_["id"])
                LokasiSurvey_.kecamatan = ObjectId(data_kecamatan["id"])
                LokasiSurvey_.desa = ObjectId(data_desa["id"])
                LokasiSurvey_.jenis = jns
                LokasiSurvey_.latitude = str(row[6].value).upper()
                LokasiSurvey_.longitude = str(row[7].value).upper()
                LokasiSurvey_.status.append(status)
                LokasiSurvey_.status.append(status_assigned)

                LokasiSurvey_.save()

                # try:
                file = request.FILES['doc']
                if not file:
                    return Response.badRequest(message='No File Upload')
                fs = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/survey/spk/',
                    base_url=f'{settings.MEDIA_URL}/survey/spk/'
                )

                status = {'status': 'created', 'date': datetime.utcnow(
                ) + timedelta(hours=7)}
                # Duplicate filtering

                # for i in range(2):
                # if i == 0:
                #    kode_survey1 = 'AI-'+str(row[8].value).upper()
                #    kode_jns = 'AI'
                # else:
                #    kode_survey1 = 'BTS-'+str(row[8].value).upper()
                #    kode_jns = 'BTS'
                kode_survey1 = jns+'-'+str(row[8].value).upper()
                try:
                    kode_survey = Penugasan.objects(
                        kode__startswith=kode_survey1.upper()).order_by('-kode').first()
                    lst_kode = kode_survey['kode'].split('-')
                    if len(lst_kode) > 0:
                        kode_Penugasan = str(int(lst_kode[2])+1).zfill(5)
                    else:
                        kode_Penugasan = str(
                            int(kode_survey['kode'])+1).zfill(5)
                except:
                    kode_Penugasan = '1'.zfill(5)

                # try:
                #    #data_jenis = JenisSurvey.objects.get(jenis=kode_jns)
                #    data_jenis = JenisSurvey.objects.get(jenis=jns)
                # except JenisSurvey.DoesNotExist:
                #    return Response.badRequest(
                #        values='null',
                #        message='jenissurvey not found'
                #    )

                try:
                    try:
                        pt_survey = Surveyor.objects.get(
                            name=str(row[9].value).upper())
                        pt_surveyid = pt_survey.id
                    except:
                        pt_surveyid = "5f1fb680171eb8928f9e7a4b"
                except Surveyor.DoesNotExist:
                    return Response.badRequest(
                        values='null',
                        message='Surveyor not found'
                    )
                print(data_provinsi.name, kabupaten_.name, data_kecamatan.name, data_desa.name, str(
                    row[6].value).upper(), str(row[7].value).upper(), kode_survey1+'-'+kode_Penugasan)
                # print(LokasiSurvey_.id)

                try:
                    filename = fs.save(file.name, file)
                    file_path = fs.url(filename)
                except:
                    pass
                doc = DocumentPenugasan(
                    name=file.name,
                    path=file_path,
                    create_date=datetime.utcnow() + timedelta(hours=7),
                    update_date=datetime.utcnow() + timedelta(hours=7)
                )
                doc.save()

                Penugasan_ = Penugasan(
                    user=ObjectId(request.POST.get('user')),
                    kode=kode_survey1+'-'+kode_Penugasan,
                    jenissurvey=ObjectId(data_jenis.id),
                    surveyor=ObjectId(pt_surveyid),
                    lokasisurvey=ObjectId(LokasiSurvey_.id),
                    tanggal_penugasan=datetime.strptime(request.POST.get(
                        'tanggal_penugasan'), '%Y-%m-%d 00:00:00'),
                    target=datetime.strptime(
                        request.POST.get('sla'), '%Y-%m-%d 00:00:00'),
                    nospk=kode_Penugasan,
                    spk=ObjectId(doc.id)
                )

                Penugasan_.status.append(status)
                Penugasan_.status.append(status_assigned)
                Penugasan_.save()
                #filename = fs.save(file.name, file)
                #file_path = fs.url(filename)
                # doc = DocumentPenugasan(
                #    name=file.name,
                #    path=file_path,
                # )
                # doc.save()

                # add status assigned di tabel lokasisurvey
                # try:
                #    lok_survey = LokasiSurvey.objects.get(id=request.POST.get('lokasisurvey'))
                #    status = {'status': 'assigned', 'date': datetime.utcnow(
                #        ) + timedelta(hours=7)}
                # except LokasiSurvey.DoesNotExist:
                #    lok_survey = None

                # if not lok_survey:
                #    return Response.badRequest(
                #        values='null',
                #        message='Lokasi Survey not found'
                #    )
                # status = {'status': 'assigned', 'date': datetime.utcnow(
                #        ) + timedelta(hours=7)}
                #statuspenugasan = [i for i,x in enumerate(lok_survey.status) if x['status']=='assigned']
                # if not statuspenugasan:
                #    lok_survey.status.append(status)
                #    lok_survey.save()
                # =============
                #Penugasan_.spk = ObjectId(doc.id)
                Penugasan_.save()

                #result = Penugasan.objects.get(id=ObjectId(Penugasan_.id)).serialize()
                # return Response.ok(
                #    values=result,
                #    message='Penugasan Created'
                # )
                # except Exception as e:
                #    return Response.badRequest(message=str(e))

        return JsonResponse({"state": "success"})


def addbatch(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    if request.method == "POST":  # Add
        try:
            #file = request.FILES['doc']
            #if not file:
            #    return Response.badRequest(message='No File Upload')

            body_data = request.POST.dict()

            judul = body_data.get('judul')
            tanggal_mulai_undangan = body_data.get('tanggal_mulai_undangan')
            tanggal_selesai_undangan = body_data.get('tanggal_selesai_undangan')
            tanggal_mulai_kerja = body_data.get('tanggal_mulai_kerja')
            tanggal_selesai_kerja = body_data.get('tanggal_selesai_kerja')
            rfi = body_data.get('rfi')
            type = body_data.get('type')
            creator = body_data.get('creator')
            penyedia_undang = body_data.get('penyedia_undang')

            status_ = {'status': 'Dibuka', 'tanggal_pembuatan': datetime.utcnow(
                    ) + timedelta(hours=7)}

            data_batch = batch(
                judul = judul,
                type = type,
                sites = [],
                creator = creator,
                rfi_no = rfi,
                tanggal_mulai_undangan = tanggal_mulai_undangan,
                tanggal_selesai_undangan = tanggal_selesai_undangan,
                tanggal_mulai_kerja = tanggal_mulai_kerja,
                tanggal_selesai_kerja = tanggal_selesai_kerja,
                penyedia_undang = penyedia_undang.split(","),
                created_at = DateTimeField(
                    default=datetime.utcnow() + timedelta(hours=7)),
                updated_at = DateTimeField(
                    default=datetime.utcnow() + timedelta(hours=7))
            )
            data_batch.status.append(status_)
            data_batch.save()
        except Exception as e:
            return Response.badRequest(message=str(e))

    else:
        return Response.badRequest(message='Hanya POST')


def getlokasisurvey(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    strjson = json.loads(request.body)
    field = strjson["field"].lower()
    value_ = strjson["value"].upper()
    jenis = strjson.get("jenis", None)
    pp_jenis = []
    if jenis:
        # try:
        #data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
        pp_jenis = [{
            '$match': {
                'jenis': jenis.upper()
            }
        }]
        # except JenisSurvey.DoesNotExist:
        #    return Response.badRequest(
        #        values=[],
        #        message='Jenis Survey not found'
        #    )

    page = int(strjson.get('page', 0)) - 1
    skip = []
    if page >= 0:
        skip = [{'$skip': 20 * page},
                {'$limit': 20}]

    pipeline = pp_jenis + [
        {
            '$lookup': {
                'from': 'provinsi',
                'localField': 'provinsi',
                'foreignField': '_id',
                'as': 'provinsi'
            }
        }, {
            '$unwind': {
                'path': '$provinsi',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'kabupaten',
                'localField': 'kabupaten',
                'foreignField': '_id',
                'as': 'kabupaten'
            }
        }, {
            '$unwind': {
                'path': '$kabupaten',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'kota',
                'localField': 'kota',
                'foreignField': '_id',
                'as': 'kota'
            }
        }, {
            '$unwind': {
                'path': '$kota',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'kecamatan',
                'localField': 'kecamatan',
                'foreignField': '_id',
                'as': 'kecamatan'
            }
        }, {
            '$unwind': {
                'path': '$kecamatan',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'desa',
                'localField': 'desa',
                'foreignField': '_id',
                'as': 'desa'
            }
        }, {
            '$unwind': {
                'path': '$desa',
                'preserveNullAndEmptyArrays': True
            }
        }
    ]

    if field == 'all':
        pipe = pipeline + skip
        agg_cursor = LokasiSurvey.objects.aggregate(*pipe)

        LokasiSurvey_ = list(agg_cursor)
    else:
        if field == 'provinsi':
            try:
                data_ = provinsi.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except provinsi.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'provinsi': ObjectId(data_.id)
                }
            }]
        if field == 'kabupaten':
            try:
                data_ = kabupaten.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except kabupaten.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'kabupaten': ObjectId(data_.id)
                }
            }]
        if field == 'kota':
            try:
                data_ = kota.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except kota.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'kota': ObjectId(data_.id)
                }
            }]
        if field == 'kecamatan':
            try:
                data_ = kecamatan.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except kecamatan.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'kecamatan': ObjectId(data_.id)
                }
            }]
        if field == 'desa':
            try:
                data_ = desa.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except desa.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'desa': ObjectId(data_.id)
                }
            }]
        # try:
        if field == 'status':
            match = [{
                '$addFields': {
                    'last': {
                        '$arrayElemAt': [
                            '$status', -1
                        ]
                    }
                }
            }, {
                '$match': {
                    'last.status': value_.lower()
                }
            }, {
                '$project': {
                    'last': 0
                }
            }]
            # if value_.lower()=='created':
            #    LokasiSurvey_ = LokasiSurvey.objects.filter(status__0__status=value_.lower(),status__1__exists=False)
            # else:
            #    LokasiSurvey_ = LokasiSurvey.objects.filter(status__status=value_.lower())
            # else:
            #    LokasiSurvey_ = LokasiSurvey.objects.filter(**{field: value_})
        if field == 'jenis':
            match = [{
                '$match': {
                    'jenis': value_.upper()
                }
            }]
        # except LokasiSurvey.DoesNotExist:
        #    return Response.badRequest(
        #        values=[],
        #        message='Lokasi Survey not found'
        #    )
        pipe = match + pipeline + skip
        agg_cursor = LokasiSurvey.objects.aggregate(*pipe)

        LokasiSurvey_ = list(agg_cursor)

    if len(LokasiSurvey_) > 0:
        return Response.ok(
            values=json.loads(json.dumps(LokasiSurvey_, default=str)),
            message=f'{len(LokasiSurvey_)} Data'
        )
    else:
        return Response.badRequest(
            values=[],
            message='Lokasi Survey not found'
        )