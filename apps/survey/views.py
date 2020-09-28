
from django.shortcuts import render
from django.http import JsonResponse
from survey.models import *
from userinfo.models import Surveyor, JenisSurvey, UserInfo, UserToken, Message
from userinfo.views import authenticate_credentials
from django.dispatch import receiver
from django.db.models.signals import post_save
from django.conf import settings
import json
from django.core.exceptions import ObjectDoesNotExist
import pandas
from .response import Response
from bson import ObjectId, json_util
from django.http import HttpResponse
#import datetime
from datetime import datetime, timedelta, timezone
from django.core.serializers import serialize
from django.core.files.storage import FileSystemStorage
import requests

from django.core.mail import EmailMultiAlternatives
from django.template.loader import get_template
from django.core import serializers
from survey.serializer import *

from itertools import groupby
from userinfo.utils.notification import Notification
import calendar


def getLaporan(request):

    if request.method == "POST":

        bulan = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
                 "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)
            month = data['month']
            year = data['year']

            pastMonth = data['month'] - 1
            pastYear = year

            if pastMonth == 0:
                pastMonth = 12
                pastYear = year-1

            ai = '5f16b4ba149882a98fc6655e'
            bts = '5f1521524f9c6764c713d73c'

            if month is None or year is None:
                return Response.badRequest(message='Need Json Body "month" & "year"')

            def pipelineDate(month, year, jenis):

                lastDay = calendar.monthrange(year=year, month=month)[1]
                return [
                    {
                        '$addFields': {
                            'last': {
                                '$arrayElemAt': [
                                    '$status', -1
                                ]
                            }
                        }
                    }, {
                        '$match': {
                            'jenissurvey': ObjectId(jenis),
                            'last.status': 'finished',
                            'last.date': {
                                '$gte': datetime(year, month, 1, 00, 00, 00, tzinfo=timezone.utc),
                                '$lte': datetime(year, month, lastDay, 23, 59, 59, tzinfo=timezone.utc)
                            }
                        }
                    }, {
                        '$count': 'count'
                    }
                ]

            def pipelineTotal(jenis, status):
                return [
                    {
                        '$match': {
                            'jenissurvey': ObjectId(jenis),
                            'status.status': status,
                        }
                    }, {
                        '$count': 'count'
                    }
                ]

            penugasanAITotal = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=ai, status='assigned')))[0]['count']
            penugasanBTSTotal = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=bts, status='assigned')))[0]['count']

            penugasanAITotalFinish = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=ai, status='finished')))[0]['count']
            penugasanBTSTotalFinish = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=bts, status='finished')))[0]['count']

            penugasanPersentaseAI = (
                penugasanAITotalFinish/penugasanAITotal) * 100
            penugasanPersentaseBTS = (
                penugasanBTSTotalFinish / penugasanBTSTotal) * 100

            penugasanAISekarangList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=ai, month=month, year=year)))
            penugasanBTSSekarangList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=bts, month=month, year=year)))

            penugasanAISebelumnyaList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=ai, month=pastMonth, year=pastYear)))
            penugasanBTSSebelumnyaList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=bts, month=pastMonth, year=pastYear)))

            penugasanAISekarang = 0 if len(
                penugasanAISekarangList) == 0 else penugasanAISekarangList[0]['count']
            penugasanBTSSekarang = 0 if len(
                penugasanBTSSekarangList) == 0 else penugasanBTSSekarangList[0]['count']

            penugasanAISebelumnya = 0 if len(
                penugasanAISebelumnyaList) == 0 else penugasanAISebelumnyaList[0]['count']
            penugasanBTSSebelumnya = 0 if len(
                penugasanBTSSebelumnyaList) == 0 else penugasanBTSSebelumnyaList[0]['count']

            persentasiKenaikanAI = ((
                abs(penugasanAISekarang - penugasanAISebelumnya)) / penugasanAISebelumnya) * 100 if penugasanAISebelumnya > 0 else 0
            persentasiKenaikanBTS = ((
                abs(penugasanBTSSekarang - penugasanBTSSebelumnya)) / penugasanBTSSebelumnya) * 100 if penugasanAISebelumnya > 0 else 0

            try:
                subject = f'Monthly Report SMASLAB {bulan[month-1]} {year}'
                text_content = ''
                htmly = get_template('email/executive/executive.html')
                d = {
                    'bulan': f'{bulan[month-1]} {year}',

                    'ai_finish': penugasanAITotalFinish,
                    'ai_total': penugasanAITotal,

                    'bts_finish': penugasanAITotalFinish,
                    'bts_total': penugasanAITotal,

                    'ai_persentase': round(penugasanPersentaseAI, 2),
                    'ai_penambahan': persentasiKenaikanAI,
                    'ai_bulan': penugasanAISekarang,

                    'bts_persentase': round(penugasanPersentaseBTS, 2),
                    'bts_penambahan': persentasiKenaikanBTS,
                    'bts_bulan': penugasanBTSSekarang,

                    'ai_perubahan': 'Kenaikan' if penugasanAISekarang > penugasanAISebelumnya else 'Penurunan' if penugasanAISekarang < penugasanAISebelumnya else 'Tidak ada Perubahan',
                    'ai_persentase_penambahan': f'{round(persentasiKenaikanAI, 2)}%' if penugasanAISekarang != penugasanAISebelumnya else '',

                    'bts_perubahan': 'Kenaikan' if penugasanBTSSekarang > penugasanBTSSebelumnya else 'Penurunan' if penugasanBTSSekarang < penugasanBTSSebelumnya else 'Tidak ada Perubahan',
                    'bts_persentase_penambahan': f'{round(persentasiKenaikanBTS, 2)}%' if penugasanBTSSekarang != penugasanBTSSebelumnya else '',

                    'media_url': settings.URL_MEDIA,
                    'survej_url': settings.URL_LOGIN,

                    'message_bottom': 'silahkan login akun SMASLAB untuk melihat detil laporan melalui aplikasi mobile ataupun website '+settings.URL_LOGIN

                }

                executive = [x.email for x in UserInfo.objects.filter(
                    role=ObjectId('5f27f7a6ea250a01d2b99d7a'))]

                html_content = htmly.render(d)
                sender = settings.EMAIL_ADMIN

                msg = EmailMultiAlternatives(
                    subject, text_content, sender, executive)
                msg.attach_alternative(html_content, "text/html")
                response = msg.send()
                # print('asdad', response)
            except:
                pass

        except Exception as e:
            print(e)
        #return HttpResponse(f' & ')
        return Response.ok(
            values=[],
            message='Berhasil'
        )
    #return HttpResponse(f'asd')
    return Response.badRequest(
        values=[],
        message='Gagal'
    )


def addjenissurvey(request):
    """
    Creator:
        1. cito
    Authorization:
        token       {string}
    Params:
        {JSON}
        user        {string}
        jenis       {string}
    Raises:
        Json     {Fail}

    Returns:
        Json     {Success}
    """
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    strjson = json.loads(request.body)
    if request.method == "POST":  # Add
        try:
            # Duplicate filtering
            jenis = strjson["jenis"].upper()
            jenissurvey = JenisSurvey.objects.get(jenis=jenis)
            if jenissurvey:
                return JsonResponse({"state": "fail", "action": "exists data"})
            # jenissurvey_ = JenisSurvey.create(user, jenis)
            # jenissurvey_.save()
            # return JsonResponse({"state": "success"})
        # except Exception as e:
        except JenisSurvey.DoesNotExist:
            # print(e)
            # print(strjson["jenis"])
            # return JsonResponse({"state": "fail"})
            jenissurvey_ = JenisSurvey()
            #jenissurvey_.user = ObjectId(strjson["user"])
            jenissurvey_.jenis = strjson["jenis"].upper()

            jenissurvey_.save()
            return JsonResponse({"state": "success"})
        return JsonResponse({"state": "fail", "action": "add success"})
    return JsonResponse({"state": "fail", "action": "none"})


def getjenissurvey(request):
    """
    Creator:
        1. cito
    Authorization:
        token       {string}
    Params:
        {JSON}
        jenis       {string}
    Raises:
        Json     {Fail}

    Returns:
        Json     {Success}
    """
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    strjson = json.loads(request.body)
    jenis = strjson["jenis"].upper()
    if jenis == 'ALL':
        jenissurvey = JenisSurvey.objects.all()
    else:
        jenissurvey = JenisSurvey.objects.filter(jenis=jenis)
    json_list = []
    for search in jenissurvey:
        json_dict = {}
        json_dict["jenis"] = search.jenis
        json_list.append(json_dict)
    return JsonResponse({"state": "success", "action": "find", "data": json_list})


def addsurveyor(request):
    if request.method == 'POST':
        try:
            req = request.body.decode('utf-8')
            data = json.loads(req)
            try:
                data_jenis = JenisSurvey.objects.get(
                    jenis=data['jenis'].upper())
            except JenisSurvey.DoesNotExist:
                return Response.badRequest(
                    values='null',
                    message='jenissurvey not found'
                )
            surveyor_ = Surveyor(
                # user=ObjectId(data['user']),
                name=data['surveyor'].upper(),
                jenissurvey=ObjectId(data_jenis.id),
            )
            surveyor_.save()
            result = surveyor_.serialize()
            #result = Penugasan.objects.get(id=ObjectId(Penugasan_.id)).serialize()
            return Response.ok(
                values=result,
                message='Surveyor Created Successfully'
            )
        except Exception as e:
            return Response.badRequest(message=str(e))
    else:
        return Response.badRequest(message='Only Post Request Accepted')


def getsurveyor(request):
    """
    Creator:
        1. cito
    Authorization:
        token       {string}
    Params:
        {JSON}
        surveyor       {string}
    Raises:
        Json     {Fail}

    Returns:
        Json     {Success}
    """
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})

    try:
        req = request.body.decode("utf-8")
        data = json.loads(req)
        surveyor = data["surveyor"].upper()
        jenissurvey = data["jenis"].upper()
        page = int(data.get('page', 0)) - 1
        skip = []
        if page >= 0:
            skip = [{'$skip': 20 * page},
                    {'$limit': 20}]

        try:
            data_jenis = JenisSurvey.objects.get(jenis=jenissurvey)
        except JenisSurvey.DoesNotExist:
            return Response.badRequest(
                values='null',
                message='jenissurvey not found'
            )

        pipeline = [
            {
                '$lookup': {
                    'from': 'jenis_survey',
                    'localField': 'jenissurvey',
                    'foreignField': '_id',
                    'as': 'jenissurvey'
                }
            }, {
                '$unwind': {
                    'path': '$jenissurvey',
                    'preserveNullAndEmptyArrays': True
                }
            }
        ]
        if surveyor.lower() == 'all':
            # surveyor_ = Surveyor.objects.filter(
            #    jenissurvey=ObjectId(data_jenis.id))
            pipe = pipeline + skip
            agg_cursor = Surveyor.objects.aggregate(*pipe)
            surveyor_ = list(agg_cursor)
        else:
            # surveyor_ = Surveyor.objects.filter(
            #    name=surveyor, jenissurvey=ObjectId(data_jenis.id))
            match = [
                {
                    '$match': {
                        'jenissurvey': ObjectId(data_jenis.id)
                    }
                }, {
                    '$match': {
                        'name': surveyor
                    }
                }
            ]
            pipe = match + pipeline + skip
            agg_cursor = Surveyor.objects.aggregate(*pipe)
            surveyor_ = list(agg_cursor)
        if not surveyor_:
            return Response.badRequest(
                values='null',
                message='Surveyor not found'
            )

        return Response.ok(
            values=json.loads(json.dumps(surveyor_, default=str)),
            message=f'{len(surveyor_)} Data'
        )
    except Exception as e:
        print(e)
        return HttpResponse(e)


def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points 
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])

    # haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    r = 6371  # Radius of earth in kilometers. Use 3956 for miles
    return c * r


def uploadlokasi(request):
    if request.method == 'POST':
        import openpyxl
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["List Lokasi Survey"]
        print(worksheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        lokasi_gagal = 'Lokasi gagal : '

        for row in worksheet.iter_rows():
            if str(row[1].value) == 'None':
                break
            if str(row[0].value) == 'NO':
                continue

            if str(row[1].value).upper() == 'PERMOHONAN AKSES INTERNET':
                jns = 'AI'
            else:
                jns = 'BTS'
            # try:
            data_provinsi = provinsi.objects.filter(
                name=str(row[2].value).upper()).first()
            # if len(data_provinsi) == 0:
            if data_provinsi is None:
                # continue
                data_provinsi = provinsi(
                    name=str(row[2].value).upper()
                )
                data_provinsi.save()
            if str(row[3].value)[0:3].upper() == 'KAB':
                kabupaten_ = kabupaten.objects.filter(
                    name=str(row[3].value).upper()).first()
                if kabupaten_ is None:
                    # continue
                    kabupaten_ = kabupaten(
                        name=str(row[3].value).upper(),
                        provinsi=ObjectId(data_provinsi.id)
                    )
                    kabupaten_.save()
            else:
                kota_ = kota.objects.filter(
                    name=str(row[3].value).upper()).first()
                if kota_ is None:
                    # continue
                    kota_ = kota(
                        name=str(row[3].value).upper(),
                        provinsi=ObjectId(data_provinsi.id)
                    )
                    kota_.save()
            data_kecamatan = kecamatan.objects.filter(
                name=str(row[4].value).upper()).first()
            if data_kecamatan is None:
                # continue
                try:
                    data_kecamatan = kecamatan(
                        name=str(row[4].value).upper(),
                        kabupaten=ObjectId(kabupaten_.id)
                    )
                    data_kecamatan.save()
                except:
                    data_kecamatan = kecamatan(
                        name=str(row[4].value).upper(),
                        kota=ObjectId(kota_.id)
                    )
                    data_kecamatan.save()
            data_desa = desa.objects.filter(
                name=str(row[5].value).upper()).first()

            if data_desa is None:
                # continue
                data_desa = desa(
                    name=str(row[5].value).upper(),
                    kecamatan=ObjectId(data_kecamatan.id)
                )
                data_desa.save()
            else:
                LokSurvey_ = LokasiSurvey.objects.filter(
                    desa=ObjectId(data_desa.id), jenis=jns)
                if len(LokSurvey_) > 0:
                    lokasi_gagal += str(row[5].value).upper()+', '
                    continue

            # if str(row[1].value).upper() == 'PERMOHONAN AKSES INTERNET':
            #    jns = 'AI'
            # else:
            #    jns = 'BTS'
            # try:
            # if str(row[3].value)[0:3].upper() == 'KAB':
            #    #LokasiSurvey_ = LokasiSurvey.objects.get(provinsi=ObjectId(data_provinsi["id"]), kabupaten=ObjectId(kabupaten_["id"]),
                #                                         kecamatan=ObjectId(data_kecamatan["id"]), desa=ObjectId(data_desa["id"]), jenis=jns)
            # else:
            #    LokasiSurvey_ = LokasiSurvey.objects.get(provinsi=ObjectId(data_provinsi["id"]), kota=ObjectId(kota_["id"]),
            #                                                kecamatan=ObjectId(data_kecamatan["id"]), desa=ObjectId(data_desa["id"]), jenis=jns)
                # return Response.badRequest(
                #    values='null',
                #    message='Data exists'
                # )
            # except LokasiSurvey.DoesNotExist:
            status = {'status': 'created', 'date': datetime.utcnow(
            ) + timedelta(hours=7)}
            LokasiSurvey_ = LokasiSurvey.objects.filter(latitude=str(
                row[6].value).upper(), longitude=str(row[7].value).upper())
            if len(LokasiSurvey_) == 0:
                LokasiSurvey_ = LokasiSurvey()
                #LokasiSurvey_.user = ObjectId("5f1aa3f07c33fe56ba294923")
                LokasiSurvey_.provinsi = ObjectId(data_provinsi["id"])
                try:
                    LokasiSurvey_.kabupaten = ObjectId(kabupaten_["id"])
                except:
                    LokasiSurvey_.kota = ObjectId(kota_["id"])
                LokasiSurvey_.kecamatan = ObjectId(data_kecamatan["id"])
                LokasiSurvey_.desa = ObjectId(data_desa["id"])
                LokasiSurvey_.jenis = jns
                LokasiSurvey_.latitude = str(row[6].value).upper()
                LokasiSurvey_.longitude = str(row[7].value).upper()
                LokasiSurvey_.status.append(status)

                LokasiSurvey_.save()
            else:
                lokasi_gagal += str(row[5].value).upper()+', '
                # return Response.badRequest(
                #    values=[str(row[5].value).upper()],
                #    message=lokasi_gagal
                # )

        # return Response.ok(
        #    values=[],
        #    message='Upload berhasil'
        # )
        if lokasi_gagal == 'Lokasi gagal : ':
            return Response.ok(
                values=[],
                message=''
            )
        else:
            return Response.ok(
                values=[],
                message=lokasi_gagal
            )
        #    return JsonResponse({"state": "success"})
        # return JsonResponse({"state": "success","gagal": lokasi_gagal})


def uploadlokasikodesurvey(request):
    if request.method == 'POST':
        import openpyxl
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["List Lokasi Survey"]

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row

        for row in worksheet.iter_rows():
            if str(row[1].value) == 'None':
                break
            if str(row[0].value) == 'NO':
                continue
            # try:
            data_provinsi = provinsi.objects.filter(
                name=str(row[2].value).upper()).first()
            # if len(data_provinsi) == 0:
            if data_provinsi is None:
                # continue
                data_provinsi = provinsi(
                    name=str(row[2].value).upper()
                )
                data_provinsi.save()

            if str(row[3].value)[0:3].upper() == 'KAB':
                kabupaten_ = kabupaten.objects.filter(
                    name=str(row[3].value).upper()).first()
                # if len(kabupaten_) == 0:
                if kabupaten_ is None:
                    # continue
                    kabupaten_ = kabupaten(
                        name=str(row[3].value).upper(),
                        provinsi=ObjectId(data_provinsi.id)
                    )
                    kabupaten_.save()
            else:
                kota_ = kota.objects.filter(
                    name=str(row[3].value).upper()).first()
                # if len(kota_) == 0:
                if kota_ is None:
                    # continue
                    kota_ = kota(
                        name=str(row[3].value).upper(),
                        provinsi=ObjectId(data_provinsi.id)
                    )
                    kota_.save()
            data_kecamatan = kecamatan.objects.filter(
                name=str(row[4].value).upper()).first()
            # if len(data_kecamatan) == 0:
            if data_kecamatan is None:
                # continue
                try:
                    data_kecamatan = kecamatan(
                        name=str(row[4].value).upper(),
                        kabupaten=ObjectId(kabupaten_.id)
                    )
                    data_kecamatan.save()
                except:
                    data_kecamatan = kecamatan(
                        name=str(row[4].value).upper(),
                        kota=ObjectId(kota_.id)
                    )
                    data_kecamatan.save()
            data_desa = desa.objects.filter(
                name=str(row[5].value).upper()).first()
            # if len(data_desa) == 0:
            if data_desa is None:
                # continue
                data_desa = desa(
                    name=str(row[5].value).upper(),
                    kecamatan=ObjectId(data_kecamatan.id)
                )
                data_desa.save()
            if str(row[1].value).upper() == 'PERMOHONAN AKSES INTERNET':
                jns = 'AI'
            else:
                jns = 'BTS'

            try:
                #data_jenis = JenisSurvey.objects.get(jenis=kode_jns)
                data_jenis = JenisSurvey.objects.get(jenis=jns)
            except JenisSurvey.DoesNotExist:
                return Response.badRequest(
                    values='null',
                    message='jenissurvey not found'
                )

            status = {'status': 'created', 'date': datetime.utcnow(
            ) + timedelta(hours=7)}
            status_assigned = {'status': 'assigned', 'date': datetime.utcnow(
            ) + timedelta(hours=7)}
            LokasiSurvey_ = LokasiSurvey.objects.filter(latitude=str(
                row[6].value).upper(), longitude=str(row[7].value).upper(), jenis=jns)
            if len(LokasiSurvey_) == 0:
                LokasiSurvey_ = LokasiSurvey()
                LokasiSurvey_.provinsi = ObjectId(data_provinsi["id"])
                try:
                    LokasiSurvey_.kabupaten = ObjectId(kabupaten_["id"])
                except:
                    LokasiSurvey_.kota = ObjectId(kota_["id"])
                LokasiSurvey_.kecamatan = ObjectId(data_kecamatan["id"])
                LokasiSurvey_.desa = ObjectId(data_desa["id"])
                LokasiSurvey_.jenis = jns
                LokasiSurvey_.latitude = str(row[6].value).upper()
                LokasiSurvey_.longitude = str(row[7].value).upper()
                LokasiSurvey_.status.append(status)
                LokasiSurvey_.status.append(status_assigned)

                LokasiSurvey_.save()

                # try:
                file = request.FILES['doc']
                if not file:
                    return Response.badRequest(message='No File Upload')
                fs = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/survey/spk/',
                    base_url=f'{settings.MEDIA_URL}/survey/spk/'
                )

                status = {'status': 'created', 'date': datetime.utcnow(
                ) + timedelta(hours=7)}
                # Duplicate filtering

                # for i in range(2):
                # if i == 0:
                #    kode_survey1 = 'AI-'+str(row[8].value).upper()
                #    kode_jns = 'AI'
                # else:
                #    kode_survey1 = 'BTS-'+str(row[8].value).upper()
                #    kode_jns = 'BTS'
                kode_survey1 = jns+'-'+str(row[8].value).upper()
                try:
                    kode_survey = Penugasan.objects(
                        kode__startswith=kode_survey1.upper()).order_by('-kode').first()
                    lst_kode = kode_survey['kode'].split('-')
                    if len(lst_kode) > 0:
                        kode_Penugasan = str(int(lst_kode[2])+1).zfill(5)
                    else:
                        kode_Penugasan = str(
                            int(kode_survey['kode'])+1).zfill(5)
                except:
                    kode_Penugasan = '1'.zfill(5)

                # try:
                #    #data_jenis = JenisSurvey.objects.get(jenis=kode_jns)
                #    data_jenis = JenisSurvey.objects.get(jenis=jns)
                # except JenisSurvey.DoesNotExist:
                #    return Response.badRequest(
                #        values='null',
                #        message='jenissurvey not found'
                #    )

                try:
                    try:
                        pt_survey = Surveyor.objects.get(
                            name=str(row[9].value).upper())
                        pt_surveyid = pt_survey.id
                    except:
                        pt_surveyid = "5f1fb680171eb8928f9e7a4b"
                except Surveyor.DoesNotExist:
                    return Response.badRequest(
                        values='null',
                        message='Surveyor not found'
                    )
                print(data_provinsi.name, kabupaten_.name, data_kecamatan.name, data_desa.name, str(
                    row[6].value).upper(), str(row[7].value).upper(), kode_survey1+'-'+kode_Penugasan)
                # print(LokasiSurvey_.id)

                try:
                    filename = fs.save(file.name, file)
                    file_path = fs.url(filename)
                except:
                    pass
                doc = DocumentPenugasan(
                    name=file.name,
                    path=file_path,
                    create_date=datetime.utcnow() + timedelta(hours=7),
                    update_date=datetime.utcnow() + timedelta(hours=7)
                )
                doc.save()

                Penugasan_ = Penugasan(
                    user=ObjectId(request.POST.get('user')),
                    kode=kode_survey1+'-'+kode_Penugasan,
                    jenissurvey=ObjectId(data_jenis.id),
                    surveyor=ObjectId(pt_surveyid),
                    lokasisurvey=ObjectId(LokasiSurvey_.id),
                    tanggal_penugasan=datetime.strptime(request.POST.get(
                        'tanggal_penugasan'), '%Y-%m-%d 00:00:00'),
                    target=datetime.strptime(
                        request.POST.get('sla'), '%Y-%m-%d 00:00:00'),
                    nospk=kode_Penugasan,
                    spk=ObjectId(doc.id)
                )

                Penugasan_.status.append(status)
                Penugasan_.status.append(status_assigned)
                Penugasan_.save()
                #filename = fs.save(file.name, file)
                #file_path = fs.url(filename)
                # doc = DocumentPenugasan(
                #    name=file.name,
                #    path=file_path,
                # )
                # doc.save()

                # add status assigned di tabel lokasisurvey
                # try:
                #    lok_survey = LokasiSurvey.objects.get(id=request.POST.get('lokasisurvey'))
                #    status = {'status': 'assigned', 'date': datetime.utcnow(
                #        ) + timedelta(hours=7)}
                # except LokasiSurvey.DoesNotExist:
                #    lok_survey = None

                # if not lok_survey:
                #    return Response.badRequest(
                #        values='null',
                #        message='Lokasi Survey not found'
                #    )
                # status = {'status': 'assigned', 'date': datetime.utcnow(
                #        ) + timedelta(hours=7)}
                #statuspenugasan = [i for i,x in enumerate(lok_survey.status) if x['status']=='assigned']
                # if not statuspenugasan:
                #    lok_survey.status.append(status)
                #    lok_survey.save()
                # =============
                #Penugasan_.spk = ObjectId(doc.id)
                Penugasan_.save()

                #result = Penugasan.objects.get(id=ObjectId(Penugasan_.id)).serialize()
                # return Response.ok(
                #    values=result,
                #    message='Penugasan Created'
                # )
                # except Exception as e:
                #    return Response.badRequest(message=str(e))

        return JsonResponse({"state": "success"})


def addlokasisurvey(request):
    """
    Creator:
        1. cito
    Authorization:
        token       {string}
    Params:
        {JSON}
        user        {string}
        provinsi    {string}
        kabupaten_kota    {string}
        kecamatan    {string}
        desa        {string}
        latitude        {string}
        longitude        {string}
    Raises:
        Json        {Fail}

    Returns:
        Json        {Success}
    """
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    strjson = json.loads(request.body)
    if request.method == "POST":  # Add
        # try:
        # Duplicate filtering
        jenis_ = strjson["jenis"].upper()
        provinsi_ = strjson["provinsi"]
        kecamatan_ = strjson["kecamatan"]
        desa_ = strjson["desa"]
        try:
            data_provinsi = provinsi.objects.filter(name=provinsi_).first()
            if len(data_provinsi) == 0:
                return Response.badRequest(
                    values='null',
                    message='Provinsi doesnt exists'
                )
            data_kecamatan = kecamatan.objects.filter(name=kecamatan_).first()
            if len(data_kecamatan) == 0:
                return Response.badRequest(
                    values='null',
                    message='Kecamatan doesnt exists'
                )
            data_desa = desa.objects.filter(name=desa_).first()
            if len(data_desa) == 0:
                return Response.badRequest(
                    values='null',
                    message='Desa doesnt exists'
                )
        except Exception as e:
            print(e)
            return Response.badRequest(
                values='null',
                message='Provinsi/Kabupaten/Kota/Kecamatan/Desa doesnt exists1'
            )
        try:
            try:
                kabupaten_kota = strjson["kabupaten"]
                kabupaten_ = kabupaten.objects.filter(
                    name=kabupaten_kota).first()
                kab_kot = 'kab'
                if len(kabupaten_) == 0:
                    return Response.badRequest(
                        values='null',
                        message='Kabupaten doesnt exists'
                    )
                # LokasiSurvey_ = LokasiSurvey.objects.get(provinsi=provinsi,kabupaten=kabupaten_kota,
                #                kecamatan=kecamatan,desa=desa)
            except Exception as e:
                return Response.badRequest(
                    values='null',
                    message='Provinsi/Kabupaten/Kota/Kecamatan/Desa doesnt exists2'
                )
        except:
            try:
                kabupaten_kota = strjson["kota"]
                kota_ = kota.objects.filter(name=kabupaten_kota).first()
                kab_kot = 'kota'
                if len(kota_) == 0:
                    return Response.badRequest(
                        values='null',
                        message='Kabupaten doesnt exists'
                    )
                # LokasiSurvey_ = LokasiSurvey.objects.get(provinsi=provinsi,kota=kabupaten_kota,
                #                kecamatan=kecamatan,desa=desa)
            except Exception as e:
                return Response.badRequest(
                    values='null',
                    message='Provinsi/Kabupaten/Kota/Kecamatan/Desa doesnt exists3'
                )
        try:
            if kab_kot == 'kab':
                LokasiSurvey_ = LokasiSurvey.objects.get(provinsi=ObjectId(data_provinsi["id"]), kabupaten=ObjectId(kabupaten_["id"]),
                                                         kecamatan=ObjectId(data_kecamatan["id"]), desa=ObjectId(data_desa["id"]), jenis=jenis_)
            else:
                LokasiSurvey_ = LokasiSurvey.objects.get(provinsi=ObjectId(data_provinsi["id"]), kota=ObjectId(kota_["id"]),
                                                         kecamatan=ObjectId(data_kecamatan["id"]), desa=ObjectId(data_desa["id"]), jenis=jenis_)
            return Response.badRequest(
                values='null',
                message='Lokasi exists'
            )
        except LokasiSurvey.DoesNotExist:
            LokasiSurvey_ = LokasiSurvey()
            LokasiSurvey_.user = ObjectId(strjson["user"])
            LokasiSurvey_.provinsi = ObjectId(data_provinsi["id"])
            try:
                LokasiSurvey_.kabupaten = ObjectId(kabupaten_["id"])
            except:
                LokasiSurvey_.kota = ObjectId(kota_["id"])
            LokasiSurvey_.kecamatan = ObjectId(data_kecamatan["id"])
            LokasiSurvey_.desa = ObjectId(data_desa["id"])
            LokasiSurvey_.jenis = strjson["jenis"].upper()
            LokasiSurvey_.latitude = strjson["latitude"]
            LokasiSurvey_.longitude = strjson["longitude"]

            LokasiSurvey_.save()
            return JsonResponse({"state": "success"})
        return JsonResponse({"state": "fail", "action": "add fail"})
    return JsonResponse({"state": "fail", "action": "none"})


def getlokasisurvey(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    strjson = json.loads(request.body)
    field = strjson["field"].lower()
    value_ = strjson["value"].upper()
    jenis = strjson.get("jenis", None)
    pp_jenis = []
    if jenis:
        # try:
        #data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
        pp_jenis = [{
            '$match': {
                'jenis': jenis.upper()
            }
        }]
        # except JenisSurvey.DoesNotExist:
        #    return Response.badRequest(
        #        values=[],
        #        message='Jenis Survey not found'
        #    )

    page = int(strjson.get('page', 0)) - 1
    skip = []
    if page >= 0:
        skip = [{'$skip': 20 * page},
                {'$limit': 20}]

    pipeline = pp_jenis + [
        {
            '$lookup': {
                'from': 'provinsi',
                'localField': 'provinsi',
                'foreignField': '_id',
                'as': 'provinsi'
            }
        }, {
            '$unwind': {
                'path': '$provinsi',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'kabupaten',
                'localField': 'kabupaten',
                'foreignField': '_id',
                'as': 'kabupaten'
            }
        }, {
            '$unwind': {
                'path': '$kabupaten',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'kota',
                'localField': 'kota',
                'foreignField': '_id',
                'as': 'kota'
            }
        }, {
            '$unwind': {
                'path': '$kota',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'kecamatan',
                'localField': 'kecamatan',
                'foreignField': '_id',
                'as': 'kecamatan'
            }
        }, {
            '$unwind': {
                'path': '$kecamatan',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'desa',
                'localField': 'desa',
                'foreignField': '_id',
                'as': 'desa'
            }
        }, {
            '$unwind': {
                'path': '$desa',
                'preserveNullAndEmptyArrays': True
            }
        }
    ]

    if field == 'all':
        pipe = pipeline + skip
        agg_cursor = LokasiSurvey.objects.aggregate(*pipe)

        LokasiSurvey_ = list(agg_cursor)
    else:
        if field == 'provinsi':
            try:
                data_ = provinsi.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except provinsi.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'provinsi': ObjectId(data_.id)
                }
            }]
        if field == 'kabupaten':
            try:
                data_ = kabupaten.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except kabupaten.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'kabupaten': ObjectId(data_.id)
                }
            }]
        if field == 'kota':
            try:
                data_ = kota.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except kota.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'kota': ObjectId(data_.id)
                }
            }]
        if field == 'kecamatan':
            try:
                data_ = kecamatan.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except kecamatan.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'kecamatan': ObjectId(data_.id)
                }
            }]
        if field == 'desa':
            try:
                data_ = desa.objects.get(name=value_)
                value_ = ObjectId(data_.id)
            except desa.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi Survey not found'
                )
            match = [{
                '$match': {
                    'desa': ObjectId(data_.id)
                }
            }]
        # try:
        if field == 'status':
            match = [{
                '$addFields': {
                    'last': {
                        '$arrayElemAt': [
                            '$status', -1
                        ]
                    }
                }
            }, {
                '$match': {
                    'last.status': value_.lower()
                }
            }, {
                '$project': {
                    'last': 0
                }
            }]
            # if value_.lower()=='created':
            #    LokasiSurvey_ = LokasiSurvey.objects.filter(status__0__status=value_.lower(),status__1__exists=False)
            # else:
            #    LokasiSurvey_ = LokasiSurvey.objects.filter(status__status=value_.lower())
            # else:
            #    LokasiSurvey_ = LokasiSurvey.objects.filter(**{field: value_})
        if field == 'jenis':
            match = [{
                '$match': {
                    'jenis': value_.upper()
                }
            }]
        # except LokasiSurvey.DoesNotExist:
        #    return Response.badRequest(
        #        values=[],
        #        message='Lokasi Survey not found'
        #    )
        pipe = match + pipeline + skip
        agg_cursor = LokasiSurvey.objects.aggregate(*pipe)

        LokasiSurvey_ = list(agg_cursor)

    if len(LokasiSurvey_) > 0:
        return Response.ok(
            values=json.loads(json.dumps(LokasiSurvey_, default=str)),
            message=f'{len(LokasiSurvey_)} Data'
        )
    else:
        return Response.badRequest(
            values=[],
            message='Lokasi Survey not found'
        )


def addpenugasan(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})

    if request.method == "POST":  # Add
        try:
            tgl_penugasan = datetime.strptime(request.POST.get(
                'tanggal_penugasan'), '%Y-%m-%d 00:00:00'),
            tgl_sla = datetime.strptime(
                request.POST.get('sla'), '%Y-%m-%d 00:00:00'),
            if tgl_sla <= tgl_penugasan:
                return Response.badRequest(
                    values=[],
                    message='Tanggal SLA harus di atas tanggal penugasan'
                )
            file = request.FILES['doc']
            if not file:
                return Response.badRequest(message='No File Upload')
            fs = FileSystemStorage(
                location=f'{settings.MEDIA_ROOT}/survey/spk/',
                base_url=f'{settings.MEDIA_URL}/survey/spk/'
            )
            try:
                lok_survey = LokasiSurvey.objects.get(
                    id=request.POST.get('lokasisurvey'))
            except LokasiSurvey.DoesNotExist:
                return Response.badRequest(
                    values=[],
                    message='Lokasi survey tidak ada'
                )

            data_provinsi = provinsi.objects.filter(
                id=lok_survey.provinsi.id).first()
            if data_provinsi is None:
                return Response.badRequest(
                    values=[],
                    message='Provinsi tidak ada'
                )
            kode_survey1 = request.POST.get(
                'jenissurvey').upper()+'-'+data_provinsi.prefix.upper()
            try:
                kode_survey = Penugasan.objects(
                    kode__startswith=kode_survey1.upper()).order_by('-kode').first()
                lst_kode = kode_survey['kode'].split('-')
                if len(lst_kode) > 0:
                    kode_Penugasan = str(int(lst_kode[2])+1).zfill(5)
                else:
                    kode_Penugasan = str(int(kode_survey['kode'])+1).zfill(5)
            except:
                kode_Penugasan = '1'.zfill(5)

            status = {'status': 'created', 'date': datetime.utcnow(
            ) + timedelta(hours=7)}
            # Duplicate filtering
            try:
                data_jenis = JenisSurvey.objects.get(
                    jenis=request.POST.get('jenissurvey').upper())
            except JenisSurvey.DoesNotExist:
                return Response.badRequest(
                    values='null',
                    message='jenissurvey not found'
                )
            Penugasan_ = Penugasan(
                user=ObjectId(request.POST.get('user')),
                kode=kode_survey1+'-'+kode_Penugasan,
                jenissurvey=ObjectId(data_jenis.id),
                surveyor=ObjectId(request.POST.get('surveyor')),
                lokasisurvey=ObjectId(request.POST.get('lokasisurvey')),
                tanggal_penugasan=datetime.strptime(request.POST.get(
                    'tanggal_penugasan'), '%Y-%m-%d 00:00:00'),
                target=datetime.strptime(
                    request.POST.get('sla'), '%Y-%m-%d 00:00:00'),
                nospk=request.POST.get('nospk').upper(),
            )

            Penugasan_.status.append(status)
            Penugasan_.save()
            filename = fs.save(file.name, file)
            file_path = fs.url(filename)
            doc = DocumentPenugasan(
                name=file.name,
                path=file_path,
            )
            doc.save()

            # add status assigned di tabel lokasisurvey
            # try:
            #    lok_survey = LokasiSurvey.objects.get(
            #        id=request.POST.get('lokasisurvey'))
            status = {'status': 'assigned', 'date': datetime.utcnow(
            ) + timedelta(hours=7)}
            # except LokasiSurvey.DoesNotExist:
            #    lok_survey = None

            # if not lok_survey:
            #    return Response.badRequest(
            #        values='null',
            #        message='Lokasi Survey not found'
            #    )
            statuspenugasan = [i for i, x in enumerate(
                lok_survey.status) if x['status'] == 'assigned']
            if not statuspenugasan:
                lok_survey.status.append(status)
                lok_survey.save()
            # =============
            Penugasan_.spk = ObjectId(doc.id)
            Penugasan_.save()

            result = Penugasan.objects.get(
                id=ObjectId(Penugasan_.id)).serialize()
            return Response.ok(
                values=result,
                message='Penugasan Created'
            )
        except Exception as e:
            return Response.badRequest(message=str(e))
    else:
        return HttpResponse('Post Only')


def editpenugasan(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret,user = authenticate_credentials(token)
    # if False == ret or None ==user:
    #    return JsonResponse({"state":"fail"})

    if request.method == 'POST':
        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)
            kode_Penugasan = data["kode"].lower()
            field = data["field"].lower()
            value = data["value"].upper()

            penugasan = Penugasan.objects.get(kode=kode_Penugasan)

            if not penugasan:
                return Response.badRequest(
                    values='null',
                    message='penugasan not found'
                )
            penugasan.update(**{field: value})
            penugasan = Penugasan.objects.get(kode=kode_Penugasan)
            return Response.ok(
                values=penugasan.serialize(),
                message='Update Success'
            )
        except Exception as e:
            print(e)
            return HttpResponse(e)
    else:
        return HttpResponse('Post Only')

    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})


def uploadspk(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})

    if request.method == "POST":
        try:
            file = request.FILES['doc']
            if not file:
                return Response.badRequest(message='No File Upload')
            fs = FileSystemStorage(
                location=f'{settings.MEDIA_ROOT}/survey/spk/',
                base_url=f'{settings.MEDIA_URL}/survey/spk/'
            )

            penugasan_ = Penugasan.objects.get(
                kode=request.POST.get('kode').upper())

            if not penugasan_:
                return Response.badRequest(
                    values='null',
                    message='penugasan not found'
                )

            filename = fs.save(file.name, file)
            file_path = fs.url(filename)
            doc = DocumentPenugasan(
                name=file.name,
                path=file_path,
            )
            doc.save()

            penugasan_.spk = ObjectId(doc.id)
            penugasan_.save()

            result = Penugasan.objects.get(
                id=ObjectId(penugasan_.id)).serialize()
            return Response.ok(
                values=result,
                message='Upload spk success'
            )
        except Exception as e:
            return Response.badRequest(message=str(e))
    else:
        return HttpResponse('Post Only')


def getpenugasan(request):
    try:
        # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
        # ret,user = authenticate_credentials(token)
        # if False == ret or None ==user:
        #    return JsonResponse({"state":"fail"})
        strjson = json.loads(request.body.decode("utf-8"))
        #user = strjson["user"]
        field = strjson["field"].lower()
        value = strjson["value"].lower()
        jenis = strjson["jenis"]
        status = strjson.get("status", None)
        page = int(strjson.get('page', 0)) - 1
        skip = []
        if page >= 0:
            skip = [{'$skip': 20 * page},
                    {'$limit': 20}]
        result = []

        pipeline = [
            {
                '$lookup': {
                    'from': 'user_info',
                    'localField': 'user',
                    'foreignField': '_id',
                    'as': 'user'
                }
            }, {
                '$unwind': {
                    'path': '$user',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'user_info',
                    'localField': 'assignfrom',
                    'foreignField': '_id',
                    'as': 'assignfrom'
                }
            }, {
                '$unwind': {
                    'path': '$assignfrom',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'surveyor',
                    'localField': 'assignto',
                    'foreignField': '_id',
                    'as': 'assignto'
                }
            }, {
                '$unwind': {
                    'path': '$assignto',
                    'preserveNullAndEmptyArrays': True
                }
            },
             {
                 '$lookup': {
                     'from': 'user_info',
                     'localField': 'assignfrom1',
                     'foreignField': '_id',
                     'as': 'assignfrom1'
                 }
             }, {
                 '$unwind': {
                     'path': '$assignfrom1',
                     'preserveNullAndEmptyArrays': True
                 }
             }, {
                 '$lookup': {
                     'from': 'user_info',
                     'localField': 'assignto1',
                     'foreignField': '_id',
                     'as': 'assignto1'
                 }
             }, {
                 '$unwind': {
                     'path': '$assignto1',
                     'preserveNullAndEmptyArrays': True
                 }
             }, 
            {
                '$lookup': {
                    'from': 'document_penugasan',
                    'localField': 'spk',
                    'foreignField': '_id',
                    'as': 'spk'
                }
            }, {
                '$unwind': {
                    'path': '$spk',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'jenis_survey',
                    'localField': 'jenissurvey',
                    'foreignField': '_id',
                    'as': 'jenissurvey'
                }
            }, {
                '$unwind': {
                    'path': '$jenissurvey',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'surveyor',
                    'localField': 'surveyor',
                    'foreignField': '_id',
                    'as': 'surveyor'
                }
            }, {
                '$unwind': {
                    'path': '$surveyor',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'lokasi_survey',
                    'localField': 'lokasisurvey',
                    'foreignField': '_id',
                    'as': 'lokasisurvey'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'provinsi',
                    'localField': 'lokasisurvey.provinsi',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.provinsi'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.provinsi',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kabupaten',
                    'localField': 'lokasisurvey.kabupaten',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.kabupaten'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.kabupaten',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kota',
                    'localField': 'lokasisurvey.kota',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.kota'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.kota',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kecamatan',
                    'localField': 'lokasisurvey.kecamatan',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.kecamatan'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.kecamatan',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'desa',
                    'localField': 'lokasisurvey.desa',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.desa'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.desa',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$addFields': {
                    'last': {
                        '$arrayElemAt': ['$status', -1]
                    },
                }
            }, {
                '$sort': {
                    'last.date': -1
                }
            }, {
                '$project': {
                    'surveyor.jenissurvey': 0,
                    'last': 0
                }
            }
        ]

        def getByStatus():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
            except JenisSurvey.DoesNotExist:
                return {}

            pp = [
                {
                    '$match': {
                        'jenissurvey': ObjectId(data_jenis.id)
                    }
                },
                {
                    '$addFields': {
                        'last': {
                            '$arrayElemAt': [
                                '$status', -1
                            ]
                        }
                    }
                }, {
                    '$match': {
                        'last.status': value
                    }
                }, {
                    '$project': {
                        'last': 0
                    }
                }
            ]

            pipe = pp + pipeline + skip

            agg_cursor = Penugasan.objects.aggregate(*pipe)
            result = list(agg_cursor)

            return result

        def getByJenis():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=value.upper())
            except JenisSurvey.DoesNotExist:
                return {}
            data = Penugasan.objects(jenissurvey=ObjectId(data_jenis.id))
            #serializer = PenugasanSerializer(data, many=True)
            # result=serializer.data
            result = []
            if page < 0:
                for _penugasan in data:
                    result.append(_penugasan.serialize())
            else:
                if page < len(data):
                    if (len(data) - page) < 20:
                        endrow = page + (len(data) - page)
                    else:
                        endrow = page + 20
                    for i in range(page, endrow):
                        result.append(data[i].serialize())
            return result

        def getByUser():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
            except JenisSurvey.DoesNotExist:
                return {}
            data = Penugasan.objects(user=ObjectId(
                value), jenissurvey=ObjectId(data_jenis.id))
            result = []
            if page < 0:
                for _penugasan in data:
                    result.append(_penugasan.serialize())
            else:
                if page < len(data):
                    if (len(data) - page) < 20:
                        endrow = page + (len(data) - page)
                    else:
                        endrow = page + 20
                    for i in range(page, endrow):
                        result.append(data[i].serialize())
            return result

        def getByField():
            pp_status = []
            if status is not None:
                pp_status = [
                    {
                        '$addFields': {
                            'last': {
                                '$arrayElemAt': [
                                    '$status', -1
                                ]
                            }
                        }
                    }, {
                        '$match': {
                            'last.status': status
                        }
                    }, {
                        '$project': {
                            'last': 0
                        }
                    }
                ]

            try:
                data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
            except JenisSurvey.DoesNotExist:
                return {}
                # return Response.badRequest(
                #    values='null',
                #    message='jenissurvey not found'
                # )

            if value.upper() == 'ALL':
                pp = [{
                    '$match': {
                        'jenissurvey': ObjectId(data_jenis.id)
                    }
                }]

                pipe = pp + pipeline + skip

                agg_cursor = Penugasan.objects.aggregate(*pipe)
                result = list(agg_cursor)
            else:
                if field.lower() == 'assignto1':
                    value_ = ObjectId(value)
                else:
                    value_ = value
                # data = Penugasan.objects.filter(
                #    **{field: value}, jenissurvey=ObjectId(data_jenis.id))
                pp = [{
                    '$match': {
                        'jenissurvey': ObjectId(data_jenis.id),
                        field: value_
                    }
                }]

                pipe = pp + pp_status + pipeline + skip

                agg_cursor = Penugasan.objects.aggregate(*pipe)
                result = list(agg_cursor)
            # result=[]

            return result

        switcher = {
            "status": getByStatus,
            "jenissurvey": getByJenis,
            "user": getByUser,
        }

        _result = switcher.get(field, getByField)

        if not _result:
            return Response.badRequest(
                message='Wrong Field Parameter'
            )

        result = _result()

        return Response.ok(
            values=json.loads(json.dumps(result, default=str)),
            message=f'{len(result)} Data'
        )
    except Exception as e:
        print(e)
        return HttpResponse(e)


def countPenugasanProvinsi(request):
    try:

        result = []

        def allLocation():
            location = Penugasan.objects.aggregate([
                # {
                #     '$lookup': {
                #         'from': 'jenis_survey',
                #         'localField': 'jenissurvey',
                #         'foreignField': '_id',
                #         'as': 'jenissurvey'
                #     }
                # },
                # {
                #     '$unwind': {
                #         'path': '$jenissurvey',
                #         'preserveNullAndEmptyArrays': True
                #     }
                # },
                # {
                #     '$lookup': {
                #         'from': 'lokasi_survey',
                #         'localField': 'lokasisurvey',
                #         'foreignField': '_id',
                #         'as': 'lokasi'
                #     }
                # },
                # {
                #     '$unwind': {
                #         'path': '$lokasi',
                #         'preserveNullAndEmptyArrays': True
                #     }
                # },
                # {
                #     '$lookup': {
                #         'from': 'hasil_survey',
                #         'localField': 'kode',
                #         'foreignField': 'kodeHasilSurvey',
                #         'as': 'hasilSurvey'
                #     }
                # },
                # {
                #     '$unwind': {
                #         'path': '$hasilSurvey',
                #         # 'preserveNullAndEmptyArrays': True
                #     }
                # },
                # {
                #     '$lookup': {
                #         'from': 'provinsi',
                #         'localField': 'lokasi.provinsi',
                #         'foreignField': '_id',
                #         'as': 'lokasi.provinsi'
                #     }
                # },
                # {
                #     '$unwind': {
                #         'path': '$lokasi.provinsi',
                #         'preserveNullAndEmptyArrays': True
                #     }
                # },
                # {
                #     '$group': {
                #         '_id': {
                #             'provinsi': '$lokasi.provinsi.name',

                #         },
                #         'count_AI': {
                #             '$sum': {
                #                 '$cond': [{
                #                     '$and': [
                #                         {
                #                             '$eq': ['$jenissurvey.jenis', 'AI']
                #                         }
                #                     ]
                #                 }, 1, 0]
                #             }
                #         },
                #         'count_BTS': {
                #             '$sum': {
                #                 '$cond': [{
                #                     '$and': [
                #                         {
                #                             '$eq': ['$jenissurvey.jenis', 'BTS']
                #                         }
                #                     ]
                #                 }, 1, 0]
                #             }
                #         },
                #         'surveyed_AI': {
                #             '$sum': {
                #                 '$cond': [{
                #                     '$and': [
                #                         {
                #                             '$eq': [
                #                                 "$kode", "$hasilSurvey.kodeHasilSurvey"
                #                             ]
                #                         },
                #                         {
                #                             '$eq': ['$jenissurvey.jenis', 'AI']
                #                         }
                #                     ]
                #                 }, 1, 0]
                #             }
                #         },
                #         'surveyed_BTS': {
                #             '$sum': {
                #                 '$cond': [{
                #                     '$and': [
                #                         {
                #                             '$eq': [
                #                                 "$kode", "$hasilSurvey.kodeHasilSurvey"
                #                             ]
                #                         },
                #                         {
                #                             '$eq': ['$jenissurvey.jenis', 'BTS']
                #                         }
                #                     ]
                #                 }, 1, 0]
                #             }
                #         },
                #         'issue_AI': {
                #             '$sum': {
                #                 '$cond': [{
                #                     '$and': [
                #                         {
                #                             '$exist': [
                #                                 "$kode", "$hasilSurvey.kodeHasilSurvey"
                #                             ]
                #                         },
                #                         {
                #                             '$eq': ['$jenissurvey.jenis', 'AI']
                #                         }
                #                     ]
                #                 }, 1, 0]
                #             }
                #         },
                #     },
                # },

                {
                    '$lookup': {
                        'from': 'hasil_survey',
                        'localField': 'kode',
                        'foreignField': 'kodeHasilSurvey',
                        'as': 'hasil'
                    }
                }, {
                    '$lookup': {
                        'from': 'hasil_surveybts',
                        'localField': 'kode',
                        'foreignField': 'kodeHasilSurvey',
                        'as': 'hasilbts'
                    }
                }, {
                    '$lookup': {
                        'from': 'jenis_survey',
                        'localField': 'jenissurvey',
                        'foreignField': '_id',
                        'as': 'jenissurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$jenissurvey',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'lokasi_survey',
                        'localField': 'lokasisurvey',
                        'foreignField': '_id',
                        'as': 'lokasisurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasisurvey',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'provinsi',
                        'localField': 'lokasisurvey.provinsi',
                        'foreignField': '_id',
                        'as': 'lokasisurvey.provinsi'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasisurvey.provinsi',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$addFields': {
                        'size': {
                            '$size': '$hasil'
                        },
                        'lastHasil': {
                            '$arrayElemAt': [
                                '$hasil', 0
                            ]
                        }
                    }
                }, {
                    '$addFields': {
                        'sizebts': {
                            '$size': '$hasilbts'
                        },
                        'lastHasilbts': {
                            '$arrayElemAt': [
                                '$hasilbts', 0
                            ]
                        }
                    }
                }, {
                    '$addFields': {
                        'issue': {
                            '$cond': {
                                'if': {
                                    '$ifNull': [
                                        '$lastHasil.issue', False
                                    ]
                                },
                                'then': '$lastHasil.issue',
                                'else': []
                            }
                        }
                    }
                }, {
                    '$addFields': {
                        'size_issue': {
                            '$size': '$issue'
                        }
                    }
                }, {
                    '$addFields': {
                        'issuebts': {
                            '$cond': {
                                'if': {
                                    '$ifNull': [
                                        '$lastHasilbts.issue', False
                                    ]
                                },
                                'then': '$lastHasilbts.issue',
                                'else': []
                            }
                        }
                    }
                }, {
                    '$addFields': {
                        'size_issuebts': {
                            '$size': '$issuebts'
                        }
                    }
                }, {
                    '$group': {
                        '_id': {
                            'provinsi': '$lokasisurvey.provinsi.name'
                        },
                        'count': {
                            '$sum': 1
                        },
                        'count_AI': {
                            '$sum': {
                                '$cond': [
                                    {
                                        '$and': [
                                            {
                                                '$eq': [
                                                    '$jenissurvey.jenis', 'AI'
                                                ]
                                            }
                                        ]
                                    }, 1, 0
                                ]
                            }
                        },
                        'count_BTS': {
                            '$sum': {
                                '$cond': [
                                    {
                                        '$and': [
                                            {
                                                '$eq': [
                                                    '$jenissurvey.jenis', 'BTS'
                                                ]
                                            }
                                        ]
                                    }, 1, 0
                                ]
                            }
                        },
                        'surveyed_AI': {
                            '$sum': {
                                '$cond': [
                                    {
                                        '$and': [
                                            {
                                                '$gt': [
                                                    '$size', 0
                                                ]
                                            }, {
                                                '$eq': [
                                                    '$jenissurvey.jenis', 'AI'
                                                ]
                                            }
                                        ]
                                    }, 1, 0
                                ]
                            }
                        },
                        'surveyed_BTS': {
                            '$sum': {
                                '$cond': [
                                    {
                                        '$and': [
                                            {
                                                '$gt': [
                                                    '$sizebts', 0
                                                ]
                                            }, {
                                                '$eq': [
                                                    '$jenissurvey.jenis', 'BTS'
                                                ]
                                            }
                                        ]
                                    }, 1, 0
                                ]
                            }
                        },
                        'issue_AI': {
                            '$sum': {
                                '$cond': [
                                    {
                                        '$and': [
                                            {
                                                '$gt': [
                                                    '$size_issue', 0
                                                ]
                                            }, {
                                                '$eq': [
                                                    '$jenissurvey.jenis', 'AI'
                                                ]
                                            }
                                        ]
                                    }, 1, 0
                                ]
                            }
                        },
                        'issue_BTS': {
                            '$sum': {
                                '$cond': [
                                    {
                                        '$and': [
                                            {
                                                '$gt': [
                                                    '$size_issuebts', 0
                                                ]
                                            }, {
                                                '$eq': [
                                                    '$jenissurvey.jenis', 'BTS'
                                                ]
                                            }
                                        ]
                                    }, 1, 0
                                ]
                            }
                        }
                    }
                },
                {
                    '$project': {
                        'provinsi': '$_id.provinsi',
                        'count_AI': '$count_AI',
                        'count_BTS': '$count_BTS',
                        'surveyed_AI': '$surveyed_AI',
                        'surveyed_BTS': '$surveyed_BTS',
                        'issue_AI': '$issue_AI',
                        'issue_BTS': '$issue_BTS',
                        'unsurveyed_AI': {'$subtract': ['$count_AI', '$surveyed_AI']},
                        'unsurveyed_BTS': {'$subtract': ['$count_BTS', '$surveyed_BTS']},
                        '_id': False
                    }
                }
            ])
            return [x for x in location]

        return Response.ok(
            values=allLocation(),
            message='Success'
        )
    except Exception as e:
        return Response.badRequest(
            values={},
            message=str(e)
        )


def countPenugasanDaily(request):
    try:
        penugasan = hasilSurvey.objects.aggregate([
            {
                "$project": {
                    'date': {
                        '$dateToString': {
                            'format': "%Y-%m-%d",
                            'date': "$tanggal_pembuatan"
                        }
                    },

                }
            },
            {
                '$group': {
                    '_id': {
                        'date': '$date'
                    },
                    'count': {
                        '$sum': 1
                    }
                }
            },

            {
                '$project': {
                    'date': '$_id.date',
                    'count': '$count',
                    '_id': False
                }
            }
        ])
        result = []
        for x in penugasan:
            result.append(x)

        return Response.ok(
            values=result,
            message='Success'
        )
    except Exception as e:
        return Response.badRequest(
            values={},
            message=str(e)
        )


def countPenugasan(request):
    try:
        def getPenugasanAI():
            hasil_survey = Penugasan.objects.filter(
                jenissurvey='5f16b4ba149882a98fc6655e').count()
            return hasil_survey

        def getPenugasanBTS():
            hasil_survey = Penugasan.objects.filter(
                jenissurvey='5f1521524f9c6764c713d73c').count()
            return hasil_survey

        def getPenugasanAISurvey():
            hasil_survey = hasilSurvey.objects.filter(
                status__status="Submitted", nomorSurvey='1')
            json_ret = []
            json_issue = []
            for k in hasil_survey:
                json_dict = {}
                json_dict["kodeSurvey"] = k
                json_ret.append(json_dict)
                if len(k.issue) > 0:
                    json_issue.append(1)

            return len(json_ret), len(json_issue)
        """
        def getPenugasanBTSSurvey():
            # try:
            hasil_survey = hasilSurveybts.objects(
                kodeHasilSurvey__startswith='BTS')
            # print(hasil_survey)
            hasil = []
            for x in list(set([x.kodeHasilSurvey for x in hasil_survey])):
                try:
                    penugasan = Penugasan.objects.get(kode=x.kodeHasilSurvey)
                except:
                    penugasan = None
                else:
                    hasil.append(penugasan)
            return len(hasil)
        """
        def getPenugasanBTSSurvey():
            hasil_survey = hasilSurveybts.objects.filter(
                status__status="Submitted", nomorSurvey='1')
            json_ret = []
            json_issue = []
            for k in hasil_survey:
                json_dict = {}
                json_dict["kodeSurvey"] = k
                json_ret.append(json_dict)
                if len(k.issue) > 0:
                    json_issue.append(1)

            return len(json_ret), len(json_issue)

        AISurvey, AISurveyIssue = getPenugasanAISurvey()
        BTSSurvey, BTSSurveyIssue = getPenugasanBTSSurvey()
        return Response.ok(
            values={
                "penugasan_ai": getPenugasanAI(),
                "penugasan_bts": getPenugasanBTS(),
                "penugasan_ai_surveyed": AISurvey,
                "penugasan_bts_surveyed": BTSSurvey,
                "penugasan_ai_issue": AISurveyIssue,
                "penugasan_bts_issue": BTSSurveyIssue,
            },
            message='Success'
        )
    except Exception as e:
        return Response.badRequest(
            values={},
            message=str(e)
        )


def assignpenugasan(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret,user = authenticate_credentials(token)
    # if False == ret or None ==user:
    #    return JsonResponse({"state":"fail"})
    # listbody=(request.body).decode("utf-8").split('&')
    # username = listbody[0].split('=')[1]
    # kode_Penugasan = (listbody[1].split('=')[1])
    # assignto = listbody[2].split('=')[1]
    # ke = listbody[3].split('=')[1]

    strjson = json.loads(request.body.decode("utf-8"))
    username = ObjectId(strjson["user"])
    kode_Penugasan = strjson["kode"].upper()
    assignto = ObjectId(strjson["assignto"])
    ke = strjson["ke"]

    if request.method == 'POST':

        dateNow = datetime.utcnow() + timedelta(hours=7)
        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)

            try:
                penugasan = Penugasan.objects.get(kode=data["kode"].upper())
            except Penugasan.DoesNotExist:
                penugasan = None

            if not penugasan:
                return Response.badRequest(
                    values='null',
                    message='penugasan not found'
                )
            if ke == '1':
                if penugasan.assignfrom or penugasan.assignto:
                    return Response.badRequest(
                        values=[],
                        message='Sudah ditugaskan'
                    )
                penugasan.assignfrom = ObjectId(data["user"])
                penugasan.assignto = ObjectId(data["assignto"])
                status = {'status': 'assigned', 'date': datetime.utcnow(
                ) + timedelta(hours=7)}
                penugasan.status.append(status)

                usersadminsurveyor = UserInfo.objects.filter(organization=ObjectId(
                    data["assignto"]), role__in=['5f13b362386bf295b4169eff'])
                #usersadmin = list(usersadmin['id'])
                userto_ = []
                for usr in usersadminsurveyor:
                    userto_.append(usr.username)

                notif = Message(
                    title='Tugaskan survey',
                    message='1 Penugasan Baru ' + kode_Penugasan,
                    userfrom=ObjectId(data["user"]),
                    userto=userto_,
                    redirect='/adminsurveyor',
                    status='new'
                )
                notif.save()

            if ke == '2':
                if penugasan.assignfrom1 or penugasan.assignto1:
                    return Response.badRequest(
                        values=[],
                        message='Sudah ditugaskan'
                    )
                penugasan.assignfrom1 = ObjectId(data["user"])
                penugasan.assignto1 = ObjectId(data["assignto"])
                status = {'status': 'on progress', 'date': datetime.utcnow(
                ) + timedelta(hours=7)}
                #penugasan.assign1 = datetime.utcnow() + timedelta(hours=7)
                # Penugasan.objects(kode = data["kode"]).update(set__assignfrom1=ObjectId(data["user"]),
                #    set__assignto1=ObjectId(data["assignto"]))
                penugasan.status.append(status)

                usersurveyor = UserInfo.objects.get(
                    id=ObjectId(data["assignto"]))

                notif = Message(
                    title='Tugaskan survey',
                    message='1 Penugasan Baru ' + kode_Penugasan,
                    userfrom=ObjectId(data["user"]),
                    userto=[usersurveyor.username],
                    redirect='/adminsurveyor',
                    status='new'
                )
                notif.save()
            penugasan.save()

            return Response.ok(
                # values=penugasan.serializeAssign(),
                values=penugasan.serialize(),
                message='Assign Success'
            )
        except Exception as e:
            print(e)
            return HttpResponse(e)
    else:
        return HttpResponse('Post Only')


def changestatuspenugasan(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret,user = authenticate_credentials(token)
    # if False == ret or None ==user:
    #    return JsonResponse({"state":"fail"})

    strjson = json.loads(request.body.decode("utf-8"))
    kode_Penugasan = strjson["kode"].lower()
    status = strjson["status"].lower()

    if request.method == 'POST':
        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)

            try:
                penugasan = Penugasan.objects.get(
                    kode=kode_Penugasan.strip().upper())
            except Penugasan.DoesNotExist:
                penugasan = None

            if not penugasan:
                return Response.badRequest(
                    values='null',
                    message='penugasan not found'
                )
            statuspenugasan = [i for i, x in enumerate(
                penugasan.status) if x['status'] == status]
            if statuspenugasan:
                # return Response.badRequest(
                #    values='null',
                #    message='status exists'
                # )
                return Response.ok(
                    values=penugasan.serialize(),
                    message='Update Success'
                )

            switcher = {
                "created": "created",
                "assigned": "assigned",
                "on progress": "on progress",
                "reviewed": "reviewed",
                "finished": "finished",
            }

            _status = switcher.get(status, 'None')
            if _status == 'None':
                return Response.badRequest(
                    values='null',
                    message='status not found'
                )
            status = {'status': _status, 'date': datetime.utcnow(
            ) + timedelta(hours=7)}
            penugasan.status.append(status)
            penugasan.save()
            return Response.ok(
                values=penugasan.serialize(),
                message='Update Success'
            )
        except Exception as e:
            print(e)
            return HttpResponse(e)
    else:
        return HttpResponse('Post Only')


def PosthasilSurvey(request):
    try:
        body_data = request.POST.dict()
        try:
            userrole = UserInfo.objects.get(id=body_data.get('userId'))
        except UserInfo.DoesNotExist:
            return Response.badRequest(
                values='null',
                message='User tidak ada'
            )

        if userrole.role.name != 'Staff Surveyor':
            return Response.badRequest(
                values='null',
                message='Anda tidak bisa submit hasil survey'
            )
        statusDefault = [{'status': 'Submitted', 'tanggal_pembuatan': datetime.utcnow(
        ) + timedelta(hours=7)}]
        if request.method == "POST":
            kode = body_data.get("kodeHasilSurvey").upper()
            nomorSurv = hasilSurvey.objects.filter(
                kodeHasilSurvey=kode)
            nomorsurvey = str(len(nomorSurv)+1)
            organization = body_data.get("organization")

            # searchKode = Penugasan.objects.filter(
            #     kode=kode)
            # if len(searchKode) < 1:  # Duplicate Kode
            #     return Response.badRequest(
            #         values='null',
            #         message='Kode Survey Tidak Ditemukan'
            #     )

            fileAkses = request.FILES['fileAkses']
            filePlang = request.FILES['filePlang']
            fileMarking = request.FILES['fileMarking']
            filePln = request.FILES['filePln']
            fileDenah = request.FILES['fileDenah']
            fileLanskap = request.FILES['fileLanskap']
            # if not file:
            #     return Response.badRequest(message='No File Upload')
            try:
                fs = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/survey/foto/',
                    base_url=f'{settings.MEDIA_URL}/survey/foto/'
                )

            except Exception as e:
                return Response.badRequest(message=str(e))
            filenameAkses = fs.save(fileAkses.name, fileAkses)
            file_pathAkses = fs.url(filenameAkses)

            filenamePlang = fs.save(filePlang.name, filePlang)
            file_pathPlang = fs.url(filenamePlang)

            filenamemarking = fs.save(fileMarking.name, fileMarking)
            file_pathmarking = fs.url(filenamemarking)

            filenamePln = fs.save(filePln.name, filePln)
            file_pathPln = fs.url(filenamePln)

            filenamegambarDenah = fs.save(fileDenah.name, fileDenah)
            file_pathgambarDenah = fs.url(filenamegambarDenah)

            filenamelanskap = fs.save(fileLanskap.name, fileLanskap)
            file_pathlanskap = fs.url(filenamelanskap)

            picData = Pic(
                namaPic=body_data.get("namaPic")if body_data.get(
                    "namaPic") != "" else '-',
                phonePic=body_data.get("phonePic")if body_data.get(
                    "phonePic") != "" else '-'
            )
            modaTransport = ModaTransportasi(
                darat=body_data.get("darat")if body_data.get(
                    "darat") != "" else '-',
                laut=body_data.get("laut")if body_data.get(
                    "laut") != "" else '-',
                udara=body_data.get("udara")if body_data.get(
                    "udara") != "" else '-',
                durasiPerjalanan=body_data.get("durasiPerjalanan")if body_data.get(
                    "durasiPerjalanan") != "" else '-',
                namaKotaKecamatan=body_data.get("namaKotaKecamatan")if body_data.get(
                    "namaKotaKecamatan") != "" else '-'
            )
            powerData = Power(
                idPelangganPLN=body_data.get("idPelangganPLN")if body_data.get(
                    "idPelangganPLN") != "" else '-',
                sumber_listrik=body_data.get("sumber_listrik")if body_data.get(
                    "sumber_listrik") != "" else '-',
                kapasitas_listrik=body_data.get("kapasitas_listrik")if body_data.get(
                    "kapasitas_listrik") != "" else '-',
                sumber_cadangan=body_data.get("sumber_cadangan")if body_data.get(
                    "sumber_cadangan") != "" else '-',
                jamOperasionalListrik=body_data.get("jamOperasionalListrik")if body_data.get(
                    "jamOperasionalListrik") != "" else '-',
                jamOperasionalLokal=body_data.get("jamOperasionalLokal")if body_data.get(
                    "jamOperasionalLokal") != "" else '-',
            )
            lisFotoData = FotoAI().instan(
                patternFoto().instan(filenameAkses, file_pathAkses),
                patternFoto().instan(filenamePlang, file_pathPlang),
                patternFoto().instan(filenamemarking, file_pathmarking),
                patternFoto().instan(filenamePln, file_pathPln),
                patternFoto().instan(filenamegambarDenah, file_pathgambarDenah),
                patternFoto().instan(filenamelanskap, file_pathlanskap),
            )

            if body_data.get("lainnya1Name") == '':
                lainya1dt = None
            else:
                lainya1dt = LainyaTemplate().instan(
                    body_data.get("lainnya1Name"),
                    body_data.get("lainnya1Qty")
                )
            if body_data.get("lainnya2Name") == '':
                lainya2dt = None
            else:
                lainya2dt = LainyaTemplate().instan(
                    body_data.get("lainnya2Name"),
                    body_data.get("lainnya2Qty")
                )
            deviceData = Device(
                pc=body_data.get("pc")if body_data.get("pc") != "" else '-',
                tablet=body_data.get("tablet")if body_data.get(
                    "tablet") != "" else '-',
                smartPhone=body_data.get("smartPhone")if body_data.get(
                    "smartPhone") != "" else '-',
                laptop=body_data.get("laptop")if body_data.get(
                    "laptop") != "" else '-',
                lainnya1=lainya1dt,
                lainnya2=lainya2dt
            )

            date_time_str = body_data.get("tanggalPelaksanaan")if body_data.get(
                "tanggalPelaksanaan") != "" else '-'
            if date_time_str == '-':
                date_time_obj = None
            else:
                date_time_obj = datetime.strptime(
                    date_time_str, '%Y-%m-%d %H:%M:%S.%f')

            networkData = Network(
                tipe=body_data.get("tipenetwork")if body_data.get(
                    "tipenetwork") != "" else '-',
                download=body_data.get("download")if body_data.get(
                    "download") != "" else '-',
                upload=body_data.get("upload")if body_data.get(
                    "upload") != "" else '-'
            )

            hasilSurv = hasilSurvey(
                kodeHasilSurvey=kode,
                nomorSurvey=nomorsurvey,
                pic=picData,
                user=body_data.get('userId'),
                namaLokasi=body_data.get("namaLokasi")if body_data.get(
                    "namaLokasi") != "" else '-',
                alamatLokasi=body_data.get("alamatLokasi")if body_data.get(
                    "alamatLokasi") != "" else '-',
                elevation=body_data.get("elevation")if body_data.get(
                    "elevation") != "" else '-',
                tipeBisnis=body_data.get("tipeBisnis")if body_data.get(
                    "tipeBisnis") != "" else '-',
                power=powerData,
                modaTransportasi=modaTransport,
                note=body_data.get("note")if body_data.get(
                    "note") != "" else '-',
                # jenisPeninjauan=body_data.get("jenisPeninjauan"),
                # solusiTeknologi=body_data.get("solusiTeknologi"),
                # catatan=body_data.get("catatan"),
                tanggalPelaksanaan=date_time_obj,
                # sisiInternalTeknisi=body_data.get("sisiInternalTeknisi"),
                # sisiPelanggan=body_data.get("sisiPelanggan"),
                # resume=body_data.get("resume"),
                longitude=body_data.get("longitude")if body_data.get(
                    "longitude") != "" else '-',
                latitude=body_data.get("latitude")if body_data.get(
                    "latitude") != "" else '-',
                status=statusDefault,
                listFoto=lisFotoData,
                device=deviceData,
                kategori=body_data.get("kategori")if body_data.get(
                    "kategori") != "" else '-',
                network=networkData,
                tanggal_pembuatan=datetime.utcnow() + timedelta(hours=7),
                tanggal_pembaruan=datetime.utcnow() + timedelta(hours=7)
            )
            serializer = hasilSurveySerializer(hasilSurv)
            result = serializer.data

            hasilSurv.save()

            usersadminsurveyor = UserInfo.objects.filter(
                organization=organization, role__in=['5f13b362386bf295b4169eff'])
            #usersadmin = list(usersadmin['id'])
            userto_ = []
            for usr in usersadminsurveyor:
                userto_.append(usr.username)

            notif = Message(
                title='Hasil survey',
                message='1 Hasil Survey Baru dari '+hasilSurv.user.username,
                userfrom=ObjectId(hasilSurv.user.id),
                userto=userto_,
                redirect='/adminsurveyor',
                status='new',
                created=datetime.utcnow() + timedelta(hours=7),
                updated=datetime.utcnow() + timedelta(hours=7)
            )
            notif.save()

            return Response.ok(
                message='Survey telah di submit',
                values=result
            )
    except Exception as e:
        return Response.badRequest(message=str(e))


def PosthasilSurveyRelokasi(request):
    try:
        body_data = request.POST.dict()
        try:
            userrole = UserInfo.objects.get(id=body_data.get('userId'))
        except UserInfo.DoesNotExist:
            return Response.badRequest(
                values='null',
                message='User tidak ada'
            )

        if userrole.role.name != 'Staff Surveyor':
            return Response.badRequest(
                values='null',
                message='Anda tidak bisa submit hasil survey'
            )
        statusDefault = [{'status': 'Submitted', 'tanggal_pembuatan': datetime.utcnow(
        ) + timedelta(hours=7)}]
        if request.method == "POST":
            koderelokasi = body_data.get("kodeRelokasi").upper()
            kode = body_data.get("kodeHasilSurvey").upper()
            nomorSurv = hasilSurvey.objects.filter(
                kodeHasilSurvey=koderelokasi)
            nomorsurvey = str(len(nomorSurv)+1)

            # searchKode = Penugasan.objects.filter(
            #     kode=kode)
            # if len(searchKode) < 1:  # Duplicate Kode
            #     return Response.badRequest(
            #         values='null',
            #         message='Kode Survey Tidak Ditemukan'
            #     )

            fileAkses = request.FILES['fileAkses']
            filePlang = request.FILES['filePlang']
            fileMarking = request.FILES['fileMarking']
            filePln = request.FILES['filePln']
            fileDenah = request.FILES['fileDenah']
            fileLanskap = request.FILES['fileLanskap']
            # if not file:
            #     return Response.badRequest(message='No File Upload')
            try:
                fs = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/survey/foto/relokasi/',
                    base_url=f'{settings.MEDIA_URL}/survey/foto/relokasi/'
                )

            except Exception as e:
                return Response.badRequest(message=str(e))
            filenameAkses = fs.save(fileAkses.name, fileAkses)
            file_pathAkses = fs.url(filenameAkses)

            filenamePlang = fs.save(filePlang.name, filePlang)
            file_pathPlang = fs.url(filenamePlang)

            filenamemarking = fs.save(fileMarking.name, fileMarking)
            file_pathmarking = fs.url(filenamemarking)

            filenamePln = fs.save(filePln.name, filePln)
            file_pathPln = fs.url(filenamePln)

            filenamegambarDenah = fs.save(fileDenah.name, fileDenah)
            file_pathgambarDenah = fs.url(filenamegambarDenah)

            filenamelanskap = fs.save(fileLanskap.name, fileLanskap)
            file_pathlanskap = fs.url(filenamelanskap)

            picData = Pic(
                namaPic=body_data.get("namaPic")if body_data.get(
                    "namaPic") != "" else '-',
                phonePic=body_data.get("phonePic")if body_data.get(
                    "phonePic") != "" else '-'
            )
            modaTransport = ModaTransportasi(
                darat=body_data.get("darat")if body_data.get(
                    "darat") != "" else '-',
                laut=body_data.get("laut")if body_data.get(
                    "laut") != "" else '-',
                udara=body_data.get("udara")if body_data.get(
                    "udara") != "" else '-',
                durasiPerjalanan=body_data.get("durasiPerjalanan")if body_data.get(
                    "durasiPerjalanan") != "" else '-',
                namaKotaKecamatan=body_data.get("namaKotaKecamatan")if body_data.get(
                    "namaKotaKecamatan") != "" else '-'
            )
            powerData = Power(
                idPelangganPLN=body_data.get("idPelangganPLN")if body_data.get(
                    "idPelangganPLN") != "" else '-',
                sumber_listrik=body_data.get("sumber_listrik")if body_data.get(
                    "sumber_listrik") != "" else '-',
                kapasitas_listrik=body_data.get("kapasitas_listrik")if body_data.get(
                    "kapasitas_listrik") != "" else '-',
                sumber_cadangan=body_data.get("sumber_cadangan")if body_data.get(
                    "sumber_cadangan") != "" else '-',
                jamOperasionalListrik=body_data.get("jamOperasionalListrik")if body_data.get(
                    "jamOperasionalListrik") != "" else '-',
                jamOperasionalLokal=body_data.get("jamOperasionalLokal")if body_data.get(
                    "jamOperasionalLokal") != "" else '-',
            )
            lisFotoData = FotoAI().instan(
                patternFoto().instan(filenameAkses, file_pathAkses),
                patternFoto().instan(filenamePlang, file_pathPlang),
                patternFoto().instan(filenamemarking, file_pathmarking),
                patternFoto().instan(filenamePln, file_pathPln),
                patternFoto().instan(filenamegambarDenah, file_pathgambarDenah),
                patternFoto().instan(filenamelanskap, file_pathlanskap),
            )

            if body_data.get("lainnya1Name") == '':
                lainya1dt = None
            else:
                lainya1dt = LainyaTemplate().instan(
                    body_data.get("lainnya1Name"),
                    body_data.get("lainnya1Qty")
                )
            if body_data.get("lainnya2Name") == '':
                lainya2dt = None
            else:
                lainya2dt = LainyaTemplate().instan(
                    body_data.get("lainnya2Name"),
                    body_data.get("lainnya2Qty")
                )
            deviceData = Device(
                pc=body_data.get("pc")if body_data.get("pc") != "" else '-',
                tablet=body_data.get("tablet")if body_data.get(
                    "tablet") != "" else '-',
                smartPhone=body_data.get("smartPhone")if body_data.get(
                    "smartPhone") != "" else '-',
                laptop=body_data.get("laptop")if body_data.get(
                    "laptop") != "" else '-',
                lainnya1=lainya1dt,
                lainnya2=lainya2dt
            )

            date_time_str = body_data.get("tanggalPelaksanaan")if body_data.get(
                "tanggalPelaksanaan") != "" else '-'
            if date_time_str == '-':
                date_time_obj = None
            else:
                date_time_obj = datetime.strptime(
                    date_time_str, '%Y-%m-%d %H:%M:%S.%f')

            if nomorsurvey == '1':
                hasilSurv_ = hasilSurvey(
                    kodeHasilSurvey=kode,
                    nomorSurvey=nomorsurvey,
                    pic=Pic(),
                    user=body_data.get('userId'),
                    namaLokasi='-',
                    alamatLokasi='-',
                    elevation='-',
                    tipeBisnis='-',
                    power=Power(),
                    modaTransportasi=ModaTransportasi(),
                    note='-',
                    # jenisPeninjauan=body_data.get("jenisPeninjauan"),
                    # solusiTeknologi=body_data.get("solusiTeknologi"),
                    # catatan=body_data.get("catatan"),
                    tanggalPelaksanaan=date_time_obj,
                    # sisiInternalTeknisi=body_data.get("sisiInternalTeknisi"),
                    # sisiPelanggan=body_data.get("sisiPelanggan"),
                    # resume=body_data.get("resume"),
                    longitude='-',
                    latitude='-',
                    status=statusDefault,
                    listFoto=FotoAI(),
                    device=Device(),
                    relokasi=Relokasi(),
                    kategori=body_data.get("kategori")if body_data.get(
                        "kategori") != "" else '-',
                    network=Network(),
                    tanggal_pembuatan=datetime.utcnow() + timedelta(hours=7),
                    tanggal_pembaruan=datetime.utcnow() + timedelta(hours=7)
                )
                hasilSurv_.save()
            try:
                _kodelama = str(hasilSurv_.id)
            except:
                hsl_sur = hasilSurvey.objects.filter(
                    kodeHasilSurvey=kode, nomorSurvey='1', latitude='-').first()
                # if len(hsl_sur)>0:
                #    _kodelama = str(hsl_sur[0].id)
                # else:
                _kodelama = str(hsl_sur.id)
            relokasiData = Relokasi(
                provinsi=body_data.get("provinsi")if body_data.get(
                    "provinsi") != "" else '-',
                kab_kota=body_data.get("kab_kota")if body_data.get(
                    "kab_kota") != "" else '-',
                kecamatan=body_data.get("kecamatan")if body_data.get(
                    "kecamatan") != "" else '-',
                desa=body_data.get("desa")if body_data.get(
                    "desa") != "" else '-',
                alasan=body_data.get("alasan")if body_data.get(
                    "alasan") != "" else '-',
                kodelama=_kodelama
            )

            networkData = Network(
                tipe=body_data.get("tipenetwork")if body_data.get(
                    "tipenetwork") != "" else '-',
                download=body_data.get("download")if body_data.get(
                    "download") != "" else '-',
                upload=body_data.get("upload")if body_data.get(
                    "upload") != "" else '-'
            )

            hasilSurv = hasilSurvey(
                kodeHasilSurvey=koderelokasi,
                nomorSurvey=nomorsurvey,
                pic=picData,
                user=body_data.get('userId'),
                namaLokasi=body_data.get("namaLokasi")if body_data.get(
                    "namaLokasi") != "" else '-',
                alamatLokasi=body_data.get("alamatLokasi")if body_data.get(
                    "alamatLokasi") != "" else '-',
                elevation=body_data.get("elevation")if body_data.get(
                    "elevation") != "" else '-',
                tipeBisnis=body_data.get("tipeBisnis")if body_data.get(
                    "tipeBisnis") != "" else '-',
                power=powerData,
                modaTransportasi=modaTransport,
                note=body_data.get("note")if body_data.get(
                    "note") != "" else '-',
                # jenisPeninjauan=body_data.get("jenisPeninjauan"),
                # solusiTeknologi=body_data.get("solusiTeknologi"),
                # catatan=body_data.get("catatan"),
                tanggalPelaksanaan=date_time_obj,
                # sisiInternalTeknisi=body_data.get("sisiInternalTeknisi"),
                # sisiPelanggan=body_data.get("sisiPelanggan"),
                # resume=body_data.get("resume"),
                longitude=body_data.get("longitude")if body_data.get(
                    "longitude") != "" else '-',
                latitude=body_data.get("latitude")if body_data.get(
                    "latitude") != "" else '-',
                status=statusDefault,
                listFoto=lisFotoData,
                device=deviceData,
                relokasi=relokasiData,
                kategori=body_data.get("kategori")if body_data.get(
                    "kategori") != "" else '-',
                network=networkData,
                tanggal_pembuatan=datetime.utcnow() + timedelta(hours=7),
                tanggal_pembaruan=datetime.utcnow() + timedelta(hours=7)
            )

            hasilSurv.save()

            serializer = hasilSurveySerializer(hasilSurv)
            result = serializer.data

            # url = "http://127.0.0.1:13000/survey/changestatuspenugasan/"
            # headers = {
            # 'Content-Type': 'application/json',
            # }
            # dicto={
            #     "kode":kode,
            #     "status":"on progress"
            # }
            # response = requests.request("POST", url, headers=headers, data = json.dumps(dicto))
            # print(response)
            return Response.ok(
                message='Survey telah di submit',
                values=result
            )
    except Exception as e:
        return Response.badRequest(message=str(e))


def getAllHasilSurveyByKode(request):
    try:
        strjson = json.loads(request.body)
        kodeSurvey = strjson['kodeSurvey'].upper()
        searchKode = hasilSurvey.objects.filter(kodeHasilSurvey=kodeSurvey)
        listResult = []
        if len(searchKode) > 0:
            for item in searchKode:
                print(item)
                serializer = hasilSurveySerializer(item)
                result = serializer.data
                listResult.append(result)
            return Response.okReturnCount(values=listResult, count=len(searchKode), message='Data berhasil terambil')
        else:
            return Response.badRequest(message='Kode '+kodeSurvey+' tidak Ditemukan')
    except Exception as e:
        return Response.badRequest(message=str(e))


def getsurveylogai(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            #status_ = body_data.get('status')
            page = int(body_data.get('page', 0)) - 1
            skip = []
            if page >= 0:
                skip = [{'$skip': 20 * page},
                        {'$limit': 20}]
            pipeline = [
                {
                    '$lookup': {
                        'from': 'user_info',
                        'localField': 'user',
                        'foreignField': '_id',
                        'as': 'user'
                    }
                }, {
                    '$unwind': {
                        'path': '$user',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'surveyor',
                        'localField': 'user.organization',
                        'foreignField': '_id',
                        'as': 'user.organization'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.organization',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'userrole',
                        'localField': 'user.role',
                        'foreignField': '_id',
                        'as': 'user.role'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.role',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'document_user',
                        'localField': 'user.doc',
                        'foreignField': '_id',
                        'as': 'user.doc'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.doc',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$group': {
                        '_id': '$kodeHasilSurvey',
                        'data': {
                            '$push': '$$ROOT'
                        }
                    }
                }, {
                    '$lookup': {
                        'from': 'penugasan',
                        'localField': '_id',
                        'foreignField': 'kode',
                        'as': 'penugasan'
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'lokasi_survey',
                        'localField': 'penugasan.lokasisurvey',
                        'foreignField': '_id',
                        'as': 'lokasisurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasisurvey',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'provinsi',
                        'localField': 'lokasisurvey.provinsi',
                        'foreignField': '_id',
                        'as': 'lokasi.provinsi'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.provinsi',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kabupaten',
                        'localField': 'lokasisurvey.kabupaten',
                        'foreignField': '_id',
                        'as': 'lokasi.kabupaten'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kabupaten',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kota',
                        'localField': 'lokasisurvey.kota',
                        'foreignField': '_id',
                        'as': 'lokasi.kota'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kota',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kecamatan',
                        'localField': 'lokasisurvey.kecamatan',
                        'foreignField': '_id',
                        'as': 'lokasi.kecamatan'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kecamatan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'desa',
                        'localField': 'lokasisurvey.desa',
                        'foreignField': '_id',
                        'as': 'lokasi.desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.desa',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$addFields': {
                        'last': {
                            '$arrayElemAt': ['$data', -1]
                        },
                    }
                }, {
                    '$sort': {
                        'last.tanggal_pembaruan': -1
                    }
                }, {
                    '$project': {
                        '_id': 0,
                        'data': 1,
                        'lokasi': '$lokasi'
                    }
                }, {
                    '$project': {
                        'data': {
                            '$slice': [
                                '$data', 1
                            ]
                        },
                        'lokasi': '$lokasi'
                    }
                }, {
                    '$unwind': {
                        'path': '$data'
                    }
                },
                {
                    '$project': {
                        'kode': '$data.kodeHasilSurvey',
                        'status': '$data.status',
                        'lokasi': '$lokasi'
                    }
                }
            ]

            # json_ret = []
            # search = hasilSurvey.objects.filter(status__status=status_,issue__0__exists=False)
            # for k,v in groupby(search,key=lambda x:x['kodeHasilSurvey'].strip()):
            #     items=[]
            #     for dt in list(v):
            #         serializer = hasilSurveySerializer(dt)
            #         result=serializer.data
            #         items.append(result)
            #     json_ret.append(items)

            pipe = pipeline + skip

            agg_cursor = hasilSurvey.objects.aggregate(*pipe)
            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Data tidak Ditemukan')
            """
            dataFrameHasil = pandas.DataFrame(agg_cursor)
            
            if len(dataFrameHasil.index)>0:
                Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
                return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
           
            else:
                return Response.badRequest(message='Status '+status_+' tidak Ditemukan')
            """
    except Exception as e:
        return Response.badRequest(message=str(e))


def getsurveyorlogai(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            #status_ = body_data.get('status')
            surveyor = body_data.get('surveyor')
            page = int(body_data.get('page', 0)) - 1
            skip = []
            if page >= 0:
                skip = [{'$skip': 20 * page},
                        {'$limit': 20}]
            pipeline = [
                {
                    '$lookup': {
                        'from': 'user_info',
                        'localField': 'user',
                        'foreignField': '_id',
                        'as': 'user'
                    }
                }, {
                    '$unwind': {
                        'path': '$user',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'surveyor',
                        'localField': 'user.organization',
                        'foreignField': '_id',
                        'as': 'user.organization'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.organization',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$match': {
                        'user.organization._id': ObjectId(surveyor)
                    }
                }, {
                    '$lookup': {
                        'from': 'userrole',
                        'localField': 'user.role',
                        'foreignField': '_id',
                        'as': 'user.role'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.role',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'document_user',
                        'localField': 'user.doc',
                        'foreignField': '_id',
                        'as': 'user.doc'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.doc',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$group': {
                        '_id': '$kodeHasilSurvey',
                        'data': {
                            '$push': '$$ROOT'
                        }
                    }
                }, {
                    '$addFields': {
                        'last': {
                            '$arrayElemAt': ['$data', -1]
                        },
                    }
                }, {
                    '$sort': {
                        'last.tanggal_pembaruan': -1
                    }
                }, {
                    '$project': {
                        '_id': 0,
                        'data': 1
                    }
                }, {
                    '$project': {
                        'data': {
                            '$slice': [
                                '$data', 1
                            ]
                        }
                    }
                }, {
                    '$unwind': {
                        'path': '$data'
                    }
                }, {
                    '$project': {
                        'kode': '$data.kodeHasilSurvey',
                        'status': '$data.status'
                    }
                }
            ]
        #     [
        #          {
        #     '$lookup': {
        #         'from': 'user_info',
        #         'localField': 'user',
        #         'foreignField': '_id',
        #         'as': 'user'
        #     }
        # }, {
        #     '$unwind': {
        #         'path': '$user',
        #         'preserveNullAndEmptyArrays': True
        #     }
        # }, {
        #     '$lookup': {
        #         'from': 'surveyor',
        #         'localField': 'user.organization',
        #         'foreignField': '_id',
        #         'as': 'user.organization'
        #     }
        # }, {
        #     '$unwind': {
        #         'path': '$user.organization',
        #         'preserveNullAndEmptyArrays': True
        #     }
        # }, {
        #     '$match': {
        #         'user.organization._id': ObjectId(surveyor)
        #     }
        # }, {
        #     '$lookup': {
        #         'from': 'userrole',
        #         'localField': 'user.role',
        #         'foreignField': '_id',
        #         'as': 'user.role'
        #     }
        # }, {
        #     '$unwind': {
        #         'path': '$user.role',
        #         'preserveNullAndEmptyArrays': True
        #     }
        # }, {
        #     '$lookup': {
        #         'from': 'document_user',
        #         'localField': 'user.doc',
        #         'foreignField': '_id',
        #         'as': 'user.doc'
        #     }
        # }, {
        #     '$unwind': {
        #         'path': '$user.doc',
        #         'preserveNullAndEmptyArrays': True
        #     }
        # }, {
        #     '$group': {
        #         '_id': '$kodeHasilSurvey',
        #         'data': {
        #             '$push': '$$ROOT'
        #         }
        #     }
        # }, {
        #     '$project': {
        #         '_id': 0,
        #         'data': 1
        #     }
        # }
        #     ]

            pipe = pipeline + skip

            agg_cursor = hasilSurvey.objects.aggregate(*pipe)

            agg_cursor = hasilSurvey.objects.aggregate(*pipe)
            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Data tidak Ditemukan')
                """
                dataFrameHasil = pandas.DataFrame(agg_cursor)
                
                if len(dataFrameHasil.index)>0:
                    Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
                    return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
                else:
                    return Response.ok(message='Tidak ada data ditemukan')
                """

                # users = list(UserInfo.objects(organization=ObjectId(surveyor)))
                # json_ret = []
                # search = hasilSurvey.objects.filter(issue__0__exists=False,user__in=users)
                # for k,v in groupby(search,key=lambda x:x['kodeHasilSurvey'].strip()):
                #     items=[]
                #     for dt in list(v):
                #         serializer = hasilSurveySerializer(dt)
                #         result=serializer.data
                #         items.append(result)
                #     json_ret.append(items)

                # if len(json_ret) > 1:
                #     #serializer = hasilSurveySerializer(json_ret,many=True)
                #     #result=serializer.data
                #     return Response.ok(values=json_ret, message=str(len(json_ret))+' Buah Data telah terambil')
                # else:
                #     return Response.badRequest(message='Data tidak Ditemukan')
    except Exception as e:
        return Response.badRequest(message=str(e))


"""
def getsurveystatusai(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            status_ = body_data.get('status')
            listJSON = []
            pipeline = [
                {'$project': {'last_status': { '$arrayElemAt': [{ '$slice': [ "$status", -1 ] }, 0 ]} }},
                {'$match': {'last_status.status':status_}}
            ]

            agg_cursor = hasilSurvey.objects.aggregate(*pipeline)

            search = [ hasilSurvey.objects.get(id=hasil_survey['_id']) for hasil_survey in agg_cursor ]

            if len(search) > 1:
                serializer = hasilSurveySerializer(search, many=True)
                result=serializer.data
                #for item in search:
                #    listJSON.append(item.serialize())
                return Response.ok(values=result, message=str(len(result))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Status '+status_+' tidak Ditemukan')
    except Exception as e:
        return Response.badRequest(message=str(e))
"""


def default(self, obj):
    if isinstance(obj, bson.ObjectId):
        return str(obj)

    return json.JSONEncoder.default(self, obj)


def getsurveystatusai(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            status_ = body_data.get('status')
            page = int(body_data.get('page', 0)) - 1
            skip = []
            if page >= 0:
                skip = [{'$skip': 20 * page},
                        {'$limit': 20}]
            listJSON = []

            pipeline = [

                {
                    '$addFields': {
                        'size_of_issue': {
                            "$size": {"$ifNull": ["$issue", []]}
                        }
                    }
                },
                {
                    '$match': {
                        'size_of_issue': 0
                    }
                },
                {
                    '$lookup': {
                        'from': 'user_info',
                        'localField': 'user',
                        'foreignField': '_id',
                        'as': 'user'
                    }
                }, {
                    '$unwind': {
                        'path': '$user',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'surveyor',
                        'localField': 'user.organization',
                        'foreignField': '_id',
                        'as': 'user.organization'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.organization',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'userrole',
                        'localField': 'user.role',
                        'foreignField': '_id',
                        'as': 'user.role'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.role',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'document_user',
                        'localField': 'user.doc',
                        'foreignField': '_id',
                        'as': 'user.doc'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.doc',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$sort': {
                        'tanggal_pembaruan': -1
                    }
                }, {
                    '$addFields': {
                        'lastStatus': {
                            '$arrayElemAt': [
                                '$status', -1
                            ]
                        }
                    }
                }, {
                    '$match': {
                        'lastStatus.status': status_
                    }
                }, {
                    '$project': {
                        'lastStatus': 0
                    }
                }, {
                    '$group': {
                        '_id': '$kodeHasilSurvey',
                        'data': {
                            '$push': '$$ROOT'
                        }
                    }
                }, {
                    '$lookup': {
                        'from': 'penugasan',
                        'localField': '_id',
                        'foreignField': 'kode',
                        'as': 'penugasan'
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'lokasi_survey',
                        'localField': 'penugasan.lokasisurvey',
                        'foreignField': '_id',
                        'as': 'lokasisurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasisurvey',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'provinsi',
                        'localField': 'lokasisurvey.provinsi',
                        'foreignField': '_id',
                        'as': 'lokasi.provinsi'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.provinsi',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kabupaten',
                        'localField': 'lokasisurvey.kabupaten',
                        'foreignField': '_id',
                        'as': 'lokasi.kabupaten'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kabupaten',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kota',
                        'localField': 'lokasisurvey.kota',
                        'foreignField': '_id',
                        'as': 'lokasi.kota'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kota',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kecamatan',
                        'localField': 'lokasisurvey.kecamatan',
                        'foreignField': '_id',
                        'as': 'lokasi.kecamatan'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kecamatan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'desa',
                        'localField': 'lokasisurvey.desa',
                        'foreignField': '_id',
                        'as': 'lokasi.desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.desa',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$addFields': {
                        'last': {
                            '$arrayElemAt': [
                                '$data', -1
                            ]
                        }
                    }
                }, {
                    '$sort': {
                        'last.tanggal_pembaruan': -1
                    }
                }, {
                    '$project': {
                        '_id': 0,
                        'kode': '$_id',
                        'lokasi': '$lokasi',
                        'data': '$data'
                    }
                }
            ]

            """
            , {
                    '$addFields': {
                        'lastStatus': {
                            '$arrayElemAt': [
                                '$status', -1
                            ]
                        }
                    }
                }, {
                    '$match': {
                        'lastStatus.status': status_
                    }
                }, {
                    '$project': {
                        'lastStatus': 0
                    }
                }
            """
            pipe = pipeline + skip
            agg_cursor = hasilSurvey.objects.aggregate(*pipe)

            #search = [x for x in agg_cursor]
            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Data tidak Ditemukan')
            """
            dataFrameHasil = pandas.DataFrame(agg_cursor)
            
            if len(dataFrameHasil.index)>0:
                Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
            # search = [ hasilSurveySerializer(hasilSurvey.objects.get(id=hasil_survey['_id'])).data for hasil_survey in agg_cursor ]
            # json_ret=[]
            # for k,v in groupby(search,key=lambda x:x['kodeHasilSurvey'].strip()):
            #     json_dict = {}
            #     #json_dict["kodeSurvey"] = k.strip()
            #     json_dict["data"] = []
            #     for dt in list(v):
            #         #serializer = hasilSurveySerializer(dt)
            #         #result=serializer.data
            #         #json_dict["data"].append(result)
            #         json_dict["data"].append(dt)
            #     #print(json_dict["kodeSurvey"],len(json_dict["data"]))
            #     json_ret.append(json_dict)
            
            # if len(json_ret) > 0:
                #serializer = hasilSurveySerializer(json_ret,many=True)
                #result=serializer.data
                #return Response.ok(values=json_ret, message=str(len(json_ret))+' Buah Data telah terambil')
            
                return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
            else:
                return Response.ok(message='Tidak ada data ditemukan')


            # else:
            #     return Response.badRequest(message='Status '+status_+' tidak Ditemukan')
            """
    except Exception as e:
        return Response.badRequest(message=str(e))


def getsurveylogbts(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            #status_ = body_data.get('status')
            page = int(body_data.get('page', 0)) - 1
            skip = []
            if page >= 0:
                skip = [{'$skip': 20 * page},
                        {'$limit': 20}]
            json_ret = []
            pipeline = [
                {
                    '$lookup': {
                        'from': 'user_info',
                        'localField': 'user',
                        'foreignField': '_id',
                        'as': 'user'
                    }
                }, {
                    '$unwind': {
                        'path': '$user',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'surveyor',
                        'localField': 'user.organization',
                        'foreignField': '_id',
                        'as': 'user.organization'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.organization',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'userrole',
                        'localField': 'user.role',
                        'foreignField': '_id',
                        'as': 'user.role'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.role',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'document_user',
                        'localField': 'user.doc',
                        'foreignField': '_id',
                        'as': 'user.doc'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.doc',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$group': {
                        '_id': '$kodeHasilSurvey',
                        'data': {
                            '$push': '$$ROOT'
                        }
                    }
                }, {
                    '$lookup': {
                        'from': 'penugasan',
                        'localField': '_id',
                        'foreignField': 'kode',
                        'as': 'penugasan'
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'lokasi_survey',
                        'localField': 'penugasan.lokasisurvey',
                        'foreignField': '_id',
                        'as': 'lokasisurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasisurvey',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'provinsi',
                        'localField': 'lokasisurvey.provinsi',
                        'foreignField': '_id',
                        'as': 'lokasi.provinsi'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.provinsi',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kabupaten',
                        'localField': 'lokasisurvey.kabupaten',
                        'foreignField': '_id',
                        'as': 'lokasi.kabupaten'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kabupaten',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kota',
                        'localField': 'lokasisurvey.kota',
                        'foreignField': '_id',
                        'as': 'lokasi.kota'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kota',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kecamatan',
                        'localField': 'lokasisurvey.kecamatan',
                        'foreignField': '_id',
                        'as': 'lokasi.kecamatan'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kecamatan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'desa',
                        'localField': 'lokasisurvey.desa',
                        'foreignField': '_id',
                        'as': 'lokasi.desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.desa',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$addFields': {
                        'last': {
                            '$arrayElemAt': ['$data', -1]
                        },
                    }
                }, {
                    '$sort': {
                        'last.tanggal_pembaruan': -1
                    }
                }, {
                    '$project': {
                        '_id': 0,
                        'data': 1,
                        'lokasi': '$lokasi'
                    }
                }, {
                    '$project': {
                        'data': {
                            '$slice': [
                                '$data', 1
                            ]
                        },
                        'lokasi': '$lokasi'
                    }
                }, {
                    '$unwind': {
                        'path': '$data'
                    }
                }, {
                    '$project': {
                        'kode': '$data.kodeHasilSurvey',
                        'status': '$data.status',
                        'lokasi': '$lokasi'
                    }
                }
            ]
            # search = hasilSurveybts.objects.filter(status__status=status_,issue__0__exists=False)
            # for k,v in groupby(search,key=lambda x:x['kodeHasilSurvey'].strip()):
            #     items=[]
            #     for dt in list(v):
            #         serializer = btsSerializer(dt)
            #         result=serializer.data
            #         items.append(result)
            #     json_ret.append(items)
            pipe = pipeline + skip
            agg_cursor = hasilSurveybts.objects.aggregate(*pipe)

            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Data tidak Ditemukan')
            """
            dataFrameHasil = pandas.DataFrame(agg_cursor)
            
            if len(dataFrameHasil.index)>0:
                Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
                return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
            else:
                return Response.ok(message='Tidak ada data ditemukan')
            """
            # if len(json_ret) > 1:
            #     #serializer = hasilSurveySerializer(json_ret,many=True)
            #     #result=serializer.data
            #     return Response.ok(values=json_ret, message=str(len(json_ret))+' Buah Data telah terambil')
            # else:
            #     return Response.badRequest(message='Status '+status_+' tidak Ditemukan')

    except Exception as e:
        return Response.badRequest(message=str(e))


"""      
def getsurveylogbts(request):
    try:
        if request.method == "POST":
            
            body_data = request.POST.dict()
            status_ = body_data.get('status')
            listJSON = []
            # return Response.ok(message='Done')
            search = hasilSurveybts.objects.filter(status__status=status_,issue__0__exists=False)
            if len(search) > 1:
                serializer = btsSerializer(search, many=True)
                result=serializer.data
                #for item in search:
                #    listJSON.append(item.serialize())
                return Response.ok(values=result, message=str(len(result))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Status '+status_+' tidak Ditemukan')
    except Exception as e:
        return Response.badRequest(message=str(e))
"""


def getsurveyorlogbts(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            #status_ = body_data.get('status')
            surveyor = body_data.get('surveyor')
            page = int(body_data.get('page', 0)) - 1
            skip = []
            if page >= 0:
                skip = [{'$skip': 20 * page},
                        {'$limit': 20}]
            # users = list(UserInfo.objects(organization=ObjectId(surveyor)))
            # json_ret = []
            # search = hasilSurveybts.objects.filter(issue__0__exists=False,user__in=users)
            pipeline = [
                {
                    '$lookup': {
                        'from': 'user_info',
                        'localField': 'user',
                        'foreignField': '_id',
                        'as': 'user'
                    }
                }, {
                    '$unwind': {
                        'path': '$user',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'surveyor',
                        'localField': 'user.organization',
                        'foreignField': '_id',
                        'as': 'user.organization'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.organization',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$match': {
                        'user.organization.name': surveyor
                    }
                }, {
                    '$lookup': {
                        'from': 'userrole',
                        'localField': 'user.role',
                        'foreignField': '_id',
                        'as': 'user.role'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.role',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'document_user',
                        'localField': 'user.doc',
                        'foreignField': '_id',
                        'as': 'user.doc'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.doc',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$group': {
                        '_id': '$kodeHasilSurvey',
                        'data': {
                            '$push': '$$ROOT'
                        }
                    }
                }, {
                    '$addFields': {
                        'last': {
                            '$arrayElemAt': ['$data', -1]
                        },
                    }
                }, {
                    '$sort': {
                        'last.tanggal_pembaruan': -1
                    }
                }, {
                    '$project': {
                        '_id': 0,
                        'data': 1
                    }
                }
            ]

            pipe = pipeline + skip

            agg_cursor = hasilSurveybts.objects.aggregate(*pipe)
            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Data tidak Ditemukan')
            """    
            dataFrameHasil = pandas.DataFrame(agg_cursor)
            
            if len(dataFrameHasil.index)>0:
                Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
                return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
            else:
                return Response.ok(message='Tidak ada data ditemukan')


            # for k,v in groupby(search,key=lambda x:x['kodeHasilSurvey'].strip()):
            #     items=[]
            #     for dt in list(v):
            #         serializer = hasilSurveySerializer(dt)
            #         result=serializer.data
            #         items.append(result)
            #     json_ret.append(items)
            
            # if len(json_ret) > 1:
            #     #serializer = hasilSurveySerializer(json_ret,many=True)
            #     #result=serializer.data
            #     return Response.ok(values=json_ret, message=str(len(json_ret))+' Buah Data telah terambil')
            # else:
            #     return Response.badRequest(message='Data tidak Ditemukan')
            """
    except Exception as e:
        return Response.badRequest(message=str(e))


"""
def getsurveystatusbts(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            status_ = body_data.get('status')
            listJSON = []
            pipeline = [
                {'$project': {'last_status': { '$arrayElemAt': [{ '$slice': [ "$status", -1 ] }, 0 ]} }},
                {'$match': {'last_status.status':status_}}
            ]

            agg_cursor = hasilSurveybts.objects.aggregate(*pipeline)

            search = [ hasilSurveybts.objects.get(id=hasil_survey['_id']) for hasil_survey in agg_cursor ]

            if len(search) > 1:
                serializer = hasilSurveySerializer(search, many=True)
                result=serializer.data
                #for item in search:
                #    listJSON.append(item.serialize())
                return Response.ok(values=result, message=str(len(result))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Status '+status_+' tidak Ditemukan')
    except Exception as e:
        return Response.badRequest(message=str(e))
"""


def getsurveystatusbts(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            status_ = body_data.get('status')
            page = int(body_data.get('page', 0)) - 1
            skip = []
            if page >= 0:
                skip = [{'$skip': 20 * page},
                        {'$limit': 20}]
            listJSON = []
            pipeline = [
                # {'$project': {'last_status': { '$arrayElemAt': [{ '$slice': [ "$status", -1 ] }, 0 ]} }},
                # {'$match': {'last_status.status':status_}}
                # {
                #     '$match': {
                #         'issue': {
                #             '$exists': False
                #         }
                #     }
                # }
                {
                    '$addFields': {
                        'size_of_issue': {
                            "$size": {"$ifNull": ["$issue", []]}
                        }
                    }
                },
                {
                    '$match': {
                        'size_of_issue': 0
                    }
                },
                {
                    '$lookup': {
                        'from': 'user_info',
                        'localField': 'user',
                        'foreignField': '_id',
                        'as': 'user'
                    }
                }, {
                    '$unwind': {
                        'path': '$user',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'surveyor',
                        'localField': 'user.organization',
                        'foreignField': '_id',
                        'as': 'user.organization'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.organization',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'userrole',
                        'localField': 'user.role',
                        'foreignField': '_id',
                        'as': 'user.role'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.role',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'document_user',
                        'localField': 'user.doc',
                        'foreignField': '_id',
                        'as': 'user.doc'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.doc',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$sort': {
                        'tanggal_pembaruan': -1
                    }
                }, {
                    '$addFields': {
                        'lastStatus': {
                            '$arrayElemAt': [
                                '$status', -1
                            ]
                        }
                    }
                }, {
                    '$match': {
                        'lastStatus.status': status_
                    }
                }, {
                    '$project': {
                        'lastStatus': 0
                    }
                },
                {
                    '$group': {
                        '_id': '$kodeHasilSurvey',
                        'data': {
                            '$push': '$$ROOT'
                        }
                    }
                }, {
                    '$lookup': {
                        'from': 'penugasan',
                        'localField': '_id',
                        'foreignField': 'kode',
                        'as': 'penugasan'
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'lokasi_survey',
                        'localField': 'penugasan.lokasisurvey',
                        'foreignField': '_id',
                        'as': 'lokasisurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasisurvey',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'provinsi',
                        'localField': 'lokasisurvey.provinsi',
                        'foreignField': '_id',
                        'as': 'lokasi.provinsi'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.provinsi',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kabupaten',
                        'localField': 'lokasisurvey.kabupaten',
                        'foreignField': '_id',
                        'as': 'lokasi.kabupaten'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kabupaten',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kota',
                        'localField': 'lokasisurvey.kota',
                        'foreignField': '_id',
                        'as': 'lokasi.kota'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kota',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kecamatan',
                        'localField': 'lokasisurvey.kecamatan',
                        'foreignField': '_id',
                        'as': 'lokasi.kecamatan'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kecamatan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'desa',
                        'localField': 'lokasisurvey.desa',
                        'foreignField': '_id',
                        'as': 'lokasi.desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.desa',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$addFields': {
                        'last': {
                            '$arrayElemAt': [
                                '$data', -1
                            ]
                        }
                    }
                }, {
                    '$sort': {
                        'last.tanggal_pembaruan': -1
                    }
                }, {
                    '$project': {
                        '_id': 0,
                        'kode': '$_id',
                        'lokasi': '$lokasi',
                        'data': '$data'
                    }
                }

            ]
            pipe = pipeline + skip
            agg_cursor = hasilSurveybts.objects.aggregate(*pipe)
            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Data tidak Ditemukan')
            """
            dataFrameHasil = pandas.DataFrame(agg_cursor)
            
            if len(dataFrameHasil.index)>0:
                Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
            
            # search = [ hasilSurveybts.objects.get(id=hasil_survey['_id']) for hasil_survey in agg_cursor ]

            # json_ret=[]
            # for k,v in groupby(search,key=lambda x:x['kodeHasilSurvey'].strip()):
            #     json_dict = {}
            #     #json_dict["kodeSurvey"] = k.strip()
            #     json_dict["data"] = []
            #     for dt in list(v):
            #         serializer = btsSerializer(dt,many=False).data
            #         #result=serializer.data
            #         json_dict["data"].append(serializer)
            #     #print(json_dict["kodeSurvey"],len(json_dict["data"]))
            #     json_ret.append(json_dict)

                return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Status '+status_+' tidak Ditemukan')
            """
    except Exception as e:
        return Response.badRequest(message=str(e))


def getsurveybydateai(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            from_ = datetime.strptime(
                body_data.get('from'), '%Y-%m-%d %H:%M:%S')
            to_ = datetime.strptime(body_data.get('to'), '%Y-%m-%d %H:%M:%S')
            if to_ < from_:
                return Response.badRequest(
                    values='null',
                    message='Tanggal akhir harus lebih besar dari tanggal awal'
                )
            jumlah_hari = (to_ - from_).days

            pipeline = [
                {
                    '$addFields': {
                        'lastStatus': {
                            '$arrayElemAt': [
                                '$status', 0
                            ]
                        }
                    }
                }, {
                    '$match': {
                        'lastStatus.tanggal_pembuatan': {
                            '$gte': from_,
                            '$lte': to_
                        }
                    }
                }, {
                    '$addFields': {
                        'date': {
                            '$dateToString': {
                                'format': '%Y-%m-%d',
                                'date': '$lastStatus.tanggal_pembuatan'
                            }
                        }
                    }
                }, {
                    '$group': {
                        '_id': {
                            'kode': '$kodeHasilSurvey',
                            'date': '$date'
                        },
                        'data': {
                            '$push': '$$ROOT'
                        }
                    }
                }, {
                    '$project': {
                        'kode': '$_id.kode',
                        'date': '$_id.date',
                        '_id': False
                    }
                }, {
                    '$group': {
                        '_id': '$date',
                        'count': {
                            '$sum': 1
                        }
                    }
                }, {
                    '$sort': {
                        '_id': 1
                    }
                }, {
                    '$project': {
                        'tanggal': '$_id',
                        '_id': False,
                        'nilai': '$count'
                    }
                }
            ]
            agg_cursor = hasilSurvey.objects.aggregate(*pipeline)
            search = list(agg_cursor)
            json_list = []
            ada = False
            for x in range(jumlah_hari+1):
                current = (from_.date() + timedelta(x)).strftime('%Y-%m-%d')
                for crr in search:
                    if crr['tanggal'] == current:
                        json_dict = {}
                        json_dict["tanggal"] = crr['tanggal']
                        json_dict["nilai"] = crr['nilai']
                        json_list.append(json_dict)
                        ada = True
                        break
                if not ada:
                    json_dict = {}
                    json_dict["tanggal"] = current
                    json_dict["nilai"] = 0
                    json_list.append(json_dict)
                else:
                    ada = False

            # if len(search) > 0:
            if len(json_list) > 0:
                # return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
                return Response.ok(values=json_list, message=str(len(json_list))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Data tidak Ditemukan')
            """
            dataFrameHasil = pandas.DataFrame(agg_cursor)
            
            if len(dataFrameHasil.index)>0:
                Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
            
                return Response.ok(values=json.loads(Hasil))
            else:
                return Response.ok(message='data tidak ditemukan')
            # #search = hasilSurvey.objects.filter(tanggal_pembuatan__gte=from_, tanggal_pembuatan__lte=to_).order_by("tanggal_pembuatan")
            # search = hasilSurvey.objects.filter(status__status="Submitted", status__0__tanggal_pembuatan__gte=from_,status__0__tanggal_pembuatan__lte=to_).order_by("tanggal_pembuatan")
            # if len(search) == 0:
            #     return Response.badRequest(message='Data tidak Ditemukan')
            
            # json_list = []
            # json_ret = []
            # count = 0
            # for i in range(0, len(search)):
            #     json_dict = {}
            #     json_dict["tanggal"] = str(search[i]['status'][0]['tanggal_pembuatan'])[:10]
            #     json_dict["nilai"] = 1
            #     json_list.append(json_dict)
            
            # for k,v in groupby(json_list,key=lambda x:x['tanggal']):
            #     json_dict = {}
            #     json_dict["tanggal"] = k
            #     json_dict["nilai"] = len(list(v))
            #     json_ret.append(json_dict)
            
            # return Response.ok(values=json_ret, message=str(len(json_ret))+' Buah Data telah terambil')
            """
    except Exception as e:
        return Response.badRequest(message=str(e))


def getsurveybydatebts(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            from_ = datetime.strptime(
                body_data.get('from'), '%Y-%m-%d %H:%M:%S')
            to_ = datetime.strptime(body_data.get('to'), '%Y-%m-%d %H:%M:%S')
            if to_ < from_:
                return Response.badRequest(
                    values='null',
                    message='Tanggal akhir harus lebih besar dari tanggal awal'
                )
            jumlah_hari = (to_ - from_).days
            #search = hasilSurveybts.objects.filter(tanggal_pembuatan__gte=from_, tanggal_pembuatan__lte=to_).order_by("tanggal_pembuatan")
            search = hasilSurveybts.objects.filter(status__status="Submitted", status__0__tanggal_pembuatan__gte=from_,
                                                   status__0__tanggal_pembuatan__lte=to_, nomorSurvey='1').order_by("tanggal_pembuatan")
            """
            if len(search) == 0:
                json_list11 = []
                json_dict = {}
                json_dict["tanggal"] = from_.date()
                json_dict["nilai"] = 0
                json_list11.append(json_dict)
                # return Response.badRequest(message='Data tidak Ditemukan')
                return Response.ok(values=json_list11, message=str(len(json_list11))+' Buah Data telah terambil')
            """
            json_list = []
            json_ret = []
            count = 0
            for i in range(0, len(search)):
                json_dict = {}
                json_dict["tanggal"] = str(
                    search[i]['status'][0]['tanggal_pembuatan'])[:10]
                json_dict["nilai"] = 1
                json_list.append(json_dict)

            for k, v in groupby(json_list, key=lambda x: x['tanggal']):
                json_dict = {}
                json_dict["tanggal"] = k
                json_dict["nilai"] = len(list(v))
                json_ret.append(json_dict)

            json_list1 = []
            ada = False
            for x in range(jumlah_hari+1):
                current = (from_.date() + timedelta(x)).strftime('%Y-%m-%d')
                for crr in json_ret:
                    if crr['tanggal'] == current:
                        json_dict = {}
                        json_dict["tanggal"] = crr['tanggal']
                        json_dict["nilai"] = crr['nilai']
                        json_list1.append(json_dict)
                        ada = True
                        break
                if not ada:
                    json_dict = {}
                    json_dict["tanggal"] = current
                    json_dict["nilai"] = 0
                    json_list1.append(json_dict)
                else:
                    ada = False

            # return Response.ok(values=json_ret, message=str(len(json_ret))+' Buah Data telah terambil')
            return Response.ok(values=json_list1, message=str(len(json_list1))+' Buah Data telah terambil')

    except Exception as e:
        return Response.badRequest(message=str(e))


def PostHasilSurveyBTS(request):
    try:
        body_data = request.POST.dict()
        if request.method == "POST":
            kode = body_data.get("kodeHasilSurvey")
            nomorSurv = body_data.get('nomorSurvey')
            searchKode = hasilSurveyBts.objects.filter(
                kodeHasilSurvey=kode, nomorSurvey=nomorSurv)
            if len(searchKode) > 0:  # Duplicate Kode
                return Response.badRequest(
                    values='null',
                    message='Hasil Survey sudah terisi di database'
                )
            try:
                fs = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/survey/foto/',
                    base_url=f'{settings.MEDIA_URL}/survey/foto/'
                )
            except Exception as e:
                return Response.badRequest(message=str(e))
            section1Data = Section1(
                tipeKawasan=body_data.get("tipeKawasan"),
                alamatLokasi=body_data.get("alamatLokasi"),
                modaTransportasi=ModaTransportasi(
                    darat=body_data.get("darat"),
                    laut=body_data.get("laut"),
                    udara=body_data.get("udara"),
                    durasiPerjalanan=body_data.get("durasiPerjalanan"),
                    namaKotaKecamatan=body_data.get("namaKotaKecamatan")
                ),
                tipeAntena=body_data.get("tipeAntena")
            )
            section2Data = Section2(
                latitude=body_data.get("latitude"),
                longitude=body_data.get("longitude"),
                ketinggianAsl=body_data.get("ketinggianAsl")
            )
            section3Data = Section3(
                posisiTower=body_data.get("posisiTower"),
                jarakPemukiman=body_data.get("jarakPemukiman"),
                kepemilikanLahan=body_data.get(
                    "kepemilikanLahan") if True else None,
                statusKondisiLahan=body_data.get("statusKondisiLahan"),
                kondisiSosial=body_data.get("kondisiSosial"),
                keamanan=body_data.get("keamanan"),
                ukuranLahan=body_data.get("ukuranLahan"),
            )
            section4Data = Section4(
                coverageRadius=body_data.get("coverageRadius"),
                levelSinyal=body_data.get("levelSinyal"),
                dbm=body_data.get("dbm"),
                callSite=body_data.get("callSite"),
                smsSite=body_data.get("smsSite"),
                namaOperator=body_data.get("namaOperator"),
            )
            section5Data = Section5(
                topografiUmum=body_data.get("topografiUmum"),
                klasifikasiLahan=body_data.get("klasifikasiLahan"),
                objekPenghalang=body_data.get("objekPenghalang"),
                rekondisiLahan=body_data.get("rekondisiLahan"),
                tipeTanah=body_data.get("tipeTanah"),
                jarakSungaiLaut=body_data.get("jarakSungaiLaut"),
            )

            if body_data.get("merk") == "":
                merkData = None
                kapasitasData = None
                kapasitasBbmData = None
            else:
                merkData = body_data.get("merk")
                kapasitasData = body_data.get("kapasitas")
                kapasitasBbmData = body_data.get("kapasitasBbm")

            section6Data = Section6(
                sumberListrik=body_data.get("sumberListrik"),
                phaseListrik=body_data.get("phaseListrik"),
                dayaListrik=body_data.get("dayaListrik"),
                jamOperasionalListrik=body_data.get("jamOperasionalListrik"),
                jarakSumberListrik=body_data.get("jarakSumberListrik"),
                generatorBackup=body_data.get("generatorBackup"),
                merk=merkData,
                kapasitas=kapasitasData,
                kapasitasBbm=kapasitasBbmData,
                pasokanBbm=body_data.get("pasokanBbm"),
                jenisBbm=body_data.get("jenisBbm"),
                harga=body_data.get("harga"),
                listrikIdeal=body_data.get("listrikIdeal"),
            )
            section7Data = Section7(
                suratTanah=body_data.get("suratTanah"),
                izinDiperlukan=body_data.get("izinDiperlukan"),
            )
            if body_data.get("namaDesa") == '':
                namaDesaData = None
            else:
                namaDesaData = body_data.get("namaDesa")

            section8Data = Section8(
                populasi=body_data.get("populasi"),
                kepadatanPenduduk=body_data.get("kepadatanPenduduk"),
                sebaranPenduduk=body_data.get("sebaranPenduduk"),
                desaTerdekat=body_data.get("desaTerdekat"),
                namaDesa=namaDesaData,
                jarakDesaTerdekat=body_data.get("jarakDesaTerdekat"),
                mataPencaharian=body_data.get("mataPencaharian"),
                penggunaHp=body_data.get("penggunaHp"),
                tipeHp=body_data.get("tipeHp"),
                providerSimCard=body_data.get("providerSimCard"),
                wargaBimtek=body_data.get("wargaBimtek"),
                aksesInternet=body_data.get("aksesInternet"),
                rumahDenganGenset=body_data.get("rumahDenganGenset"),
            )

            fileLahan = request.FILES['fotoLahan']
            filenameLahan = fs.save(fileLahan.name, fileLahan)
            file_pathLahan = fs.url(filenameLahan)

            section9Data = Section9(
                fotoLahan=patternFoto(
                    nama=filenameLahan, url=file_pathLahan),
            )

            fileDetailLahan = request.FILES['fotoDetailLahan']
            filenameDetailLahan = fs.save(
                fileDetailLahan.name, fileDetailLahan)
            file_pathDetailLahan = fs.url(filenameDetailLahan)

            fileDetailMarking = request.FILES['fotoDetailMarking']
            filenameDetailMarking = fs.save(
                fileDetailMarking.name, fileDetailMarking)
            file_pathDetailMarking = fs.url(filenameDetailMarking)

            fileDetailSisiUtara = request.FILES['fotoDetailSisiUtara']
            filenameDetailSisiUtara = fs.save(
                fileDetailSisiUtara.name, fileDetailSisiUtara)
            file_pathDetailSisiUtara = fs.url(filenameDetailSisiUtara)

            fileDetailSisiTimur = request.FILES['fotoDetailSisiTimur']
            filenameDetailSisiTimur = fs.save(
                fileDetailSisiTimur.name, fileDetailSisiTimur)
            file_pathDetailSisiTimur = fs.url(filenameDetailSisiTimur)

            fileDetailSisiBarat = request.FILES['fotoDetailSisiBarat']
            filenameDetailSisiBarat = fs.save(
                fileDetailSisiBarat.name, fileDetailSisiBarat)
            file_pathDetailSisiBarat = fs.url(filenameDetailSisiBarat)

            fileDetailSisiSelatan = request.FILES['fotoDetailSisiSelatan']
            filenameDetailSisiSelatan = fs.save(
                fileDetailSisiSelatan.name, fileDetailSisiSelatan)
            file_pathDetailSisiSelatan = fs.url(filenameDetailSisiSelatan)

            section10Data = Section10(
                fotoDetailLahan=patternFoto(
                    nama=filenameDetailLahan, url=file_pathDetailLahan),
                fotoDetailMarking=patternFoto(
                    nama=filenameDetailMarking, url=file_pathDetailMarking),
                fotoDetailSisiUtara=patternFoto(
                    nama=filenameDetailSisiUtara, url=file_pathDetailSisiTimur),
                fotoDetailSisiTimur=patternFoto(
                    nama=filenameDetailSisiTimur, url=file_pathDetailSisiTimur),
                fotoDetailSisiBarat=patternFoto(
                    nama=filenameDetailSisiBarat, url=file_pathDetailSisiBarat),
                fotoDetailSisiSelatan=patternFoto(
                    nama=filenameDetailSisiSelatan, url=file_pathDetailSisiSelatan),

            )
            section11Data = Section11(
                coverage0N=body_data.get("coverage0N"),
                coverage0Ndesc=body_data.get("coverage0Ndesc"),
                coverage45N=body_data.get("coverage45N"),
                coverage45Ndesc=body_data.get("coverage45Ndesc"),
                coverage90N=body_data.get("coverage90N"),
                coverage90Ndesc=body_data.get("coverage90Ndesc"),
                coverage135N=body_data.get("coverage135N"),
                coverage135Ndesc=body_data.get("coverage135Ndesc"),
                coverage180N=body_data.get("coverage180N"),
                coverage180Ndesc=body_data.get("coverage180Ndesc"),
                coverage225N=body_data.get("coverage225N"),
                coverage225Ndesc=body_data.get("coverage225Ndesc"),
                coverage270N=body_data.get("coverage270N"),
                coverage270Ndesc=body_data.get("coverage270Ndesc"),
                coverage315N=body_data.get("coverage315N"),
                coverage315Ndesc=body_data.get("coverage315Ndesc"),
            )
            fileGnetTrack0N2Km = request.FILES['fotoGnetTrack0N2Km']
            filenameGnetTrack0N2Km = fs.save(
                fileGnetTrack0N2Km.name, fileGnetTrack0N2Km)
            file_pathGnetTrack0N2Km = fs.url(filenameGnetTrack0N2Km)

            fileGnetTrack45N2Km = request.FILES['fotoGnetTrack45N2Km']
            filenameGnetTrack45N2Km = fs.save(
                fileGnetTrack45N2Km.name, fileGnetTrack45N2Km)
            file_pathGnetTrack45N2Km = fs.url(filenameGnetTrack45N2Km)

            fileGnetTrack90N2Km = request.FILES['fotoGnetTrack90N2Km']
            filenameGnetTrack90N2Km = fs.save(
                fileGnetTrack90N2Km.name, fileGnetTrack90N2Km)
            file_pathGnetTrack90N2Km = fs.url(filenameGnetTrack90N2Km)

            fileGnetTrack135N2Km = request.FILES['fotoGnetTrack135N2Km']
            filenameGnetTrack135N2Km = fs.save(
                fileGnetTrack135N2Km.name, fileGnetTrack135N2Km)
            file_pathGnetTrack135N2Km = fs.url(filenameGnetTrack135N2Km)

            fileGnetTrack180N2Km = request.FILES['fotoGnetTrack180N2Km']
            filenameGnetTrack180N2Km = fs.save(
                fileGnetTrack180N2Km.name, fileGnetTrack180N2Km)
            file_pathGnetTrack180N2Km = fs.url(filenameGnetTrack180N2Km)

            fileGnetTrack225N2Km = request.FILES['fotoGnetTrack225N2Km']
            filenameGnetTrack225N2Km = fs.save(
                fileGnetTrack225N2Km.name, fileGnetTrack225N2Km)
            file_pathGnetTrack225N2Km = fs.url(filenameGnetTrack225N2Km)

            fileGnetTrack270N2Km = request.FILES['fotoGnetTrack270N2Km']
            filenameGnetTrack270N2Km = fs.save(
                fileGnetTrack270N2Km.name, fileGnetTrack270N2Km)
            file_pathGnetTrack270N2Km = fs.url(filenameGnetTrack270N2Km)

            fileGnetTrack315N2Km = request.FILES['fotoGnetTrack315N2Km']
            filenameGnetTrack315N2Km = fs.save(
                fileGnetTrack315N2Km.name, fileGnetTrack315N2Km)
            file_pathGnetTrack315N2Km = fs.url(filenameGnetTrack315N2Km)

            section12Data = Section12(
                fotoGnetTrack0N2Km=patternFoto(
                    nama=filenameGnetTrack0N2Km, url=file_pathGnetTrack0N2Km),
                fotoGnetTrack45N2Km=patternFoto(
                    nama=filenameGnetTrack45N2Km, url=file_pathGnetTrack45N2Km),
                fotoGnetTrack90N2Km=patternFoto(
                    nama=filenameGnetTrack90N2Km, url=file_pathGnetTrack90N2Km),
                fotoGnetTrack135N2Km=patternFoto(
                    nama=filenameGnetTrack135N2Km, url=file_pathGnetTrack135N2Km),
                fotoGnetTrack180N2Km=patternFoto(
                    nama=filenameGnetTrack180N2Km, url=file_pathGnetTrack180N2Km),
                fotoGnetTrack225N2Km=patternFoto(
                    nama=filenameGnetTrack225N2Km, url=file_pathGnetTrack225N2Km),
                fotoGnetTrack270N2Km=patternFoto(
                    nama=filenameGnetTrack270N2Km, url=file_pathGnetTrack270N2Km),
                fotoGnetTrack315N2Km=patternFoto(
                    nama=filenameGnetTrack315N2Km, url=file_pathGnetTrack315N2Km),
            )
            fileGnetTrack0N5Km = request.FILES['fotoGnetTrack0N5Km']
            filenameGnetTrack0N5Km = fs.save(
                fileGnetTrack0N5Km.name, fileGnetTrack0N5Km)
            file_pathGnetTrack0N5Km = fs.url(filenameGnetTrack0N5Km)

            fileGnetTrack45N5Km = request.FILES['fotoGnetTrack45N5Km']
            filenameGnetTrack45N5Km = fs.save(
                fileGnetTrack45N5Km.name, fileGnetTrack45N5Km)
            file_pathGnetTrack45N5Km = fs.url(filenameGnetTrack45N5Km)

            fileGnetTrack90N5Km = request.FILES['fotoGnetTrack90N5Km']
            filenameGnetTrack90N5Km = fs.save(
                fileGnetTrack90N5Km.name, fileGnetTrack90N5Km)
            file_pathGnetTrack90N5Km = fs.url(filenameGnetTrack90N5Km)

            fileGnetTrack135N5Km = request.FILES['fotoGnetTrack135N5Km']
            filenameGnetTrack135N5Km = fs.save(
                fileGnetTrack135N5Km.name, fileGnetTrack135N5Km)
            file_pathGnetTrack135N5Km = fs.url(filenameGnetTrack135N5Km)

            fileGnetTrack180N5Km = request.FILES['fotoGnetTrack180N5Km']
            filenameGnetTrack180N5Km = fs.save(
                fileGnetTrack180N5Km.name, fileGnetTrack180N5Km)
            file_pathGnetTrack180N5Km = fs.url(filenameGnetTrack180N5Km)

            fileGnetTrack225N5Km = request.FILES['fotoGnetTrack225N5Km']
            filenameGnetTrack225N5Km = fs.save(
                fileGnetTrack225N5Km.name, fileGnetTrack225N5Km)
            file_pathGnetTrack225N5Km = fs.url(filenameGnetTrack225N5Km)

            fileGnetTrack270N5Km = request.FILES['fotoGnetTrack270N5Km']
            filenameGnetTrack270N5Km = fs.save(
                fileGnetTrack270N5Km.name, fileGnetTrack270N5Km)
            file_pathGnetTrack270N5Km = fs.url(filenameGnetTrack270N5Km)

            fileGnetTrack315N5Km = request.FILES['fotoGnetTrack315N5Km']
            filenameGnetTrack315N5Km = fs.save(
                fileGnetTrack315N5Km.name, fileGnetTrack315N5Km)
            file_pathGnetTrack315N5Km = fs.url(filenameGnetTrack315N5Km)

            section13Data = Section13(
                fotoGnetTrack0N5Km=patternFoto(
                    nama=filenameGnetTrack0N5Km, url=file_pathGnetTrack0N5Km),
                fotoGnetTrack45N5Km=patternFoto(
                    nama=filenameGnetTrack45N5Km, url=file_pathGnetTrack45N5Km),
                fotoGnetTrack90N5Km=patternFoto(
                    nama=filenameGnetTrack90N5Km, url=file_pathGnetTrack90N5Km),
                fotoGnetTrack135N5Km=patternFoto(
                    nama=filenameGnetTrack135N5Km, url=file_pathGnetTrack135N5Km),
                fotoGnetTrack180N5Km=patternFoto(
                    nama=filenameGnetTrack180N5Km, url=file_pathGnetTrack180N5Km),
                fotoGnetTrack225N5Km=patternFoto(
                    nama=filenameGnetTrack225N5Km, url=file_pathGnetTrack225N5Km),
                fotoGnetTrack270N5Km=patternFoto(
                    nama=filenameGnetTrack270N5Km, url=file_pathGnetTrack270N5Km),
                fotoGnetTrack315N5Km=patternFoto(
                    nama=filenameGnetTrack315N5Km, url=file_pathGnetTrack315N5Km),

            )

            if body_data.get("latitudeAlt1") == "":
                latitudeAlt1Data = None
                longitudeAlt1Data = None
                elevasiAlt1Data = None
            else:
                latitudeAlt1Data = body_data.get("latitudeAlt1")
                longitudeAlt1Data = body_data.get("longitudeAlt1")
                elevasiAlt1Data = body_data.get("elevasiAlt1")

            if body_data.get("latitudeAlt2") == "":
                latitudeAlt2Data = None
                longitudeAlt2Data = None
                elevasiAlt2Data = None
            else:
                latitudeAlt2Data = body_data.get("latitudeAlt2")
                longitudeAlt2Data = body_data.get("longitudeAlt2")
                elevasiAlt2Data = body_data.get("elevasiAlt2")

            section14Data = Section14(
                latitudeMapping=body_data.get("latitudeMapping"),
                longitudeMapping=body_data.get("longitudeMapping"),
                elevasiMapping=body_data.get("elevasiMapping"),
                latitudeAlt1=latitudeAlt1Data,
                longitudeAlt1=longitudeAlt1Data,
                elevasiAlt1=elevasiAlt1Data,
                latitudeAlt2=latitudeAlt2Data,
                longitudeAlt2=longitudeAlt2Data,
                elevasiAlt2=elevasiAlt2Data,
            )

            section15Data = Section15(
                topografiSektor0N=body_data.get("topografiSektor0N"),
                landscapeSektor0N=body_data.get("landscapeSektor0N"),
                demografiSektor0N=body_data.get("demografiSektor0N"),
                topografiSektor45N=body_data.get("topografiSektor45N"),
                landscapeSektor45N=body_data.get("landscapeSektor45N"),
                demografiSektor45N=body_data.get("demografiSektor45N"),
                topografiSektor90N=body_data.get("topografiSektor90N"),
                landscapeSektor90N=body_data.get("landscapeSektor90N"),
                demografiSektor90N=body_data.get("demografiSektor90N"),
                topografiSektor135N=body_data.get("topografiSektor135N"),
                landscapeSektor135N=body_data.get("landscapeSektor135N"),
                demografiSektor135N=body_data.get("demografiSektor135N"),
                topografiSektor180N=body_data.get("topografiSektor180N"),
                landscapeSektor180N=body_data.get("landscapeSektor180N"),
                demografiSektor180N=body_data.get("demografiSektor180N"),
                topografiSektor225N=body_data.get("topografiSektor225N"),
                landscapeSektor225N=body_data.get("landscapeSektor225N"),
                demografiSektor225N=body_data.get("demografiSektor225N"),
                topografiSektor270N=body_data.get("topografiSektor270N"),
                landscapeSektor270N=body_data.get("landscapeSektor270N"),
                demografiSektor270N=body_data.get("demografiSektor270N"),
                topografiSektor315N=body_data.get("topografiSektor315N"),
                landscapeSektor315N=body_data.get("landscapeSektor315N"),
                demografiSektor315N=body_data.get("demografiSektor315N"),
            )
            fileSektor0N = request.FILES['fotoSektor0N']
            filenameSektor0N = fs.save(fileSektor0N.name, fileSektor0N)
            file_pathSektor0N = fs.url(filenameSektor0N)

            fileSektor45N = request.FILES['fotoSektor45N']
            filenameSektor45N = fs.save(fileSektor45N.name, fileSektor45N)
            file_pathSektor45N = fs.url(filenameSektor45N)

            fileSektor90N = request.FILES['fotoSektor90N']
            filenameSektor90N = fs.save(fileSektor90N.name, fileSektor90N)
            file_pathSektor90N = fs.url(filenameSektor90N)

            fileSektor135N = request.FILES['fotoSektor135N']
            filenameSektor135N = fs.save(fileSektor135N.name, fileSektor135N)
            file_pathSektor135N = fs.url(filenameSektor135N)

            fileSektor180N = request.FILES['fotoSektor180N']
            filenameSektor180N = fs.save(fileSektor180N.name, fileSektor180N)
            file_pathSektor180N = fs.url(filenameSektor180N)

            fileSektor225N = request.FILES['fotoSektor225N']
            filenameSektor225N = fs.save(fileSektor225N.name, fileSektor225N)
            file_pathSektor225N = fs.url(filenameSektor225N)

            fileSektor270N = request.FILES['fotoSektor270N']
            filenameSektor270N = fs.save(fileSektor270N.name, fileSektor270N)
            file_pathSektor270N = fs.url(filenameSektor270N)

            fileSektor315N = request.FILES['fotoSektor315N']
            filenameSektor315N = fs.save(fileSektor315N.name, fileSektor315N)
            file_pathSektor315N = fs.url(filenameSektor315N)

            section16Data = Section16(
                fotoSektor0N=patternFoto(
                    nama=filenameSektor0N, url=file_pathSektor0N),
                fotoSektor45N=patternFoto(
                    nama=filenameSektor45N, url=file_pathSektor45N),
                fotoSektor90N=patternFoto(
                    nama=filenameSektor90N, url=file_pathSektor90N),
                fotoSektor135N=patternFoto(
                    nama=filenameSektor135N, url=file_pathSektor135N),
                fotoSektor180N=patternFoto(
                    nama=filenameSektor180N, url=file_pathSektor180N),
                fotoSektor225N=patternFoto(
                    nama=filenameSektor225N, url=file_pathSektor225N),
                fotoSektor270N=patternFoto(
                    nama=filenameSektor270N, url=file_pathSektor270N),
                fotoSektor315N=patternFoto(
                    nama=filenameSektor315N, url=file_pathSektor315N),
                tempatFotoSektor=body_data.get("tempatFotoSektor"),
            )

            filePenggunaPotensial1 = request.FILES['fotoPenggunaPotensial1']
            filenamePenggunaPotensial1 = fs.save(
                filePenggunaPotensial1.name, filePenggunaPotensial1)
            file_pathPenggunaPotensial1 = fs.url(filenamePenggunaPotensial1)

            filePenggunaPotensial2 = request.FILES['fotoPenggunaPotensial2']
            filenamePenggunaPotensial2 = fs.save(
                filePenggunaPotensial2.name, filePenggunaPotensial2)
            file_pathPenggunaPotensial2 = fs.url(filenamePenggunaPotensial2)

            filePenggunaPotensial3 = request.FILES['fotoPenggunaPotensial3']
            filenamePenggunaPotensial3 = fs.save(
                filePenggunaPotensial3.name, filePenggunaPotensial3)
            file_pathPenggunaPotensial3 = fs.url(filenamePenggunaPotensial3)

            filePenggunaPotensial4 = request.FILES['fotoPenggunaPotensial4']
            filenamePenggunaPotensial4 = fs.save(
                filePenggunaPotensial4.name, filePenggunaPotensial4)
            file_pathPenggunaPotensial4 = fs.url(filenamePenggunaPotensial4)

            filePenggunaPotensial5 = request.FILES['fotoPenggunaPotensial5']
            filenamePenggunaPotensial5 = fs.save(
                filePenggunaPotensial5.name, filePenggunaPotensial5)
            file_pathPenggunaPotensial5 = fs.url(filenamePenggunaPotensial5)

            section17Data = Section17(
                fotoPenggunaPotensial1=patternFoto(
                    nama=filenamePenggunaPotensial1, url=file_pathPenggunaPotensial1),
                fotoPenggunaPotensial2=patternFoto(
                    nama=filenamePenggunaPotensial2, url=file_pathPenggunaPotensial2),
                fotoPenggunaPotensial3=patternFoto(
                    nama=filenamePenggunaPotensial3, url=file_pathPenggunaPotensial3),
                fotoPenggunaPotensial4=patternFoto(
                    nama=filenamePenggunaPotensial4, url=file_pathPenggunaPotensial4),
                fotoPenggunaPotensial5=patternFoto(
                    nama=filenamePenggunaPotensial5, url=file_pathPenggunaPotensial5),
            )

            fileAksesSite1 = request.FILES['fotoAksesSite1']
            filenameAksesSite1 = fs.save(fileAksesSite1.name, fileAksesSite1)
            file_pathAksesSite1 = fs.url(filenameAksesSite1)

            fileAksesSite2 = request.FILES['fotoAksesSite2']
            filenameAksesSite2 = fs.save(fileAksesSite2.name, fileAksesSite2)
            file_pathAksesSite2 = fs.url(filenameAksesSite2)

            fileAksesSite3 = request.FILES['fotoAksesSite3']
            filenameAksesSite3 = fs.save(fileAksesSite3.name, fileAksesSite3)
            file_pathAksesSite3 = fs.url(filenameAksesSite3)

            fileAksesSite4 = request.FILES['fotoAksesSite4']
            filenameAksesSite4 = fs.save(fileAksesSite4.name, fileAksesSite4)
            file_pathAksesSite4 = fs.url(filenameAksesSite4)

            fileAksesSite5 = request.FILES['fotoAksesSite5']
            filenameAksesSite5 = fs.save(fileAksesSite5.name, fileAksesSite5)
            file_pathAksesSite5 = fs.url(filenameAksesSite5)

            section18Data = Section18(
                fotoAksesSite1=patternFoto(
                    nama=filenameAksesSite1, url=file_pathAksesSite1),
                fotoAksesSite2=patternFoto(
                    nama=filenameAksesSite2, url=file_pathAksesSite2),
                fotoAksesSite3=patternFoto(
                    nama=filenameAksesSite3, url=file_pathAksesSite3),
                fotoAksesSite4=patternFoto(
                    nama=filenameAksesSite4, url=file_pathAksesSite4),
                fotoAksesSite5=patternFoto(
                    nama=filenameAksesSite5, url=file_pathAksesSite5),

            )

            hasilBts = hasilSurveyBts(
                user=body_data.get("userId"),
                kodeHasilSurvey=body_data.get("kodeHasilSurvey"),
                nomorSurvey=body_data.get("nomorSurvey"),
                section1=section1Data,
                comment=body_data.get("comment"),
                section2=section2Data,
                section3=section3Data,
                section4=section4Data,
                section5=section5Data,
                section6=section6Data,
                section7=section7Data,
                section8=section8Data,
                section9=section9Data,
                section10=section10Data,
                section11=section11Data,
                section12=section12Data,
                section13=section13Data,
                section14=section14Data,
                section15=section15Data,
                section16=section16Data,
                section17=section17Data,
                section18=section18Data
            )
            serializer = hasilSurveyBtsSerializer(hasilBts)
            result = serializer.data

            hasilBts.save()
            return Response.ok(
                message='Survey telah di submit',
                values=result
            )

    except Exception as e:
        return Response.badRequest(message=str(e))

        strjson = json.loads(request.body)
        userId = strjson["userId"]
        resultHasilSurvey = hasilSurvey.objects.filter(user=userId)
        serializer = hasilSurveySerializer(resultHasilSurvey, many=True)
        result = serializer.data
        return Response.ok(values=result, message=str(len(result))+' Buah Data telah terambil')

    except Exception as e:
        return Response.badRequest(message=str(e))


def getHasilSurveyByUserIdBTS(request):
    try:
        strjson = json.loads(request.body)
        userId = strjson["userId"]
        resultHasilSurveyBts = hasilSurveybts.objects.filter(user=userId)
        serializer = btsSerializer(resultHasilSurveyBts, many=True)
        result = serializer.data
        return Response.ok(values=result, message=str(len(result))+' Buah Data telah terambil')

    except Exception as e:
        return Response.badRequest(message=str(e))


# def logHasilSurvey(request):
#     try:


#     except Exception as e:
#         return Response.badRequest(message=str(e))


def postBTS(request):
    try:
        body_data = request.POST.dict()
        try:
            userrole = UserInfo.objects.get(id=body_data.get("userId"))
        except UserInfo.DoesNotExist:
            return Response.badRequest(
                values='null',
                message='User tidak ada'
            )

        if userrole.role.name != 'Staff Surveyor':
            return Response.badRequest(
                values='null',
                message='Anda tidak bisa submit hasil survey'
            )

        if request.method == "POST":
            statusDefault = [{'status': 'Submitted', 'tanggal_pembuatan': datetime.utcnow(
            ) + timedelta(hours=7)}]
            kode = body_data.get("kodeHasilSurvey").upper()

            nomorSurv = hasilSurveybts.objects.filter(
                kodeHasilSurvey=kode
            )
            nomorsurvey = str(len(nomorSurv)+1)
            organization = body_data.get("organization")
            # searchKode = Penugasan.objects.filter(
            #     kode=kode)
            # if len(searchKode) < 1:  # Duplicate Kode
            #     return Response.badRequest(
            #         values='null',
            #         message='Kode Survey Tidak Ditemukan'
            #     )
            try:
                fs = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/survey/foto/',
                    base_url=f'{settings.MEDIA_URL}/survey/foto/'
                )
            except Exception as e:
                return Response.badRequest(message=str(e))

            fileKandidatLahan = request.FILES['fotoKandidatLahan']
            filenameKandidatLahan = fs.save(
                fileKandidatLahan.name, fileKandidatLahan)
            file_pathKandidatLahan = fs.url(filenameKandidatLahan)

            fileMarkingGps = request.FILES['fotoMarkingGps']
            filenameMarkingGps = fs.save(fileMarkingGps.name, fileMarkingGps)
            file_pathMarkingGps = fs.url(filenameMarkingGps)

            fileUtaraTitik = request.FILES['fotoUtaraTitik']
            filenameUtaraTitik = fs.save(fileUtaraTitik.name, fileUtaraTitik)
            file_pathUtaraTitik = fs.url(filenameUtaraTitik)

            fileTimurTitik = request.FILES['fotoTimurTitik']
            filenameTimurTitik = fs.save(fileTimurTitik.name, fileTimurTitik)
            file_pathTimurTitik = fs.url(filenameTimurTitik)

            fileBaratTitik = request.FILES['fotoBaratTitik']
            filenameBaratTitik = fs.save(fileBaratTitik.name, fileBaratTitik)
            file_pathBaratTitik = fs.url(filenameBaratTitik)

            fileSelatanTitik = request.FILES['fotoSelatanTitik']
            filenameSelatanTitik = fs.save(
                fileSelatanTitik.name, fileSelatanTitik)
            file_pathSelatanTitik = fs.url(filenameSelatanTitik)

            date_time_str = body_data.get("tanggalPelaksanaan")if body_data.get(
                "tanggalPelaksanaan") != "" else '-'
            if date_time_str == '-':
                date_time_obj = None
            else:
                date_time_obj = datetime.strptime(
                    date_time_str, '%Y-%m-%d %H:%M:%S.%f')

            networkData = Network(
                tipe=body_data.get("tipenetwork")if body_data.get(
                    "tipenetwork") != "" else '-',
                download=body_data.get("download")if body_data.get(
                    "download") != "" else '-',
                upload=body_data.get("upload")if body_data.get(
                    "upload") != "" else '-'
            )

            hasilBts = hasilSurveybts(
                user=body_data.get("userId"),
                kodeHasilSurvey=kode,
                nomorSurvey=nomorsurvey,
                pic=Pic(
                    namaPic=body_data.get("namaPic")if body_data.get(
                        "namaPic") != "" else '-',
                    phonePic=body_data.get("phonePic")if body_data.get(
                        "phonePic") != "" else '-',
                ),
                note=body_data.get("note")if body_data.get(
                    "note") != "" else '-',
                tanggalPelaksanaan=date_time_obj,
                namaLokasi=body_data.get("namaLokasi")if body_data.get(
                    "namaLokasi") != "" else '-',
                latitude=body_data.get("latitude")if body_data.get(
                    "latitude") != "" else '-',
                longitude=body_data.get("longitude")if body_data.get(
                    "longitude") != "" else '-',
                fotoKandidatLahan=patternFoto(
                    nama=filenameKandidatLahan, url=file_pathKandidatLahan),
                fotoMarkingGps=patternFoto(
                    nama=filenameMarkingGps, url=file_pathMarkingGps),
                fotoUtaraTitik=patternFoto(
                    nama=filenameUtaraTitik, url=file_pathUtaraTitik),
                fotoTimurTitik=patternFoto(
                    nama=filenameTimurTitik, url=file_pathTimurTitik),
                fotoSelatanTitik=patternFoto(
                    nama=filenameSelatanTitik, url=file_pathSelatanTitik),
                fotoBaratTitik=patternFoto(
                    nama=filenameBaratTitik, url=file_pathBaratTitik),
                status=statusDefault,
                kategori=body_data.get("kategori")if body_data.get(
                    "kategori") != "" else '-',
                network=networkData,
                tanggal_pembuatan=datetime.utcnow() + timedelta(hours=7),
                tanggal_pembaruan=datetime.utcnow() + timedelta(hours=7)
            )
            hasilBts.save()

            usersadminsurveyor = UserInfo.objects.filter(
                organization=organization, role__in=['5f13b362386bf295b4169eff'])
            userto_ = []
            for usr in usersadminsurveyor:
                userto_.append(usr.username)

            notif = Message(
                title='Hasil survey',
                message='1 Hasil Survey Baru dari '+hasilBts.user.username,
                userfrom=ObjectId(hasilBts.user.id),
                userto=userto_,
                redirect='/adminsurveyor',
                status='new',
                created=datetime.utcnow() + timedelta(hours=7),
                updated=datetime.utcnow() + timedelta(hours=7)
            )
            notif.save()

            serializer = btsSerializer(hasilBts)
            result = serializer.data
            return Response.ok(
                message='Survey berhasil di submit',
                values=result
            )

    except Exception as e:
        return Response.badRequest(message=str(e))


def postBTSRelokasi(request):
    try:
        body_data = request.POST.dict()
        try:
            userrole = UserInfo.objects.get(id=body_data.get("userId"))
        except UserInfo.DoesNotExist:
            return Response.badRequest(
                values='null',
                message='User tidak ada'
            )

        if userrole.role.name != 'Staff Surveyor':
            return Response.badRequest(
                values='null',
                message='Anda tidak bisa submit hasil survey'
            )
        if request.method == "POST":
            statusDefault = [{'status': 'Submitted', 'tanggal_pembuatan': datetime.utcnow(
            ) + timedelta(hours=7)}]
            koderelokasi = body_data.get("kodeRelokasi").upper()
            kode = body_data.get("kodeHasilSurvey").upper()
            statusDefault = [{'status': 'Submitted', 'tanggal_pembuatan': datetime.utcnow(
            ) + timedelta(hours=7)}]
            nomorSurv = hasilSurveybts.objects.filter(
                kodeHasilSurvey=koderelokasi
            )
            nomorsurvey = str(len(nomorSurv)+1)
            # searchKode = Penugasan.objects.filter(
            #     kode=kode)
            # if len(searchKode) < 1:  # Duplicate Kode
            #     return Response.badRequest(
            #         values='null',
            #         message='Kode Survey Tidak Ditemukan'
            #     )
            try:
                fs = FileSystemStorage(
                    location=f'{settings.MEDIA_ROOT}/survey/foto/relokasi/',
                    base_url=f'{settings.MEDIA_URL}/survey/foto/relokasi/'
                )
            except Exception as e:
                return Response.badRequest(message=str(e))

            fileKandidatLahan = request.FILES['fotoKandidatLahan']
            filenameKandidatLahan = fs.save(
                fileKandidatLahan.name, fileKandidatLahan)
            file_pathKandidatLahan = fs.url(filenameKandidatLahan)

            fileMarkingGps = request.FILES['fotoMarkingGps']
            filenameMarkingGps = fs.save(fileMarkingGps.name, fileMarkingGps)
            file_pathMarkingGps = fs.url(filenameMarkingGps)

            fileUtaraTitik = request.FILES['fotoUtaraTitik']
            filenameUtaraTitik = fs.save(fileUtaraTitik.name, fileUtaraTitik)
            file_pathUtaraTitik = fs.url(filenameUtaraTitik)

            fileTimurTitik = request.FILES['fotoTimurTitik']
            filenameTimurTitik = fs.save(fileTimurTitik.name, fileTimurTitik)
            file_pathTimurTitik = fs.url(filenameTimurTitik)

            fileBaratTitik = request.FILES['fotoBaratTitik']
            filenameBaratTitik = fs.save(fileBaratTitik.name, fileBaratTitik)
            file_pathBaratTitik = fs.url(filenameBaratTitik)

            fileSelatanTitik = request.FILES['fotoSelatanTitik']
            filenameSelatanTitik = fs.save(
                fileSelatanTitik.name, fileSelatanTitik)
            file_pathSelatanTitik = fs.url(filenameSelatanTitik)

            date_time_str = body_data.get("tanggalPelaksanaan")if body_data.get(
                "tanggalPelaksanaan") != "" else '-'
            if date_time_str == '-':
                date_time_obj = None
            else:
                date_time_obj = datetime.strptime(
                    date_time_str, '%Y-%m-%d %H:%M:%S.%f')

            print('here')
            if nomorsurvey == '1':
                hasilSurv_ = hasilSurveybts(
                    user=body_data.get("userId"),
                    kodeHasilSurvey=kode,
                    nomorSurvey=nomorsurvey,
                    pic=Pic(),
                    note='-',
                    tanggalPelaksanaan=date_time_obj,
                    namaLokasi='-',
                    latitude='-',
                    longitude='-',
                    fotoKandidatLahan=patternFoto(),
                    fotoMarkingGps=patternFoto(),
                    fotoUtaraTitik=patternFoto(),
                    fotoTimurTitik=patternFoto(),
                    fotoSelatanTitik=patternFoto(),
                    fotoBaratTitik=patternFoto(),
                    relokasi=Relokasi(),
                    status=statusDefault,
                    kategori=body_data.get("kategori")if body_data.get(
                        "kategori") != "" else '-',
                    network=Network(),
                    tanggal_pembuatan=datetime.utcnow() + timedelta(hours=7),
                    tanggal_pembaruan=datetime.utcnow() + timedelta(hours=7)
                )

                hasilSurv_.save()

            try:
                _kodelama = str(hasilSurv_.id)
            except:
                hsl_sur = hasilSurveybts.objects.filter(
                    kodeHasilSurvey=kode, nomorSurvey='1', latitude='-'
                ).first()
                _kodelama = str(hsl_sur.id)

            relokasiData = Relokasi(
                provinsi=body_data.get("provinsi")if body_data.get(
                    "provinsi") != "" else '-',
                kab_kota=body_data.get("kab_kota")if body_data.get(
                    "kab_kota") != "" else '-',
                kecamatan=body_data.get("kecamatan")if body_data.get(
                    "kecamatan") != "" else '-',
                desa=body_data.get("desa")if body_data.get(
                    "desa") != "" else '-',
                alasan=body_data.get("alasan")if body_data.get(
                    "alasan") != "" else '-',
                kodelama=_kodelama
            )

            networkData = Network(
                tipe=body_data.get("tipenetwork")if body_data.get(
                    "tipenetwork") != "" else '-',
                download=body_data.get("download")if body_data.get(
                    "download") != "" else '-',
                upload=body_data.get("upload")if body_data.get(
                    "upload") != "" else '-'
            )

            hasilBts = hasilSurveybts(
                user=body_data.get("userId"),
                kodeHasilSurvey=koderelokasi,
                nomorSurvey=nomorsurvey,
                pic=Pic(
                    namaPic=body_data.get("namaPic")if body_data.get(
                        "namaPic") != "" else '-',
                    phonePic=body_data.get("phonePic")if body_data.get(
                        "phonePic") != "" else '-',
                ),
                note=body_data.get("note")if body_data.get(
                    "note") != "" else '-',
                tanggalPelaksanaan=date_time_obj,
                namaLokasi=body_data.get("namaLokasi")if body_data.get(
                    "namaLokasi") != "" else '-',
                latitude=body_data.get("latitude")if body_data.get(
                    "latitude") != "" else '-',
                longitude=body_data.get("longitude")if body_data.get(
                    "longitude") != "" else '-',
                fotoKandidatLahan=patternFoto(
                    nama=filenameKandidatLahan, url=file_pathKandidatLahan),
                fotoMarkingGps=patternFoto(
                    nama=filenameMarkingGps, url=file_pathMarkingGps),
                fotoUtaraTitik=patternFoto(
                    nama=filenameUtaraTitik, url=file_pathUtaraTitik),
                fotoTimurTitik=patternFoto(
                    nama=filenameTimurTitik, url=file_pathTimurTitik),
                fotoSelatanTitik=patternFoto(
                    nama=filenameSelatanTitik, url=file_pathSelatanTitik),
                fotoBaratTitik=patternFoto(
                    nama=filenameBaratTitik, url=file_pathBaratTitik),
                relokasi=relokasiData,
                status=statusDefault,
                kategori=body_data.get("kategori")if body_data.get(
                    "kategori") != "" else '-',
                network=networkData,
                tanggal_pembuatan=datetime.utcnow() + timedelta(hours=7),
                tanggal_pembaruan=datetime.utcnow() + timedelta(hours=7)
            )
            hasilBts.save()

            # url = "http://127.0.0.1:13000/survey/changestatuspenugasan/"
            # headers = {
            # 'Content-Type': 'application/json',
            # }
            # dicto={
            #     "kode":kode,
            #     "status":"on progress"
            # }
            # response = requests.request("POST", url, headers=headers, data = json.dumps(dicto))
            # print(response)
            serializer = btsSerializer(hasilBts)
            result = serializer.data
            return Response.ok(
                message='Survey berhasil di submit',
                values=result
            )

    except Exception as e:
        return Response.badRequest(message=str(e))


"""
def declinesurvey(request):
    strjson = json.loads(request.body.decode("utf-8"))
    kode_survey = strjson["kode"].upper()
    status = "Declined"
    jenis = strjson["jenis"]
    nomor_survey = strjson["nomorsurvey"]

    if request.method == 'POST':
        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)

            if jenis.upper() == 'AI':
                hasil_survey = hasilSurvey.objects.get(kodeHasilSurvey=kode_survey,nomorSurvey=nomor_survey)
            else:
                hasil_survey = hasilSurveybts.objects.get(kodeHasilSurvey=kode_survey,nomorSurvey=nomor_survey)
            if not hasil_survey:
                return Response.badRequest(
                    values='null',
                    message='Survey tidak ada'
                )
            statussurvey = [i for i, x in enumerate(
                hasil_survey.status) if x['status'] == status]
            if statussurvey:
                if jenis.upper() == 'AI':  
                    serializer = hasilSurveySerializer(hasil_survey)
                else:
                    serializer = btsSerializer(hasil_survey)
                result=serializer.data
                return Response.ok(
                    values=result,
                    message='Kode survey ditolak'
                )

            switcher = {
                "Declined": "Declined",
            }

            _status = switcher.get(status, 'None')
            if _status == 'None':
                return Response.badRequest(
                    values='null',
                    message='status tidak ada'
                )
            status = {'status': _status, 'tanggal_pembuatan': datetime.utcnow(
            ) + timedelta(hours=7)}
            hasil_survey.status.append(status)
            hasil_survey.save()
            if jenis.upper() == 'AI':  
                serializer = hasilSurveySerializer(hasil_survey)
            else:
                serializer = btsSerializer(hasil_survey)
            result=serializer.data
            return Response.ok(
                values=result,
                message='Kode survey ditolak'
            )
        except Exception as e:
            print(e)
            return HttpResponse(e)
    else:
        return HttpResponse('Post Only')
"""


def declinesurvey(request):
    strjson = json.loads(request.body.decode("utf-8"))
    kode_survey = strjson["kode"].upper()
    status_ = "Declined"
    jenis = strjson["jenis"]

    if request.method == 'POST':
        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)

            status = {'status': status_,
                      'tanggal_pembuatan': datetime.utcnow() + timedelta(hours=7)}
            if jenis.upper() == 'AI':
                hasil_survey = hasilSurvey.objects.filter(
                    kodeHasilSurvey=kode_survey)
            else:
                hasil_survey = hasilSurveybts.objects.filter(
                    kodeHasilSurvey=kode_survey)
            if len(hasil_survey) == 0:
                return Response.badRequest(
                    values='null',
                    message='Kode Survey tidak ada'
                )

            for h_sur in hasil_survey:
                statussurvey = [i for i, x in enumerate(
                    h_sur.status) if x['status'] == status_]
                if statussurvey:
                    continue

                # switcher = {
                #    "Decline": "Declined",
                # }

                #_status = switcher.get(status, "None")
                # if _status == "None":
                #    continue
                # print(_status)
                #status = {'status': status, 'tanggal_pembuatan': datetime.utcnow() + timedelta(hours=7)}
                h_sur.status.append(status)
                h_sur.tanggal_pembaruan = datetime.utcnow() + timedelta(hours=7)
                h_sur.save()
            if jenis.upper() == 'AI':
                serializer = hasilSurveySerializer(hasil_survey, many=True)
            else:
                serializer = btsSerializer(hasil_survey, many=True)
            result = serializer.data
            return Response.ok(
                values=result,
                message='Kode survey ditolak'
            )
        except Exception as e:
            print(e)
            return HttpResponse(e)
    else:
        return HttpResponse('Post Only')


def approvesurvey(request):
    strjson = json.loads(request.body.decode("utf-8"))
    kode_survey = strjson["kode"].upper()
    status_ = "Finished"
    jenis = strjson["jenis"]
    usersetujui = strjson["usersetujui"]
    #nomor_survey = strjson["nomorsurvey"]

    if request.method == 'POST':
        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)

            if jenis.upper() == 'AI':
                hasil_survey = hasilSurvey.objects.filter(
                    kodeHasilSurvey=kode_survey)
            else:
                hasil_survey = hasilSurveybts.objects.filter(
                    kodeHasilSurvey=kode_survey)
            if len(hasil_survey) == 0:
                return Response.badRequest(
                    values='null',
                    message='Kode Survey tidak ada'
                )

            status = {"status": status_,
                      "tanggal_pembuatan": datetime.utcnow() + timedelta(hours=7)}

            for h_sur in hasil_survey:
                statussurvey = [i for i, x in enumerate(
                    h_sur.status) if x['status'] == status_]

                if statussurvey:
                    continue

                #status = {"status": status, "tanggal_pembuatan": datetime.utcnow() + timedelta(hours=7)}
                h_sur.status.append(status)
                h_sur.tanggal_pembaruan = datetime.utcnow() + timedelta(hours=7)
                h_sur.save()

            status_penugasan = {"status": status_.lower(),
                                "date": datetime.utcnow() + timedelta(hours=7),
                                "user": usersetujui}

            try:
                penugasan = Penugasan.objects.get(kode=kode_survey)
                penugasan.status.append(status_penugasan)
                penugasan.save()
            except Penugasan.DoesNotExist:
                return Response.badRequest(
                    values='null',
                    message='Penugasan tidak ada'
                )
                # pass

            usersadminsurveyor = UserInfo.objects.filter(organization=ObjectId(
                penugasan.surveyor.id), role__in=['5f13b362386bf295b4169eff'])
            userto_ = []
            for usr in usersadminsurveyor:
                userto_.append(usr.username)

            notif = Message(
                title='Setujui Survey',
                message='Survei '+kode_survey+' Telah Disetujui',
                userfrom=usersetujui,
                userto=userto_,
                redirect='/adminsurveyor',
                status='new',
                created=datetime.utcnow() + timedelta(hours=7),
                updated=datetime.utcnow() + timedelta(hours=7)
            )
            notif.save()

            if jenis.upper() == 'AI':
                serializer = hasilSurveySerializer(hasil_survey, many=True)
            else:
                serializer = btsSerializer(hasil_survey, many=True)
            result = serializer.data
            return Response.ok(
                values=result,
                message='Kode survey disetujui'
            )
        except Exception as e:
            print(e)
            return HttpResponse(e)
    else:
        return HttpResponse('Post Only')


def setujuisurvey(request):
    strjson = json.loads(request.body.decode("utf-8"))
    kode_survey = strjson["kode"].upper()
    status_ = "Done"
    jenis = strjson["jenis"]
    userfrom = strjson.get("userfrom", None)
    #nomor_survey = strjson["nomorsurvey"]

    if request.method == 'POST':
        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)

            status = {"status": status_,
                      "tanggal_pembuatan": datetime.utcnow() + timedelta(hours=7)}

            if jenis.upper() == 'AI':
                hasil_survey = hasilSurvey.objects.filter(
                    kodeHasilSurvey=kode_survey)
            else:
                hasil_survey = hasilSurveybts.objects.filter(
                    kodeHasilSurvey=kode_survey)
            if len(hasil_survey) == 0:
                return Response.badRequest(
                    values='null',
                    message='Kode Survey tidak ada'
                )

            for h_sur in hasil_survey:
                statussurvey = [i for i, x in enumerate(
                    h_sur.status) if x['status'] == status_]
                if statussurvey:
                    continue

                #status = {"status": status, "tanggal_pembuatan": datetime.utcnow() + timedelta(hours=7)}
                h_sur.status.append(status)
                h_sur.tanggal_pembaruan = datetime.utcnow() + timedelta(hours=7)
                h_sur.save()

                if h_sur.nomorSurvey == '1':
                    if not userfrom:
                        usersadminsurveyor = UserInfo.objects.get(
                            username='adminsurveyor')
                    else:
                        usersadminsurveyor = UserInfo.objects.get(id=userfrom)
                    userfrom = usersadminsurveyor.id
                    usersadmin = UserInfo.objects.filter(
                        role__in=['5f13b1fa478ef95f4f0a83a7', '5f13b353386bf295b4169efe'])
                    tokens = list(usersadmin)
                    userto_ = []

                    for usr in usersadmin:
                        userto_.append(usr.username)

                    try:
                        token = UserToken.objects.get(user__in=tokens)
                        token_list = []
                        for tk in token:
                            token_list.append(token.key)
                        Notification(
                            users=token_list,
                            # users=['ef0D86AAQRG8Tu9ZXdEv2D:APA91bFmQDKHVjaTlRpUuHXEbXOVjywVyJuEoSrzjKLPqIrON4fviP9uJapeyZGQGFJ3WBODB_7xzFSeuNLpDZC0E_TMBH6jo8oJ5_QCF_qHCjBwxa7uQtacQGPgLgiI4DxoAKhJ1FcM'],
                            title='Setujui Survey',
                            message='1 Hasil Survey Baru dari '+usersadminsurveyor.name,
                        ).send_message()
                    except UserToken.DoesNotExist:
                        pass

                    notif = Message(
                        title='Setujui Survey',
                        message='1 Hasil Survey Baru dari ' + usersadminsurveyor.name,
                        userfrom=userfrom,
                        userto=userto_,
                        redirect='/admin',
                        status='new',
                        created=datetime.utcnow() + timedelta(hours=7),
                        updated=datetime.utcnow() + timedelta(hours=7)
                    )
                    notif.save()

                    notif_s = Message(
                        title='Setujui Survey',
                        message='Hasil Survey '+kode_survey +
                        ' berhasil disetujui ' + usersadminsurveyor.username,
                        userfrom=userfrom,
                        userto=[h_sur.user.username],
                        redirect='/admin',
                        status='new',
                        created=datetime.utcnow() + timedelta(hours=7),
                        updated=datetime.utcnow() + timedelta(hours=7)
                    )
                    notif_s.save()

            if jenis.upper() == 'AI':
                serializer = hasilSurveySerializer(hasil_survey, many=True)
            else:
                serializer = btsSerializer(hasil_survey, many=True)
            result = serializer.data
            return Response.ok(
                values=result,
                message='Kode survey disetujui'
            )
        except Exception as e:
            print(e)
            return HttpResponse(e)
    else:
        return HttpResponse('Post Only')


def tandaisurvey(request):
    strjson = json.loads(request.body.decode("utf-8"))
    kode_survey = strjson["kode"].upper()
    status_ = "Done"
    jenis = strjson["jenis"]
    alasan = strjson["alasan"]
    userfrom = strjson.get("userfrom", None)
    #nomor_survey = strjson["nomorsurvey"]

    if request.method == 'POST':
        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)

            if jenis.upper() == 'AI':
                hasil_survey = hasilSurvey.objects.filter(
                    kodeHasilSurvey=kode_survey)
            else:
                hasil_survey = hasilSurveybts.objects.filter(
                    kodeHasilSurvey=kode_survey)

            if len(hasil_survey) == 0:
                return Response.badRequest(
                    values='null',
                    message='Kode Survey tidak ada'
                )
            status = {"status": status_,
                      "tanggal_pembuatan": datetime.utcnow() + timedelta(hours=7)}
            status_issue = {"status": "Issued", "alasan": alasan,
                            "tanggal_pembuatan": datetime.utcnow() + timedelta(hours=7), }
            for h_sur in hasil_survey:
                statussurvey = [i for i, x in enumerate(
                    h_sur.status) if x['status'] == status_]
                if statussurvey:
                    continue

                #status = {"status": status, "tanggal_pembuatan": datetime.utcnow() + timedelta(hours=7)}
                h_sur.status.append(status)
                #status_issue = {"status": "Issued", "alasan": alasan, "tanggal_pembuatan": datetime.utcnow() + timedelta(hours=7), }
                h_sur.issue.append(status_issue)
                h_sur.tanggal_pembaruan = datetime.utcnow() + timedelta(hours=7)
                h_sur.save()

                if h_sur.nomorSurvey == '1':
                    if not userfrom:
                        usersadminsurveyor = UserInfo.objects.get(
                            username='adminsurveyor')
                    else:
                        usersadminsurveyor = UserInfo.objects.get(id=userfrom)
                    userfrom = usersadminsurveyor.id

                    usersadmin = UserInfo.objects.filter(
                        role__in=['5f13b1fa478ef95f4f0a83a7', '5f13b353386bf295b4169efe'])
                    tokens = list(usersadmin)
                    userto_ = []
                    for usr in usersadmin:
                        userto_.append(usr.username)
                    try:
                        token = UserToken.objects.get(user__in=tokens)
                        token_list = []
                        for tk in token:
                            token_list.append(token.key)
                        Notification(
                            users=token_list,
                            # users=['ef0D86AAQRG8Tu9ZXdEv2D:APA91bFmQDKHVjaTlRpUuHXEbXOVjywVyJuEoSrzjKLPqIrON4fviP9uJapeyZGQGFJ3WBODB_7xzFSeuNLpDZC0E_TMBH6jo8oJ5_QCF_qHCjBwxa7uQtacQGPgLgiI4DxoAKhJ1FcM'],
                            title='Issue Survey',
                            message='1 Temuan hasil survey baru dari '+usersadminsurveyor.name,
                        ).send_message()
                    except UserToken.DoesNotExist:
                        pass

                    notif = Message(
                        title='Issue Survey',
                        message='1 Temuan hasil survey baru dari ' + usersadminsurveyor.name,
                        userfrom=userfrom,
                        userto=userto_,
                        redirect='/admin',
                        status='new',
                        created=datetime.utcnow() + timedelta(hours=7),
                        updated=datetime.utcnow() + timedelta(hours=7)
                    )
                    notif.save()

                    notif_s = Message(
                        title='Issue Survey',
                        message='Hasil Survey '+kode_survey +
                        ' berhasil ditandai ' + usersadminsurveyor.username,
                        userfrom=userfrom,
                        userto=[h_sur.user.username],
                        redirect='/admin',
                        status='new',
                        created=datetime.utcnow() + timedelta(hours=7),
                        updated=datetime.utcnow() + timedelta(hours=7)
                    )
                    notif_s.save()
            if jenis.upper() == 'AI':
                serializer = hasilSurveySerializer(hasil_survey, many=True)
            else:
                serializer = btsSerializer(hasil_survey, many=True)
            result = serializer.data

            return Response.ok(
                values=result,
                message='Kode survey disetujui'
            )
        except Exception as e:
            print(e)
            return HttpResponse(e)
    else:
        return HttpResponse('Post Only')


def changeStatusAI(request):
    body_data = request.POST.dict()
    if request.method == "POST":
        kode = body_data.get("kodeHasilSurvey")
        statuss = body_data.get("status")
        #noSurvey = body_data.get("nomorSurvey")
        # try:
        searchKode = hasilSurvey.objects.filter(
            kodeHasilSurvey=kode)  # , nomorSurvey=noSurvey
        # except hasilSurvey.DoesNotExist:
        #    searchKode = None
        status = 0
        if len(searchKode) > 0:
            for dt in searchKode:
                """
                if len(dt.status) == 0:
                    statusDefault = {'status': 'Submitted', 'tanggal_pembuatan': datetime.utcnow(
                    ) + timedelta(hours=7)}
                    dt.status.append(statusDefault)
                elif len(dt.status) == 1:
                    statusReview = {'status': 'Reviewed', 'tanggal_pembuatan': datetime.utcnow(
                    ) + timedelta(hours=7)}
                    dt.status.append(statusReview)
                elif len(dt.status) == 2:

                    statusDone = {'status': 'Done', 'tanggal_pembuatan': datetime.utcnow(
                    ) + timedelta(hours=7)}
                    dt.status.append(statusDone)
                """
                statussurvey = [i for i, x in enumerate(
                    dt.status) if x['status'] == statuss]
                if statussurvey:
                    continue
                status = {'status': statuss, 'tanggal_pembuatan': datetime.utcnow(
                ) + timedelta(hours=7)}
                dt.status.append(status)
                dt.tanggal_pembaruan = datetime.utcnow() + \
                    timedelta(hours=7)
                dt.save()
            serializer = hasilSurveySerializer(searchKode, many=True)
            result = serializer.data
            return Response.ok(
                values=result,
                message='Status telah terupdate'
            )
        else:
            return Response.badRequest(
                values='null',
                message='Kode Hasil Survey tidak ditemukan'
            )


def changeStatusBTS(request):
    body_data = request.POST.dict()
    if request.method == "POST":
        kode = body_data.get("kodeHasilSurvey")
        statuss = body_data.get("status")
        #noSurvey = body_data.get("nomorSurvey")
        # try:
        searchKode = hasilSurveybts.objects.filter(
            kodeHasilSurvey=kode)  # , nomorSurvey=noSurvey
        # except hasilSurvey.DoesNotExist:
        #    searchKode = None
        status = 0
        if len(searchKode) > 0:
            for dt in searchKode:
                """
                if len(dt.status) == 0:
                    statusDefault = {'status': 'Submitted', 'tanggal_pembuatan': datetime.utcnow(
                    ) + timedelta(hours=7)}
                    dt.status.append(statusDefault)
                elif len(dt.status) == 1:
                    statusReview = {'status': 'Reviewed', 'tanggal_pembuatan': datetime.utcnow(
                    ) + timedelta(hours=7)}
                    dt.status.append(statusReview)
                elif len(dt.status) == 2:

                    statusDone = {'status': 'Done', 'tanggal_pembuatan': datetime.utcnow(
                    ) + timedelta(hours=7)}
                    dt.status.append(statusDone)
                """
                statussurvey = [i for i, x in enumerate(
                    dt.status) if x['status'] == statuss]
                if statussurvey:
                    continue
                status = {'status': statuss, 'tanggal_pembuatan': datetime.utcnow(
                ) + timedelta(hours=7)}
                dt.status.append(status)
                dt.tanggal_pembaruan = datetime.utcnow() + \
                    timedelta(hours=7)
                dt.save()
            serializer = btsSerializer(searchKode, many=True)
            result = serializer.data
            return Response.ok(
                values=result,
                message='Status telah terupdate'
            )
        else:
            return Response.badRequest(
                values='null',
                message='Kode Hasil Survey tidak ditemukan'
            )


def changeStatus(request):
    body_data = request.POST.dict()
    if request.method == "POST":
        kode = body_data.get("kodeHasilSurvey")
        noSurvey = body_data.get("nomorSurvey")
        try:
            searchKode = hasilSurvey.objects.get(
                kodeHasilSurvey=kode, nomorSurvey=noSurvey)
        except hasilSurvey.DoesNotExist:
            searchKode = None

        if searchKode != None:
            if len(searchKode.status) == 0:
                statusDefault = {'status': 'Submitted', 'tanggal_pembuatan': datetime.utcnow(
                ) + timedelta(hours=7)}
                searchKode.status.append(statusDefault)
            elif len(searchKode.status) == 1:
                statusReview = {'status': 'Reviewed', 'tanggal_pembuatan': datetime.utcnow(
                ) + timedelta(hours=7)}
                searchKode.status.append(statusReview)
            elif len(searchKode.status) == 2:

                statusDone = {'status': 'Done', 'tanggal_pembuatan': datetime.utcnow(
                ) + timedelta(hours=7)}
                searchKode.status.append(statusDone)
            else:
                return Response.badRequest(
                    values=searchKode.serializeStatusOnly(),
                    message='status hasil survey sudah selesai'
                )
            searchKode.tanggal_pembaruan = datetime.utcnow() + \
                timedelta(hours=7)
            searchKode.save()
            return Response.ok(
                values=searchKode.serializeStatusOnly(),
                message='Status telah terupdate'
            )
        else:
            return Response.badRequest(
                values='null',
                message='Kode Hasil Survey tidak ditemukan'
            )


def getHasilSurveyByUserIdAI(request):
    # try:
    body_data = request.POST.dict()
    userId = body_data.get("userId")
    resultHasilSurvey = hasilSurvey.objects.filter(
        user=userId, issue__0__exists=False)
    setab = set()
    listResult = []
    for i in resultHasilSurvey:
        # print(len(resultHasilSurvey.issue))
        # if len(resultHasilSurvey.issue)==0:
        awal = len(setab)
        setab.add(i['kodeHasilSurvey'])
        if len(setab) > awal:
            listResult.append(i)
    if len(resultHasilSurvey) > 0:
        serializer = hasilSurveySerializer(listResult, many=True)
        result = serializer.data
        for item in result:
            if item['listFoto']['plang'] == None:
                item['listFoto']['aksesJalan'] = {'nama': '-', 'url': '-'}
                item['listFoto']['plang'] = {'nama': '-', 'url': '-'}
                item['listFoto']['markingPerangkat'] = {
                    'nama': '-', 'url': '-'}
                item['listFoto']['kwhMeter'] = {'nama': '-', 'url': '-'}
                item['listFoto']['gambarDenah'] = {'nama': '-', 'url': '-'}
                item['listFoto']['lanskapBangunan'] = {'nama': '-', 'url': '-'}
        return Response.ok(values=result, message=str(len(result))+' Buah Data telah terambil')
    else:
        return Response.ok(values=[], message='0 Buah Data telah terambil')
    # except Exception as e:
    #    return Response.badRequest(message=str(e))


def getHasilSurveyByUserAI(request):
    try:
        body_data = request.POST.dict()
        userId = body_data.get("userId")
        pipeline = [
            {
                '$match': {
                    'user': ObjectId(userId),
                    'issue.0': {
                        '$exists': False
                    }
                }
            }, {
                '$lookup': {
                    'from': 'user_info',
                    'localField': 'user',
                    'foreignField': '_id',
                    'as': 'user'
                }
            }, {
                '$unwind': {
                    'path': '$user',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'surveyor',
                    'localField': 'user.organization',
                    'foreignField': '_id',
                    'as': 'user.organization'
                }
            }, {
                '$unwind': {
                    'path': '$user.organization',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'userrole',
                    'localField': 'user.role',
                    'foreignField': '_id',
                    'as': 'user.role'
                }
            }, {
                '$unwind': {
                    'path': '$user.role',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'document_user',
                    'localField': 'user.doc',
                    'foreignField': '_id',
                    'as': 'user.doc'
                }
            }, {
                '$unwind': {
                    'path': '$user.doc',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$sort': {
                    'tanggal_pembaruan': -1
                }
            },  # {
            #    '$addFields': {
            #        'lastStatus': {
            #            '$arrayElemAt': [
            #                '$status', -1
            #            ]
            #        }
            #    }
            # }, {
            #    '$match': {
            #        'lastStatus.status': 'Submitted'
            #    }
            # }, {
            #    '$project': {
            #        'lastStatus': 0
            #    }
            # },
            {
                '$group': {
                    '_id': '$kodeHasilSurvey',
                    'data': {
                        '$push': '$$ROOT'
                    }
                }
            }, {
                '$lookup': {
                    'from': 'penugasan',
                    'localField': '_id',
                    'foreignField': 'kode',
                    'as': 'penugasan'
                }
            }, {
                '$unwind': {
                    'path': '$penugasan',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'lokasi_survey',
                    'localField': 'penugasan.lokasisurvey',
                    'foreignField': '_id',
                    'as': 'lokasisurvey'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'provinsi',
                    'localField': 'lokasisurvey.provinsi',
                    'foreignField': '_id',
                    'as': 'lokasi.provinsi'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.provinsi',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kabupaten',
                    'localField': 'lokasisurvey.kabupaten',
                    'foreignField': '_id',
                    'as': 'lokasi.kabupaten'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.kabupaten',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kota',
                    'localField': 'lokasisurvey.kota',
                    'foreignField': '_id',
                    'as': 'lokasi.kota'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.kota',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kecamatan',
                    'localField': 'lokasisurvey.kecamatan',
                    'foreignField': '_id',
                    'as': 'lokasi.kecamatan'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.kecamatan',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'desa',
                    'localField': 'lokasisurvey.desa',
                    'foreignField': '_id',
                    'as': 'lokasi.desa'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.desa',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$addFields': {
                    'last': {
                        '$arrayElemAt': [
                            '$data', -1
                        ]
                    }
                }
            }, {
                '$sort': {
                    'last.tanggal_pembaruan': -1
                }
            }, {
                '$project': {
                    '_id': 0,
                    'kode': '$_id',
                    'lokasi': '$lokasi',
                    'data': '$data'
                }
            }
        ]

        resultHasilSurvey = list(hasilSurvey.objects.aggregate(pipeline))

        return Response.ok(values=json.loads(json.dumps(resultHasilSurvey, default=str)), message=str(len(resultHasilSurvey))+' Buah Data telah terambil')
    except Exception as e:
        return Response.badRequest(values=[], message=str(e))


def getHasilSurveyByUserIdBTS(request):
    try:
        body_data = request.POST.dict()
        userId = body_data.get("userId")
        resultHasilSurveyBts = hasilSurveybts.objects.filter(
            user=userId, issue__0__exists=False)
        setab = set()
        listResult = []
        for i in resultHasilSurveyBts:
            awal = len(setab)
            setab.add(i['kodeHasilSurvey'])
            if len(setab) > awal:
                listResult.append(i)
        if len(resultHasilSurveyBts) > 0:
            serializer = btsSerializer(listResult, many=True)
            result = serializer.data
            for item in result:
                if item['fotoKandidatLahan'] == None:
                    item['fotoKandidatLahan'] = {'nama': '-', 'url': '-'}
                    item['fotoMarkingGps'] = {'nama': '-', 'url': '-'}
                    item['fotoUtaraTitik'] = {'nama': '-', 'url': '-'}
                    item['fotoTimurTitik'] = {'nama': '-', 'url': '-'}
                    item['fotoSelatanTitik'] = {'nama': '-', 'url': '-'}
                    item['fotoBaratTitik'] = {'nama': '-', 'url': '-'}
            return Response.ok(values=result, message=str(len(result))+' Buah Data telah terambil')
        else:
            return Response.ok(values=[], message='0 Buah Data telah terambil')

    except Exception as e:
        return Response.badRequest(message=str(e))


def getHasilSurveyByUserBTS(request):
    try:
        body_data = request.POST.dict()
        userId = body_data.get("userId")
        pipeline = [
            {
                '$match': {
                    'user': ObjectId(userId),
                    'issue.0': {
                        '$exists': False
                    }
                }
            }, {
                '$lookup': {
                    'from': 'user_info',
                    'localField': 'user',
                    'foreignField': '_id',
                    'as': 'user'
                }
            }, {
                '$unwind': {
                    'path': '$user',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'surveyor',
                    'localField': 'user.organization',
                    'foreignField': '_id',
                    'as': 'user.organization'
                }
            }, {
                '$unwind': {
                    'path': '$user.organization',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'userrole',
                    'localField': 'user.role',
                    'foreignField': '_id',
                    'as': 'user.role'
                }
            }, {
                '$unwind': {
                    'path': '$user.role',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'document_user',
                    'localField': 'user.doc',
                    'foreignField': '_id',
                    'as': 'user.doc'
                }
            }, {
                '$unwind': {
                    'path': '$user.doc',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$sort': {
                    'tanggal_pembaruan': -1
                }
            },  # {
            #    '$addFields': {
            #        'lastStatus': {
            #            '$arrayElemAt': [
            #                '$status', -1
            #            ]
            #        }
            #    }
            # }, {
            #    '$match': {
            #        'lastStatus.status': 'Submitted'
            #    }
            # }, {
            #    '$project': {
            #        'lastStatus': 0
            #    }
            # },
            {
                '$group': {
                    '_id': '$kodeHasilSurvey',
                    'data': {
                        '$push': '$$ROOT'
                    }
                }
            }, {
                '$lookup': {
                    'from': 'penugasan',
                    'localField': '_id',
                    'foreignField': 'kode',
                    'as': 'penugasan'
                }
            }, {
                '$unwind': {
                    'path': '$penugasan',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'lokasi_survey',
                    'localField': 'penugasan.lokasisurvey',
                    'foreignField': '_id',
                    'as': 'lokasisurvey'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'provinsi',
                    'localField': 'lokasisurvey.provinsi',
                    'foreignField': '_id',
                    'as': 'lokasi.provinsi'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.provinsi',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kabupaten',
                    'localField': 'lokasisurvey.kabupaten',
                    'foreignField': '_id',
                    'as': 'lokasi.kabupaten'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.kabupaten',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kota',
                    'localField': 'lokasisurvey.kota',
                    'foreignField': '_id',
                    'as': 'lokasi.kota'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.kota',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kecamatan',
                    'localField': 'lokasisurvey.kecamatan',
                    'foreignField': '_id',
                    'as': 'lokasi.kecamatan'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.kecamatan',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'desa',
                    'localField': 'lokasisurvey.desa',
                    'foreignField': '_id',
                    'as': 'lokasi.desa'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.desa',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$addFields': {
                    'last': {
                        '$arrayElemAt': [
                            '$data', -1
                        ]
                    }
                }
            }, {
                '$sort': {
                    'last.tanggal_pembaruan': -1
                }
            }, {
                '$project': {
                    '_id': 0,
                    'kode': '$_id',
                    'lokasi': '$lokasi',
                    'data': '$data'
                }
            }
        ]

        resultHasilSurvey = list(hasilSurveybts.objects.aggregate(pipeline))

        return Response.ok(values=json.loads(json.dumps(resultHasilSurvey, default=str)), message=str(len(resultHasilSurvey))+' Buah Data telah terambil')
    except Exception as e:
        return Response.badRequest(values=[], message=str(e))


def getHasilSurveyByUserIdAIdetail(request):
    try:
        body_data = request.POST.dict()
        kodeSurvey = body_data.get('kode')
        userId = body_data.get("userId")
        resultHasilSurvey = hasilSurvey.objects.filter(
            kodeHasilSurvey=kodeSurvey, user=ObjectId(userId))
        if len(resultHasilSurvey) > 0:
            serializer = hasilSurveySerializer(resultHasilSurvey, many=True)
            result = serializer.data
            for item in result:
                if item['listFoto']['plang'] == None:
                    item['listFoto']['aksesJalan'] = {'nama': '-', 'url': '-'}
                    item['listFoto']['plang'] = {'nama': '-', 'url': '-'}
                    item['listFoto']['markingPerangkat'] = {
                        'nama': '-', 'url': '-'}
                    item['listFoto']['kwhMeter'] = {'nama': '-', 'url': '-'}
                    item['listFoto']['gambarDenah'] = {'nama': '-', 'url': '-'}
                    item['listFoto']['lanskapBangunan'] = {
                        'nama': '-', 'url': '-'}
            return Response.ok(values=result, message=str(len(result))+' Buah Data telah terambil')
        else:
            return Response.ok(values=[], message='0 Buah Data telah terambil')
    except Exception as e:
        return Response.badRequest(message=str(e))


def getHasilSurveyByUserIdBtsdetail(request):
    try:
        body_data = request.POST.dict()
        kodeSurvey = body_data.get('kode')
        userId = body_data.get("userId")
        resultHasilSurveyBts = hasilSurveybts.objects.filter(
            kodeHasilSurvey=kodeSurvey, user=ObjectId(userId))
        if len(resultHasilSurveyBts) > 0:
            serializer = btsSerializer(resultHasilSurveyBts, many=True)
            result = serializer.data
            for item in result:
                if item['fotoKandidatLahan'] == None:
                    item['fotoKandidatLahan'] = {'nama': '-', 'url': '-'}
                    item['fotoMarkingGps'] = {'nama': '-', 'url': '-'}
                    item['fotoUtaraTitik'] = {'nama': '-', 'url': '-'}
                    item['fotoTimurTitik'] = {'nama': '-', 'url': '-'}
                    item['fotoSelatanTitik'] = {'nama': '-', 'url': '-'}
                    item['fotoBaratTitik'] = {'nama': '-', 'url': '-'}

            return Response.ok(values=result, message=str(len(result))+' Buah Data telah terambil')
        else:
            return Response.ok(values=[], message='0 Buah Data telah terambil')
    except Exception as e:
        return Response.badRequest(message=str(e))


def getHasilSurveyAiById(request):
    try:
        jsonResponse = json.loads(request.body.decode('utf-8'))
        idhasilSurvey = jsonResponse['id']
        resultHasilSurvey = hasilSurvey.objects.filter(
            id=ObjectId(idhasilSurvey))
        if len(resultHasilSurvey) > 0:
            serializer = hasilSurveySerializer(resultHasilSurvey, many=True)
            result = serializer.data
            return Response.ok(values=result, message='success')
        else:
            return Response.ok(values='', message='No data')
    except expression as identifier:
        return Response.badRequest(message=str(e))


def getHasilSurveyBtsById(request):
    try:
        jsonResponse = json.loads(request.body.decode('utf-8'))
        idhasilSurvey = jsonResponse['id']
        resultHasilSurvey = hasilSurveybts.objects.filter(
            id=ObjectId(idhasilSurvey))
        if len(resultHasilSurvey) > 0:
            serializer = btsSerializer(resultHasilSurvey, many=True)
            result = serializer.data
            return Response.ok(values=result, message='success')
        else:
            return Response.ok(values='', message='No data')
    except expression as identifier:
        return Response.badRequest(message=str(e))


def getissuebysurveyor(request):
    try:
        req = request.body.decode("utf-8")
        data = json.loads(req)
        # status_ = data['status']
        jenis = data['jenis'].upper()
        surveyor = data['surveyor']
        page = int(data.get('page', 0)) - 1
        skip = []
        if page >= 0:
            skip = [{'$skip': 20 * page},
                    {'$limit': 20}]
        pipeline = [
            {
                '$match': {
                    'issue': {
                        '$exists': True
                    }
                }
            }, {
                '$addFields': {
                    'size_of_issue': {
                        '$size': '$issue'
                    }
                }
            }, {
                '$match': {
                    'size_of_issue': {
                        '$gt': 0
                    }
                }
            }, {
                '$addFields': {
                    'lastStatus': {
                        '$arrayElemAt': ['$status', -1]
                    }
                }
            }, {
                '$match': {
                    'lastStatus.status': 'Done'
                }
            }, {
                '$project': {
                    'size_of_issue': 0
                }
            }, {
                '$lookup': {
                    'from': 'user_info',
                    'localField': 'user',
                    'foreignField': '_id',
                    'as': 'user'
                }
            }, {
                '$unwind': {
                    'path': '$user',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$match': {
                    'user.organization': ObjectId(surveyor)
                }
            }, {
                '$lookup': {
                    'from': 'surveyor',
                    'localField': 'user.organization',
                    'foreignField': '_id',
                    'as': 'user.organization'
                }
            }, {
                '$unwind': {
                    'path': '$user.organization',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'userrole',
                    'localField': 'user.role',
                    'foreignField': '_id',
                    'as': 'user.role'
                }
            }, {
                '$unwind': {
                    'path': '$user.role',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'document_user',
                    'localField': 'user.doc',
                    'foreignField': '_id',
                    'as': 'user.doc'
                }
            }, {
                '$unwind': {
                    'path': '$user.doc',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$group': {
                    '_id': '$kodeHasilSurvey',
                    'data': {
                        '$push': '$$ROOT'
                    }
                }
            }, {
                '$lookup': {
                    'from': 'penugasan',
                    'localField': '_id',
                    'foreignField': 'kode',
                    'as': 'penugasan'
                }
            }, {
                '$unwind': {
                    'path': '$penugasan',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'lokasi_survey',
                    'localField': 'penugasan.lokasisurvey',
                    'foreignField': '_id',
                    'as': 'lokasisurvey'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'provinsi',
                    'localField': 'lokasisurvey.provinsi',
                    'foreignField': '_id',
                    'as': 'lokasi.provinsi'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.provinsi',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kabupaten',
                    'localField': 'lokasisurvey.kabupaten',
                    'foreignField': '_id',
                    'as': 'lokasi.kabupaten'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.kabupaten',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kota',
                    'localField': 'lokasisurvey.kota',
                    'foreignField': '_id',
                    'as': 'lokasi.kota'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.kota',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kecamatan',
                    'localField': 'lokasisurvey.kecamatan',
                    'foreignField': '_id',
                    'as': 'lokasi.kecamatan'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.kecamatan',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'desa',
                    'localField': 'lokasisurvey.desa',
                    'foreignField': '_id',
                    'as': 'lokasi.desa'
                }
            }, {
                '$unwind': {
                    'path': '$lokasi.desa',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$addFields': {
                    'last': {
                        '$arrayElemAt': ['$data', -1]
                    },
                }
            }, {
                '$sort': {
                    'last.tanggal_pembaruan': -1
                }
            },
            {
                '$project': {
                    '_id': 0,
                    'kode': '$_id',
                    'lokasi': '$lokasi',
                    'data': '$data'
                }
            }
        ]
        pipe = pipeline + skip
        if jenis.upper() == 'AI':
            # hasil_survey = hasilSurvey.objects.filter(issue__0__exists=True)#likes__21__exists=True
            agg_cursor = hasilSurvey.objects.aggregate(*pipeline)
        else:
            # hasil_survey = hasilSurveybts.objects.filter(issue__0__exists=True)
            agg_cursor = hasilSurveybts.objects.aggregate(*pipe)

        search = list(agg_cursor)
        if len(search) > 0:
            return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
        else:
            return Response.badRequest(message='Data tidak Ditemukan')
        """
        dataFrameHasil = pandas.DataFrame(agg_cursor)
        if len(dataFrameHasil.index)>0:
            Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
            return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
        else:
            return Response.ok(message='Data tidak Ditemukan')
        """
    except Exception as e:
        print(e)
        return HttpResponse(e)


def getsurveyissue(request):
    # if request.method == 'POST':
    try:
        req = request.body.decode("utf-8")
        data = json.loads(req)
        # status_ = data['status']
        jenis = data['jenis'].upper()
        page = int(data.get('page', 0)) - 1
        skip = []
        if page >= 0:
            skip = [{'$skip': 20 * page},
                    {'$limit': 20}]
        pipeline = [{
            '$match': {
                'issue': {
                    '$exists': True
                }
            }
        }, {
            '$addFields': {
                'size_of_issue': {
                    '$size': '$issue'
                }
            }
        }, {
            '$match': {
                'size_of_issue': {
                    '$gt': 0
                }
            }
        }, {
            '$addFields': {
                'lastStatus': {
                    '$arrayElemAt': ['$status', -1]
                }
            }
        }, {
            '$match': {
                'lastStatus.status': 'Done'
            }
        },
            {
            '$project': {
                'size_of_issue': 0,
                'lastStatus': 0
            }
        }, {
            '$lookup': {
                'from': 'user_info',
                'localField': 'user',
                'foreignField': '_id',
                'as': 'user'
            }
        }, {
            '$unwind': {
                'path': '$user',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'surveyor',
                'localField': 'user.organization',
                'foreignField': '_id',
                'as': 'user.organization'
            }
        }, {
            '$unwind': {
                'path': '$user.organization',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'userrole',
                'localField': 'user.role',
                'foreignField': '_id',
                'as': 'user.role'
            }
        }, {
            '$unwind': {
                'path': '$user.role',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'document_user',
                'localField': 'user.doc',
                'foreignField': '_id',
                'as': 'user.doc'
            }
        }, {
            '$unwind': {
                'path': '$user.doc',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$sort': {
                'tanggal_pembaruan': -1
            }
        }, {
            '$group': {
                '_id': '$kodeHasilSurvey',
                'data': {
                    '$push': '$$ROOT'
                }
            }
        }, {
            '$lookup': {
                'from': 'penugasan',
                'localField': '_id',
                'foreignField': 'kode',
                'as': 'penugasan'
            }
        }, {
            '$unwind': {
                'path': '$penugasan',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'lokasi_survey',
                'localField': 'penugasan.lokasisurvey',
                'foreignField': '_id',
                'as': 'lokasisurvey'
            }
        }, {
            '$unwind': {
                'path': '$lokasisurvey',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'provinsi',
                'localField': 'lokasisurvey.provinsi',
                'foreignField': '_id',
                'as': 'lokasi.provinsi'
            }
        }, {
            '$unwind': {
                'path': '$lokasi.provinsi',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'kabupaten',
                'localField': 'lokasisurvey.kabupaten',
                'foreignField': '_id',
                'as': 'lokasi.kabupaten'
            }
        }, {
            '$unwind': {
                'path': '$lokasi.kabupaten',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'kota',
                'localField': 'lokasisurvey.kota',
                'foreignField': '_id',
                'as': 'lokasi.kota'
            }
        }, {
            '$unwind': {
                'path': '$lokasi.kota',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'kecamatan',
                'localField': 'lokasisurvey.kecamatan',
                'foreignField': '_id',
                'as': 'lokasi.kecamatan'
            }
        }, {
            '$unwind': {
                'path': '$lokasi.kecamatan',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$lookup': {
                'from': 'desa',
                'localField': 'lokasisurvey.desa',
                'foreignField': '_id',
                'as': 'lokasi.desa'
            }
        }, {
            '$unwind': {
                'path': '$lokasi.desa',
                'preserveNullAndEmptyArrays': True
            }
        }, {
            '$addFields': {
                'last': {
                    '$arrayElemAt': [
                        '$data', -1
                    ]
                }
            }
        }, {
            '$sort': {
                'last.tanggal_pembaruan': -1
            }
        }, {
            '$project': {
                '_id': 0,
                'kode': '$_id',
                'lokasi': '$lokasi',
                'data': '$data'
            }
        }]
        pipe = pipeline + skip
        if jenis.upper() == 'AI':
            # hasil_survey = hasilSurvey.objects.filter(issue__0__exists=True)#likes__21__exists=True
            agg_cursor = hasilSurvey.objects.aggregate(*pipe)
        else:
            # hasil_survey = hasilSurveybts.objects.filter(issue__0__exists=True)
            agg_cursor = hasilSurveybts.objects.aggregate(*pipe)

        search = list(agg_cursor)
        if len(search) > 0:
            return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
        else:
            return Response.badRequest(values='null', message='Data tidak Ditemukan')
        """
        dataFrameHasil = pandas.DataFrame(agg_cursor)
        if len(dataFrameHasil.index)>0:
            Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
            return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
        else:
            return Response.ok(message='Data tidak Ditemukan')
        """
        # if len(hasil_survey)==0:
        #     return Response.badRequest(
        #         values='null',
        #         message='Issued Survey tidak ada'
        #     )
        # json_ret = []
        # for k,v in groupby(hasil_survey,key=lambda x:x['kodeHasilSurvey'].strip()):
        #         json_dict = {}
        #         json_dict["kodeSurvey"] = k.strip()
        #         json_dict["data"] = []
        #         for dt in list(v):
        #             if jenis.upper() == 'AI':
        #                 serializer = hasilSurveySerializer(dt)
        #             else:
        #                 serializer = btsSerializer(json_ret)
        #             result=serializer.data
        #             json_dict["data"].append(result)
        #         #print(json_dict["kodeSurvey"],len(json_dict["data"]))
        #         json_ret.append(json_dict)

        # #if jenis.upper() == 'AI':
        # #    serializer = hasilSurveySerializer(json_ret)
        # #else:
        # #    serializer = btsSerializer(json_ret)#,many=True
        # #result=serializer.data
        # return Response.ok(
        #     values=json_ret,
        #     message=f'{len(json_ret)} Data'
        # )

    except Exception as e:
        print(e)
        return HttpResponse(e)
    # else:
    #    return HttpResponse('Post Only')


def getsurveybyprovinsiai(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            provinsi = body_data.get('provinsi').upper()

            pipeline = [
                {
                    '$lookup': {
                        'from': 'penugasan',
                        'localField': 'kodeHasilSurvey',
                        'foreignField': 'kode',
                        'as': 'penugasan'
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan'
                    }
                }, {
                    '$unwind': {
                        'path': '$status'
                    }
                }, {
                    '$lookup': {
                        'from': 'lokasi_survey',
                        'localField': 'penugasan.lokasisurvey',
                        'foreignField': '_id',
                        'as': 'lokasiSurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasiSurvey'
                    }
                }, {
                    '$lookup': {
                        'from': 'provinsi',
                        'localField': 'lokasiSurvey.provinsi',
                        'foreignField': '_id',
                        'as': 'provinsi'
                    }
                }, {
                    '$unwind': {
                        'path': '$provinsi'
                    }
                }, {
                    '$lookup': {
                        'from': 'kabupaten',
                        'localField': 'lokasiSurvey.kabupaten',
                        'foreignField': '_id',
                        'as': 'kabupaten'
                    }
                }, {
                    '$unwind': {
                        'path': '$kabupaten'
                    }
                }, {
                    '$lookup': {
                        'from': 'kecamatan',
                        'localField': 'lokasiSurvey.kecamatan',
                        'foreignField': '_id',
                        'as': 'kecamatan'
                    }
                }, {
                    '$unwind': {
                        'path': '$kecamatan'
                    }
                }, {
                    '$lookup': {
                        'from': 'desa',
                        'localField': 'lokasiSurvey.desa',
                        'foreignField': '_id',
                        'as': 'desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan.status'
                    }
                }, {
                    '$project': {
                        '_id': False,
                        'kode survey': '$kodeHasilSurvey',
                        'nomorsurvey': '$nomorSurvey',
                        'provinsi': '$provinsi.name',
                        'kabupaten': '$kabupaten.name',
                        'kecamatan': '$kecamatan.name',
                        'desa': '$desa.name',
                        'longitude': '$longitude',
                        'latitude': '$latitude',
                        'nama PIC': '$pic.namaPic',
                        'phone PIC': '$pic.phonePic',
                        'status penugasan': '$penugasan.status.status',
                        'survey status': '$status.status',
                        'survey tgl': '$status.tanggal_pembuatan',
                        'fotoKandidatLahan nama': '$fotoKandidatLahan.nama',
                        'fotoKandidatLahan url': '$fotoKandidatLahan.url',
                        'fotoMarkingGps nama': '$fotoMarkingGps.nama',
                        'fotoMarkingGps url': '$fotoMarkingGps.url',
                        'fotoUtaraTitik nama': '$fotoUtaraTitik.nama',
                        'fotoUtaraTitik url': '$fotoUtaraTitik.url',
                        'fotoTimurTitik nama': '$fotoTimurTitik.nama',
                        'fotoTimurTitik url': '$fotoTimurTitik.url',
                        'fotoSelatanTitik nama': '$fotoSelatanTitik.nama',
                        'fotoSelatanTitik url': '$fotoSelatanTitik.url',
                        'fotoBaratTitik nama': '$fotoBaratTitik.nama',
                        'fotoBaratTitik url': '$fotoBaratTitik.url'
                    }
                }, {
                    '$match': {
                        'status penugasan': 'assigned',
                        'survey status': 'Submitted',
                        'provinsi': provinsi,
                        'nomorsurvey': '1',
                    }
                }
            ]

            agg_cursor = hasilSurvey.objects.aggregate(*pipeline)
            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message=provinsi+' tidak Ditemukan')
            """
            search = [x for x in agg_cursor]
            if len(search) > 0:
                return Response.ok(values=search, message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message=provinsi+' tidak Ditemukan')
            """
    except Exception as e:
        return Response.badRequest(message=str(e))


def getsurveybyprovinsibts(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            provinsi = body_data.get('provinsi').upper()

            pipeline = [
                {
                    '$lookup': {
                        'from': 'penugasan',
                        'localField': 'kodeHasilSurvey',
                        'foreignField': 'kode',
                        'as': 'penugasan'
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan'
                    }
                }, {
                    '$unwind': {
                        'path': '$status'
                    }
                }, {
                    '$lookup': {
                        'from': 'lokasi_survey',
                        'localField': 'penugasan.lokasisurvey',
                        'foreignField': '_id',
                        'as': 'lokasiSurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasiSurvey'
                    }
                }, {
                    '$lookup': {
                        'from': 'provinsi',
                        'localField': 'lokasiSurvey.provinsi',
                        'foreignField': '_id',
                        'as': 'provinsi'
                    }
                }, {
                    '$unwind': {
                        'path': '$provinsi'
                    }
                }, {
                    '$lookup': {
                        'from': 'kabupaten',
                        'localField': 'lokasiSurvey.kabupaten',
                        'foreignField': '_id',
                        'as': 'kabupaten'
                    }
                }, {
                    '$unwind': {
                        'path': '$kabupaten'
                    }
                }, {
                    '$lookup': {
                        'from': 'kecamatan',
                        'localField': 'lokasiSurvey.kecamatan',
                        'foreignField': '_id',
                        'as': 'kecamatan'
                    }
                }, {
                    '$unwind': {
                        'path': '$kecamatan'
                    }
                }, {
                    '$lookup': {
                        'from': 'desa',
                        'localField': 'lokasiSurvey.desa',
                        'foreignField': '_id',
                        'as': 'desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan.status'
                    }
                }, {
                    '$project': {
                        '_id': False,
                        'kode survey': '$kodeHasilSurvey',
                        'nomorsurvey': '$nomorSurvey',
                        'provinsi': '$provinsi.name',
                        'kabupaten': '$kabupaten.name',
                        'kecamatan': '$kecamatan.name',
                        'desa': '$desa.name',
                        'longitude': '$longitude',
                        'latitude': '$latitude',
                        'nama PIC': '$pic.namaPic',
                        'phone PIC': '$pic.phonePic',
                        'status penugasan': '$penugasan.status.status',
                        'survey status': '$status.status',
                        'survey tgl': '$status.tanggal_pembuatan',
                        'fotoKandidatLahan nama': '$fotoKandidatLahan.nama',
                        'fotoKandidatLahan url': '$fotoKandidatLahan.url',
                        'fotoMarkingGps nama': '$fotoMarkingGps.nama',
                        'fotoMarkingGps url': '$fotoMarkingGps.url',
                        'fotoUtaraTitik nama': '$fotoUtaraTitik.nama',
                        'fotoUtaraTitik url': '$fotoUtaraTitik.url',
                        'fotoTimurTitik nama': '$fotoTimurTitik.nama',
                        'fotoTimurTitik url': '$fotoTimurTitik.url',
                        'fotoSelatanTitik nama': '$fotoSelatanTitik.nama',
                        'fotoSelatanTitik url': '$fotoSelatanTitik.url',
                        'fotoBaratTitik nama': '$fotoBaratTitik.nama',
                        'fotoBaratTitik url': '$fotoBaratTitik.url'
                    }
                }, {
                    '$match': {
                        'status penugasan': 'assigned',
                        'survey status': 'Submitted',
                        'provinsi': provinsi,
                        'nomorsurvey': '1',
                    }
                }
            ]

            agg_cursor = hasilSurveybts.objects.aggregate(*pipeline)
            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message=provinsi+' tidak Ditemukan')
            """
            search = [x for x in agg_cursor]
            if len(search) > 0:
                return Response.ok(values=search, message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message=provinsi+' tidak Ditemukan')
            """
    except Exception as e:
        return Response.badRequest(message=str(e))


def getsurveyorsubmitai(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            surveyor = body_data.get('surveyor')
            page = int(body_data.get('page', 0)) - 1
            skip = []
            if page >= 0:
                skip = [{'$skip': 20 * page},
                        {'$limit': 20}]
            pipeline = [
                {
                    '$addFields': {
                        'lastStatus': {
                            '$arrayElemAt': ['$status', -1]
                        }
                    }
                },
                # {
                #     '$match': {
                #         'size_of_status': 1
                #     }
                # },
                # {
                #     '$unwind': {
                #         'path': '$status',
                #         'preserveNullAndEmptyArrays': True
                #     }
                # },
                {
                    '$match': {
                        # ,
                        'lastStatus.status': {'$in': ['Submitted', 'Reviewed']}
                        # 'status.status': 'Reviewed'
                    }
                },
                # {
                #     '$unwind': {
                #         'path': '$issue',
                #         'preserveNullAndEmptyArrays': True
                #     }
                # },
                # {
                #     '$match': {
                #         'issue': {
                #             '$exists': False
                #         }
                #     }
                # }
                {
                    '$addFields': {
                        'size_of_issue': {
                            "$size": {"$ifNull": ["$issue", []]}
                        }
                    }
                },
                {
                    '$match': {
                        'size_of_issue': 0
                    }
                }, {
                    '$project': {
                        'size_of_issue': 0,
                        'lastStatus': 0,
                    }
                },
                {
                    '$lookup': {
                        'from': 'user_info',
                        'localField': 'user',
                        'foreignField': '_id',
                        'as': 'user'
                    }
                }, {
                    '$unwind': {
                        'path': '$user',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'surveyor',
                        'localField': 'user.organization',
                        'foreignField': '_id',
                        'as': 'user.organization'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.organization',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$match': {
                        'user.organization._id': ObjectId(surveyor)
                    }
                }, {
                    '$group': {
                        '_id': '$kodeHasilSurvey',
                        'data': {
                            '$push': '$$ROOT'
                        }
                    }
                }, {
                    '$lookup': {
                        'from': 'penugasan',
                        'localField': '_id',
                        'foreignField': 'kode',
                        'as': 'penugasan'
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'lokasi_survey',
                        'localField': 'penugasan.lokasisurvey',
                        'foreignField': '_id',
                        'as': 'lokasisurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasisurvey',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'provinsi',
                        'localField': 'lokasisurvey.provinsi',
                        'foreignField': '_id',
                        'as': 'lokasi.provinsi'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.provinsi',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kabupaten',
                        'localField': 'lokasisurvey.kabupaten',
                        'foreignField': '_id',
                        'as': 'lokasi.kabupaten'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kabupaten',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kota',
                        'localField': 'lokasisurvey.kota',
                        'foreignField': '_id',
                        'as': 'lokasi.kota'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kota',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kecamatan',
                        'localField': 'lokasisurvey.kecamatan',
                        'foreignField': '_id',
                        'as': 'lokasi.kecamatan'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kecamatan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'desa',
                        'localField': 'lokasisurvey.desa',
                        'foreignField': '_id',
                        'as': 'lokasi.desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.desa',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$addFields': {
                        'last': {
                            '$arrayElemAt': ['$data', -1]
                        },
                    }
                }, {
                    '$sort': {
                        'last.tanggal_pembaruan': -1
                    }
                }, {
                    '$project': {
                        '_id': 0,
                        'kode': '$_id',
                        'lokasi': '$lokasi',
                        'data': '$data'
                    }
                }
            ]
            pipe = pipeline + skip
            agg_cursor = hasilSurvey.objects.aggregate(*pipe)

            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                # return Response.badRequest(message='Data tidak Ditemukan')
                return Response.ok(message='Data tidak Ditemukan')
            """
            dataFrameHasil = pandas.DataFrame(agg_cursor)
            if len(dataFrameHasil.index)>0:
                Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
                return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
            else:
                return Response.ok(message='Data tidak Ditemukan')


            # users = list(UserInfo.objects(organization=ObjectId(surveyor)))
            # hasil_survey = hasilSurvey.objects.filter(status__1__exists=False,status__0__status='Submitted',user__in=users)
            # json_ret=[]
            # for k,v in groupby(hasil_survey,key=lambda x:x['kodeHasilSurvey'].strip()):
            #     items=[]
            #     for dt in list(v):
            #         serializer = hasilSurveySerializer(dt)
            #         result=serializer.data
            #         items.append(result)
            #     json_ret.append(items)
            # if len(json_ret) > 0:
            #     #serializer = hasilSurveySerializer(hasil_survey,many=True)
            #     #result=serializer.data
            #     return Response.ok(values=json_ret, message=str(len(json_ret))+' Buah Data telah terambil')
            # else:
            #     return Response.badRequest(message='Data tidak Ditemukan')
            """
    except Exception as e:
        return Response.badRequest(message=str(e))


def getsurveyorsubmitbts(request):
    try:
        if request.method == "POST":
            body_data = request.POST.dict()
            surveyor = body_data.get('surveyor')
            page = int(body_data.get('page', 0)) - 1
            skip = []
            if page >= 0:
                skip = [{'$skip': 20 * page},
                        {'$limit': 20}]
            pipeline = [
                # {
                #     '$addFields': {
                #         'size_of_status': {
                #             '$size': '$status'
                #         }
                #     }
                # }, {
                #     '$match': {
                #         'size_of_status': 1
                #     }
                # }, {
                #     '$unwind': {
                #         'path': '$status',
                #         'preserveNullAndEmptyArrays': True
                #     }
                # },
                {
                    '$addFields': {
                        'lastStatus': {
                            '$arrayElemAt': ['$status', -1]
                        }
                    }
                },
                {
                    '$match': {
                        # 'status.status': 'Submitted'
                        'lastStatus.status': {'$in': ['Submitted', 'Reviewed']}
                    }
                },
                # {
                #     '$unwind': {
                #         'path': '$issue',
                #         'preserveNullAndEmptyArrays': True
                #     }
                # },
                {
                    '$addFields': {
                        'size_of_issue': {
                            "$size": {"$ifNull": ["$issue", []]}
                        }
                    }
                },
                {
                    '$match': {
                        'size_of_issue': 0
                    }
                }               # {
                #     '$match': {
                #         'issue': {
                #             '$exists': False
                #         }
                #     }
                # }
                , {
                    '$project': {
                        'size_of_issue': 0,
                        'lastStatus': 0,
                    }
                }, {
                    '$lookup': {
                        'from': 'user_info',
                        'localField': 'user',
                        'foreignField': '_id',
                        'as': 'user'
                    }
                }, {
                    '$unwind': {
                        'path': '$user',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'surveyor',
                        'localField': 'user.organization',
                        'foreignField': '_id',
                        'as': 'user.organization'
                    }
                }, {
                    '$unwind': {
                        'path': '$user.organization',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$match': {
                        'user.organization._id': ObjectId(surveyor)
                    }
                }, {
                    '$group': {
                        '_id': '$kodeHasilSurvey',
                        'data': {
                            '$push': '$$ROOT'
                        }
                    }
                }, {
                    '$unwind': {
                        'path': '$penugasan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'lokasi_survey',
                        'localField': 'penugasan.lokasisurvey',
                        'foreignField': '_id',
                        'as': 'lokasisurvey'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasisurvey',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'provinsi',
                        'localField': 'lokasisurvey.provinsi',
                        'foreignField': '_id',
                        'as': 'lokasi.provinsi'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.provinsi',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kabupaten',
                        'localField': 'lokasisurvey.kabupaten',
                        'foreignField': '_id',
                        'as': 'lokasi.kabupaten'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kabupaten',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kota',
                        'localField': 'lokasisurvey.kota',
                        'foreignField': '_id',
                        'as': 'lokasi.kota'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kota',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'kecamatan',
                        'localField': 'lokasisurvey.kecamatan',
                        'foreignField': '_id',
                        'as': 'lokasi.kecamatan'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.kecamatan',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$lookup': {
                        'from': 'desa',
                        'localField': 'lokasisurvey.desa',
                        'foreignField': '_id',
                        'as': 'lokasi.desa'
                    }
                }, {
                    '$unwind': {
                        'path': '$lokasi.desa',
                        'preserveNullAndEmptyArrays': True
                    }
                }, {
                    '$addFields': {
                        'last': {
                            '$arrayElemAt': [
                                '$data', -1
                            ]
                        }
                    }
                }, {
                    '$sort': {
                        'last.tanggal_pembaruan': -1
                    }
                }, {
                    '$project': {
                        '_id': 0,
                        'kode': '$_id',
                        'lokasi': '$lokasi',
                        'data': '$data'
                    }
                }
            ]
            pipe = pipeline + skip
            agg_cursor = hasilSurveybts.objects.aggregate(*pipe)

            search = list(agg_cursor)
            if len(search) > 0:
                return Response.ok(values=json.loads(json.dumps(search, default=str)), message=str(len(search))+' Buah Data telah terambil')
            else:
                return Response.badRequest(message='Data tidak Ditemukan')
            """
            dataFrameHasil = pandas.DataFrame(agg_cursor)
            if len(dataFrameHasil.index)>0:
                Hasil = dataFrameHasil.to_json(orient='records', default_handler=str)
                return Response.ok(values=json.loads(Hasil), message=str(len(dataFrameHasil.index))+' Buah Data telah terambil')
            else:
                return Response.ok(message='Data tidak Ditemukan')


            # users = list(UserInfo.objects(organization=ObjectId(surveyor)))
            # hasil_survey = hasilSurveybts.objects.filter(status__1__exists=False,status__0__status='Submitted',user__in=users)
            # json_ret=[]
            # for k,v in groupby(hasil_survey,key=lambda x:x['kodeHasilSurvey'].strip()):
            #     items=[]
            #     for dt in list(v):
            #         serializer = btsSerializer(dt)
            #         result=serializer.data
            #         items.append(result)
            #     json_ret.append(items)
            # if len(json_ret) > 0:
            #     #serializer = hasilSurveySerializer(hasil_survey,many=True)
            #     #result=serializer.data
            #     return Response.ok(values=json_ret, message=str(len(json_ret))+' Buah Data telah terambil')
            # else:
            #     return Response.badRequest(message='Data tidak Ditemukan')
            """
    except Exception as e:
        return Response.badRequest(message=str(e))


def countPenugasanSurveyor(request):
    try:
        def getPenugasanAI(surveyor):
            hasil_survey = Penugasan.objects.filter(
                jenissurvey='5f16b4ba149882a98fc6655e', surveyor=ObjectId(surveyor)).count()
            return hasil_survey

        def getPenugasanBTS(surveyor):
            hasil_survey = Penugasan.objects.filter(
                jenissurvey='5f1521524f9c6764c713d73c', surveyor=ObjectId(surveyor)).count()
            return hasil_survey

        def getPenugasanAISurvey(users):
            hasil_survey = hasilSurvey.objects.filter(
                status__status="Submitted", nomorSurvey='1', user__in=users)
            json_ret = []
            json_issue = []
            for k in hasil_survey:
                json_dict = {}
                json_dict["kodeSurvey"] = k
                json_ret.append(json_dict)
                if len(k.issue) > 0:
                    json_issue.append(1)

            return len(json_ret), len(json_issue)

        def getPenugasanBTSSurvey(users):
            hasil_survey = hasilSurveybts.objects.filter(
                status__status="Submitted", nomorSurvey='1', user__in=users)
            json_ret = []
            json_issue = []
            for k in hasil_survey:
                json_dict = {}
                json_dict["kodeSurvey"] = k
                json_ret.append(json_dict)
                if len(k.issue) > 0:
                    json_issue.append(1)

            return len(json_ret), len(json_issue)

        body_data = request.POST.dict()
        surveyor = body_data.get('surveyor')
        users = list(UserInfo.objects(organization=ObjectId(surveyor)))
        AISurvey, AISurveyIssue = getPenugasanAISurvey(users)
        BTSSurvey, BTSSurveyIssue = getPenugasanBTSSurvey(users)
        return Response.ok(
            values={
                "penugasan_ai": getPenugasanAI(surveyor),
                "penugasan_bts": getPenugasanBTS(surveyor),
                "penugasan_ai_surveyed": AISurvey,
                "penugasan_bts_surveyed": BTSSurvey,
                "penugasan_ai_issue": AISurveyIssue,
                "penugasan_bts_issue": BTSSurveyIssue,
            },
            message='Success'
        )
    except Exception as e:
        return Response.badRequest(
            values={},
            message=str(e)
        )


def getpenugasansurveyor_ori(request):
    try:
        # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
        # ret,user = authenticate_credentials(token)
        # if False == ret or None ==user:
        #    return JsonResponse({"state":"fail"})
        strjson = json.loads(request.body.decode("utf-8"))
        #user = strjson["user"]
        field = strjson["field"].lower()
        value = strjson["value"].lower()
        jenis = strjson["jenis"]
        surveyor = strjson["surveyor"]
        page = (int(strjson.get('page', 0)) - 1)
        result = []

        def getByStatus():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
            except JenisSurvey.DoesNotExist:
                return {}
            data = Penugasan.objects(
                status__status=value, jenissurvey=ObjectId(data_jenis.id), surveyor=ObjectId(surveyor))
            #serializer = PenugasanSerializer(data, many=True)
            # result=serializer.data
            # for _penugasan in data:
            #    result.append(_penugasan.serialize())
            # return result

        def getByJenis():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=value.upper())
            except JenisSurvey.DoesNotExist:
                return {}
            data = Penugasan.objects(jenissurvey=ObjectId(
                data_jenis.id), surveyor=ObjectId(surveyor))
            #serializer = PenugasanSerializer(data, many=True)
            # result=serializer.data
            result = []
            if page < 0:
                for _penugasan in data:
                    result.append(_penugasan.serialize())
            else:
                if page < len(data):
                    if (len(data) - page) < 20:
                        endrow = page + (len(data) - page)
                    else:
                        endrow = page + 20
                    for i in range(page, endrow):
                        result.append(data[i].serialize())
            return result

        def getByUser():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
            except JenisSurvey.DoesNotExist:
                return {}
            data = Penugasan.objects(user=ObjectId(
                value), jenissurvey=ObjectId(data_jenis.id), surveyor=ObjectId(surveyor))
            result = []
            if page < 0:
                for _penugasan in data:
                    result.append(_penugasan.serialize())
            else:
                if page < len(data):
                    if (len(data) - page) < 20:
                        endrow = page + (len(data) - page)
                    else:
                        endrow = page + 20
                    for i in range(page, endrow):
                        result.append(data[i].serialize())
            return result

        def getByField():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
            except JenisSurvey.DoesNotExist:
                return {}
                # return Response.badRequest(
                #    values='null',
                #    message='jenissurvey not found'
                # )
            if value.upper() == 'ALL':
                data = Penugasan.objects.filter(
                    jenissurvey=ObjectId(data_jenis.id), surveyor=ObjectId(surveyor))
            else:
                data = Penugasan.objects.filter(
                    **{field: value}, jenissurvey=ObjectId(data_jenis.id), surveyor=ObjectId(surveyor))
            result = []
            if page < 0:
                # for _penugasan in data:
                #    result.append(_penugasan.serialize())
                serializer = PenugasanSerializer(data, many=True)
                result = serializer.data
            else:
                if page < len(data):
                    if (len(data) - page) < 20:
                        endrow = page + (len(data) - page)
                    else:
                        endrow = page + 20
                    for i in range(page, endrow):
                        result.append(PenugasanSerializer(data[i]).data)
            return result

        switcher = {
            "status": getByStatus,
            "jenissurvey": getByJenis,
            "user": getByUser,
        }

        _result = switcher.get(field, getByField)

        if not _result:
            return Response.badRequest(
                message='Wrong Field Parameter'
            )

        result = _result()
        return Response.ok(
            values=result,
            message=f'{len(result)} Data'
        )
    except Exception as e:
        print(e)
        return HttpResponse(e)


def getpenugasansurveyor(request):
    try:
        # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
        # ret,user = authenticate_credentials(token)
        # if False == ret or None ==user:
        #    return JsonResponse({"state":"fail"})
        strjson = json.loads(request.body.decode("utf-8"))
        #user = strjson["user"]
        field = strjson["field"].lower()
        value = strjson["value"].lower()
        jenis = strjson["jenis"]
        surveyor = strjson["surveyor"]
        page = int(strjson.get('page', 0)) - 1
        skip = []
        if page >= 0:
            skip = [{'$skip': 20 * page},
                    {'$limit': 20}]
        result = []

        pipeline = [
            {
                '$lookup': {
                    'from': 'user_info',
                    'localField': 'user',
                    'foreignField': '_id',
                    'as': 'user'
                }
            }, {
                '$unwind': {
                    'path': '$user',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'user_info',
                    'localField': 'assignfrom1',
                    'foreignField': '_id',
                    'as': 'assignfrom1'
                }
            }, {
                '$unwind': {
                    'path': '$assignfrom1',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'user_info',
                    'localField': 'assignto1',
                    'foreignField': '_id',
                    'as': 'assignto1'
                }
            }, {
                '$unwind': {
                    'path': '$assignto1',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'document_penugasan',
                    'localField': 'spk',
                    'foreignField': '_id',
                    'as': 'spk'
                }
            }, {
                '$unwind': {
                    'path': '$spk',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'jenis_survey',
                    'localField': 'jenissurvey',
                    'foreignField': '_id',
                    'as': 'jenissurvey'
                }
            }, {
                '$unwind': {
                    'path': '$jenissurvey',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'surveyor',
                    'localField': 'surveyor',
                    'foreignField': '_id',
                    'as': 'surveyor'
                }
            }, {
                '$unwind': {
                    'path': '$surveyor',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'lokasi_survey',
                    'localField': 'lokasisurvey',
                    'foreignField': '_id',
                    'as': 'lokasisurvey'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'provinsi',
                    'localField': 'lokasisurvey.provinsi',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.provinsi'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.provinsi',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kabupaten',
                    'localField': 'lokasisurvey.kabupaten',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.kabupaten'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.kabupaten',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kota',
                    'localField': 'lokasisurvey.kota',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.kota'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.kota',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'kecamatan',
                    'localField': 'lokasisurvey.kecamatan',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.kecamatan'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.kecamatan',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$lookup': {
                    'from': 'desa',
                    'localField': 'lokasisurvey.desa',
                    'foreignField': '_id',
                    'as': 'lokasisurvey.desa'
                }
            }, {
                '$unwind': {
                    'path': '$lokasisurvey.desa',
                    'preserveNullAndEmptyArrays': True
                }
            }, {
                '$addFields': {
                    'last': {
                        '$arrayElemAt': ['$status', -1]
                    },
                }
            }, {
                '$sort': {
                    'last.date': -1
                }
            }, {
                '$project': {
                    'surveyor.jenissurvey': 0,
                    'last': 0
                }
            }
        ]

        def getByStatus():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
            except JenisSurvey.DoesNotExist:
                return {}

            match = [
                {
                    '$match': {
                        'jenissurvey': ObjectId(data_jenis.id),
                        'surveyor': ObjectId(surveyor)
                    }
                }, {
                    '$addFields': {
                        'last_status': {
                            '$arrayElemAt': ['$status', -1]
                        }
                    }
                }, {
                    '$match': {
                        'last_status.status': value
                    }
                }, {
                    '$project': {
                        'last_status': 0
                    }
                }]
            pipe = match + pipeline + skip

            agg_cursor = Penugasan.objects.aggregate(*pipe)
            result = list(agg_cursor)
            return result

        def getByUser():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
            except JenisSurvey.DoesNotExist:
                return {}

            match = [{
                '$match': {
                    'jenissurvey': ObjectId(data_jenis.id),
                    'surveyor': ObjectId(surveyor),
                    user: ObjectId(value)}
            }]

            pipe = match + pipeline + skip

            agg_cursor = Penugasan.objects.aggregate(*pipe)
            result = list(agg_cursor)

            return result

        def getByField():
            try:
                data_jenis = JenisSurvey.objects.get(jenis=jenis.upper())
            except JenisSurvey.DoesNotExist:
                return {}

            if value.upper() == 'ALL':
                match = [{
                    '$match': {
                        'jenissurvey': ObjectId(data_jenis.id),
                        'surveyor': ObjectId(surveyor)
                    }
                }]

            else:
                if field.lower() == 'assignto1':
                    value_ = ObjectId(value)
                else:
                    value_ = value
                # data = Penugasan.objects.filter(
                #    **{c: value}, jenissurvey=ObjectId(data_jenis.id), surveyor=ObjectId(surveyor))
                match = [{
                    '$match': {
                        'jenissurvey': ObjectId(data_jenis.id),
                        'surveyor': ObjectId(surveyor),
                        field: value_
                    }
                }]
            pipe = match + pipeline + skip
            agg_cursor = Penugasan.objects.aggregate(*pipe)
            result = list(agg_cursor)

            return result

        switcher = {
            "status": getByStatus,
            "user": getByUser,
        }

        _result = switcher.get(field, getByField)

        if not _result:
            return Response.badRequest(
                message='Wrong Field Parameter'
            )

        result = _result()
        return Response.ok(
            values=json.loads(json.dumps(result, default=str)),
            message=f'{len(result)} Data'
        )
    except Exception as e:
        print(e)
        return HttpResponse(e)
