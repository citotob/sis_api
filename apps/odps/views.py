from django.shortcuts import render
from odps.models import *
from sites.models import kabupaten, kota
from userinfo.models import vendor
from odps.serializer import *
from .response import Response
import json
from operator import itemgetter

from geojson import Feature, Point
from turfpy.measurement import distance, rhumb_distance, boolean_point_in_polygon
from turfpy.transformation import circle

import openpyxl

from bson import ObjectId

def uploadodp(request):
    if request.method == 'POST':
        
        lokasi_gagal = ''

        odp_file = request.FILES["odp_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(odp_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row

        for row in worksheet.iter_rows():
            #data_odp1 = Odp.objects.filter(latitude=str(row[6].value).replace(',','.'),
            #            longitude=str(row[5].value).replace(',','.'))
            #if data_odp1:
            #    continue
            vndr=str(row[7].value)
            if "TELKOM" in vndr:
                vndr = "TELKOM"
            try:
                #data_vendor = vendor.objects.get(
                #    name__iexact=str(row[7].value))
                data_vendor = vendor.objects.get(
                    name__iexact=vndr)
            except vendor.DoesNotExist:
                data_vendor = vendor(
                    name=str(row[7].value),
                    latitude='0',
                    longitude='0',
                    longlat=[0, 0],
                )
                data_vendor.save()

                #data_VPScore = VPScore(vendor=data_vendor.id)
                #data_VPScore.save()

            lanjut = True
            if str(row[1].value) == 'None':
                break
            if str(row[0].value) == 'NAMA LOKASI':
                continue
            tekno = str(row[9].value)
            if "VSAT" in str(row[9].value):
                tekno = "VSAT"

            try:
                data_odp = Odp(
                    latitude=str(row[6].value).replace(',','.'),
                    longitude=str(row[5].value).replace(',','.'),
                    longlat=[float(str(row[5].value).replace(',','.')), float(str(row[6].value).replace(',','.'))],
                    teknologi=tekno,
                    nama=str(row[0].value),
                    desa_kelurahan=ObjectId(str(row[4].value)),
                    kecamatan=ObjectId(str(row[3].value)),
                    provinsi=ObjectId(str(row[1].value)),
                    vendorid=data_vendor.id,
                )

                try:
                    data_kab_kot = kabupaten.objects.get(id=ObjectId(str(row[2].value)))
                    data_odp.kabupaten = data_kab_kot.id
                except kabupaten.DoesNotExist:
                    try:
                        data_kab_kot = kota.objects.get(id=ObjectId(str(row[2].value)))
                        data_odp.kota = data_kab_kot.id
                    except kota.DoesNotExist:
                        #return Response.ok(
                        #    values=[],
                        #    message="kabkot " + str(row[2].value) +" tidak ada"
                        #)
                        return Response().base(
                            success=False,
                            values=[],
                            message="kabkot " + str(row[2].value) +" tidak ada",
                            status=404
                        )
                
                data_odp.save()
            except:
                continue

        return Response.ok(
            values=[],
            message="OK"
        )

def uploadodp1(request):
    if request.method == 'POST':
        
        lokasi_gagal = ''

        odp_file = request.FILES["odp_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(odp_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["AKSES INTERNET - 27092020"]

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        id_gagal = []
        for row in worksheet.iter_rows():
            lanjut = True
            if str(row[1].value) == 'None':
                break
            if str(row[0].value).upper() == 'NO URUT':
                continue
            try:
                data_vendor = vendor.objects.get(
                    name__iexact=str(row[8].value).strip())
                #data_vendor = vendor.objects.get(
                #    name__iexact=vndr)
            except vendor.DoesNotExist:
                data_vendor = vendor(
                    name=str(row[8].value),
                    latitude='0',
                    longitude='0',
                    longlat=[0, 0],
                )
                data_vendor.save()

            tekno = str(row[10].value).strip()
            if "VSAT" in str(row[10].value):
                tekno = "VSAT"
            else:
                if "RADIO" in str(row[10].value):
                    tekno = "RL"
                else:
                    if "FIBER" in str(row[10].value):
                        tekno = "FO"

            data_prov = provinsi.objects.filter(
                name=str(row[2].value).strip()).first()
            if not data_prov:
                """
                json_dict = {}
                json_dict["No Urut"] = str(row[0].value).strip()
                json_dict["provinsi"] = str(row[2].value).strip()
                id_gagal.append(json_dict)
                continue
                """
                data_prov = provinsi(
                    name=str(row[2].value).upper()
                )
                data_prov.save()

            data_kab = kabupaten.objects.filter(
                name='KAB. '+str(row[3].value).strip(), provinsi=data_prov.id).first()
            if not data_kab:
                data_kota = kota.objects.filter(
                    name='KOTA '+str(row[3].value).strip(), provinsi=data_prov.id).first()
                if not data_kota:
                    """
                    json_dict = {}
                    json_dict["No Urut"] = str(row[0].value).strip()
                    json_dict["provinsi_id"] = data_prov.id
                    json_dict["provinsi"] = data_prov.name
                    json_dict["kab_kota"] = str(row[3].value).strip()
                    id_gagal.append(json_dict)
                    continue
                    """
                    data_kab = kabupaten(
                        name=str(row[3].value).strip().upper(),
                        provinsi=ObjectId(data_prov.id)
                    )
                    data_kab.save()
            
            if data_kab:
                data_kec = kecamatan.objects.filter(
                    name=str(row[4].value).strip(),kabupaten=data_kab.id).first()
                kab_kot_id = data_kab.id
                kab_kot_name = data_kab.name
            else:
                data_kec = kecamatan.objects.filter(
                    name=str(row[4].value).strip(),kota=data_kota.id).first()
                kab_kot_id = data_kota.id
                kab_kot_name = data_kota.name
            if not data_kec:
                """
                json_dict = {}
                json_dict["No Urut"] = str(row[0].value).strip()
                json_dict["kab_kota_id"] = kab_kot_id
                json_dict["kab_kota"] = kab_kot_name
                json_dict["kecamatan"] = str(row[4].value).strip()
                id_gagal.append(json_dict)
                continue
                """
                if data_kab:
                    data_kec = kecamatan(
                        name=str(row[4].value).strip().upper(),
                        kabupaten=ObjectId(kab_kot_id)
                    )
                    data_kec.save()
                else:
                    data_kec = kecamatan(
                        name=str(row[4].value).strip().upper(),
                        kota=ObjectId(kab_kot_id)
                    )
                    data_kec.save()

            data_desa = desa.objects.filter(
                name=str(row[5].value).strip(),kecamatan=data_kec.id).first()
            if not data_desa:
                """
                json_dict = {}
                json_dict["No Urut"] = str(row[0].value).strip()
                json_dict["kecamatan_id"] = data_kec.id
                json_dict["kecamatan"] = data_kec.name
                json_dict["desa"] = str(row[5].value).strip()
                id_gagal.append(json_dict)
                continue
                """
                data_desa = desa(
                    name=str(row[5].value).strip().upper(),
                    kecamatan=ObjectId(data_kec.id)
                )
                data_desa.save()

            try:
                create_date = datetime.strptime(
                    str(row[13].value), '%Y-%m-%d 00:00:00')
            except:
                json_dict = {}
                json_dict["No Urut"] = str(row[0].value).strip()
                json_dict["tanggal"] = str(row[13].value)
                id_gagal.append(json_dict)
                continue
            try:
                data_odp = Odp_backup(
                    latitude=str(row[7].value).replace(',','.'),
                    longitude=str(row[6].value).replace(',','.'),
                    longlat=[float(str(row[6].value).replace(',','.')), float(str(row[7].value).replace(',','.'))],
                    teknologi=tekno,
                    nama=str(row[1].value).strip(),
                    desa_kelurahan=data_desa.name,
                    kecamatan=data_kec.name,
                    provinsi=data_prov.name,
                    vendorid=data_vendor.name,
                    created_at=create_date,
                    updated_at=create_date
                )

                if data_kab:
                    data_odp.kabupaten=data_kab.name
                else:
                    data_odp.kota=data_kota.name
                data_odp.save()
            except Exception as e:
                json_dict = {}
                json_dict["No Urut"] = str(row[0].value).strip()
                json_dict["error"] = str(e)
                id_gagal.append(json_dict)
                continue
        #print(id_gagal)
        return Response.ok(
            values=json.loads(json.dumps(id_gagal, default=str)),
            message="OK"
        )

def uploadbts(request):
    if request.method == 'POST':
        
        lokasi_gagal = ''

        odp_file = request.FILES["bts_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(odp_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        id_gagal = []
        for row in worksheet.iter_rows():
            lanjut = True
            if str(row[1].value) == 'None':
                break
            if str(row[0].value).lower() == 'unik_id':
                continue

            data_prov = provinsi.objects.filter(
                name=str(row[4].value).strip()).first()
            if not data_prov:
                
                json_dict = {}
                json_dict["unik_id"] = str(row[0].value).strip()
                json_dict["provinsi"] = str(row[4].value).strip()
                id_gagal.append(json_dict)
                continue
                """
                data_prov = provinsi(
                    name=str(row[2].value).upper()
                )
                data_prov.save()
                """
            data_kab = kabupaten.objects.filter(
                name=str(row[5].value).strip(), provinsi=data_prov.id).first()
            if not data_kab:
                data_kota = kota.objects.filter(
                    name=str(row[5].value).strip(), provinsi=data_prov.id).first()
                if not data_kota:
                    
                    json_dict = {}
                    json_dict["unik_id"] = str(row[0].value).strip()
                    json_dict["provinsi_id"] = data_prov.id
                    json_dict["provinsi"] = data_prov.name
                    json_dict["kab_kota"] = str(row[5].value).strip()
                    id_gagal.append(json_dict)
                    continue
                    """
                    data_kab = kabupaten(
                        name=str(row[3].value).strip().upper(),
                        provinsi=ObjectId(data_prov.id)
                    )
                    data_kab.save()
                    """
            
            if data_kab:
                data_kec = kecamatan.objects.filter(
                    name=str(row[6].value).strip(),kabupaten=data_kab.id).first()
                kab_kot_id = data_kab.id
                kab_kot_name = data_kab.name
            else:
                data_kec = kecamatan.objects.filter(
                    name=str(row[6].value).strip(),kota=data_kota.id).first()
                kab_kot_id = data_kota.id
                kab_kot_name = data_kota.name
            if not data_kec:
                
                json_dict = {}
                json_dict["unik_id"] = str(row[0].value).strip()
                json_dict["provinsi"] = data_prov.name
                json_dict["kab_kota_id"] = kab_kot_id
                json_dict["kab_kota"] = kab_kot_name
                json_dict["kecamatan"] = str(row[6].value).strip()
                id_gagal.append(json_dict)
                continue
                """
                if data_kab:
                    data_kec = kecamatan(
                        name=str(row[4].value).strip().upper(),
                        kabupaten=ObjectId(kab_kot_id)
                    )
                    data_kec.save()
                else:
                    data_kec = kecamatan(
                        name=str(row[4].value).strip().upper(),
                        kota=ObjectId(kab_kot_id)
                    )
                    data_kec.save()
                """
            data_desa = desa.objects.filter(
                name=str(row[7].value).strip(),kecamatan=data_kec.id).first()
            if not data_desa:
                
                json_dict = {}
                json_dict["unik_id"] = str(row[0].value).strip()
                json_dict["provinsi"] = data_prov.name
                json_dict["kab_kota"] = kab_kot_name
                json_dict["kecamatan_id"] = data_kec.id
                json_dict["kecamatan"] = data_kec.name
                json_dict["desa"] = str(row[7].value).strip()
                id_gagal.append(json_dict)
                continue
                """
                data_desa = desa(
                    name=str(row[5].value).strip().upper(),
                    kecamatan=ObjectId(data_kec.id)
                )
                data_desa.save()
                """
            try:
                create_date = datetime.strptime(
                    str(row[11].value)[:19].replace('.',':'), '%Y-%m-%d %H:%M:%S')
            except:
                json_dict = {}
                json_dict["unik_id"] = str(row[0].value).strip()
                json_dict["tanggal_pembuatan"] = str(row[11].value)
                id_gagal.append(json_dict)
                continue
            try:
                access_date = datetime.strptime(
                    str(row[12].value).replace('.',':'), '%Y-%m-%d %H:%M:%S')
            except:
                json_dict = {}
                json_dict["unik_id"] = str(row[0].value).strip()
                json_dict["tanggal_akses"] = str(row[12].value)
                id_gagal.append(json_dict)
                continue
            try:
                data_odp = bts_onair(
                    unik_id=str(row[0].value).strip(),
                    latitude=str(row[9].value).replace(',','.'),
                    longitude=str(row[10].value).replace(',','.'),
                    longlat=[float(str(row[10].value).replace(',','.')), float(str(row[9].value).replace(',','.'))],
                    nama=str(row[8].value).strip(),
                    desa_kelurahan=data_desa.name,
                    kecamatan_name=data_kec.name,
                    provinsi_name=data_prov.name,
                    created_at=create_date,
                    updated_at=create_date,
                    access_date=access_date,
                    desa=data_desa.id,
                    kecamatan=data_kec.id,
                    provinsi=data_prov.id,
                )

                if data_kab:
                    data_odp.kabupaten_name=data_kab.name
                    data_odp.kabupaten=data_kab.id
                else:
                    data_odp.kota_name=data_kota.name
                    data_odp.kota=data_kota.id
                data_odp.save()
            except Exception as e:
                json_dict = {}
                json_dict["unik_id"] = str(row[0].value).strip()
                json_dict["error"] = str(e)
                id_gagal.append(json_dict)
                continue
        #print(id_gagal)
        return Response.ok(
            values=json.loads(json.dumps(id_gagal, default=str)),
            message="OK"
        )

def addodp(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    if request.method == "POST":  # Add
        try:
            body_data = json.loads(request.body)

            longitude = body_data.get('longitude')
            latitude = body_data.get('latitude')
            teknologi = body_data.get('teknologi')
            vendorid = body_data.get('vendorid')

            data_odp = Odp(
                latitude=latitude,
                longitude=longitude,
                longlat=[float(longitude), float(latitude)],
                teknologi=teknologi,
                vendorid=vendorid
            )

            data_odp.save()
            result = []
            # results = site.objects.get(id=ObjectId(data_site.id))

            # result = results.serialize()

            return Response.ok(
                values=result,
                message='Berhasil'
            )
        except Exception as e:
            return Response.badRequest(message=str(e))

    else:
        return Response.badRequest(message='Hanya POST')

def getRecommendTech(request):

    try:

        if not request.body:
            return Response.badRequest(
                values=[],
                message="Need Json Body longitude & latitude"
            )
        body = json.loads(request.body)
        longitude = body.get('longitude', None)
        latitude = body.get('latitude', None)

        if not (longitude and latitude):
            return Response.badRequest(
                values=[],
                message="Need Json Body longitude & latitude"
            )

        coordinates = [float(longitude), float(latitude)]
        # print(getRecommendTechnologi(longitude, latitude))
        start = Feature(geometry=Point(coordinates=coordinates))
        # data = Odp.objects.aggregate([
        #     {
        #         '$match': {
        #             "longlat": {"$geoWithin":
        #                         {"$center": [[121.2866, 39.984], (116 / 111.32)]}}
        #         }
        #     }
        # ])
        data = Odp.objects(
            longlat__geo_within_sphere=[coordinates, (10 / 6378.1)])
        # data = Odp.objects(
        #     longlat__near=[122.2866, -1.14911], longlat__max_distance=115199)
        # radius = circle(center=end, radius=85, units='km')
        # print(boolean_point_in_polygon(start, radius))
        # print(rhumb_distance(start, end, units='km'))
        # print(distance(start, end, units='km'))
        # print(list(data))

        serializer = ODPSerializer(data, many=True)

        if len(serializer.data) > 0:
            results = []
            datas = serializer.data.copy()
            for x in datas:
                end = Feature(geometry=Point(x["longlat"]["coordinates"]))
                x['distance'] = f'{int(distance(start, end, units="km"))} km'
                results.append(x)
            results.sort(key=itemgetter('distance', 'created_at'))
            return Response.ok(
                values=json.loads(json.dumps(serializer.data, default=str)),
                message=f'{len(serializer.data)} Data'
            )
        else:
            #return Response.ok(
            #    values=[],
            #    message='Data tidak ada'
            #)
            return Response().base(
                success=False,
                values=[],
                message='Data tidak ada',
                status=404
            )

    except Exception as e:
        print(e)
        return Response.badRequest(message=str(e))

def getodp(request):
    try:
        req_fields = ['latitude', 'longitude', 'teknologi']
        try:
            start = int(request.GET.get('start')) - 1
            end = int(request.GET.get('end'))

            if start < 0:
                start = 0
        
            data = Odp.objects.all().only(*req_fields)[start:end]
        except:
            data = Odp.objects.all().only(*req_fields)
        
        serializer = siteonairSerializer(data, many=True)
        if len(serializer.data) > 0:
            return Response.ok(
                values=json.loads(json.dumps(serializer.data, default=str)),
                message=f'{len(serializer.data)} Data'
            )
        else:
            #return Response.ok(
            #    values=[],
            #    message='Data tidak ada'
            #)
            return Response().base(
                success=False,
                values=[],
                message='Data tidak ada',
                status=404
            )
    except Exception as e:
        #print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )

