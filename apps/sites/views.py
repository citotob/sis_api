
import os
from django.shortcuts import render
from django.http import JsonResponse
from sites.models import *
#from apps.sites.models import Odp
from vendor.models import *
from userinfo.models import *
from odps.models import desa, kecamatan, kota, kabupaten, provinsi
from userinfo.views import authenticate_credentials
from apps.vendorperformance.serializer import VPSerializer
from vendorperformance.models import *
from apps.userinfo.serializer import VendorScoreSerializer
from django.dispatch import receiver
from django.db.models.signals import post_save
from django.conf import settings
import json
from django.core.exceptions import ObjectDoesNotExist
import pandas
from .utils import getRecommendTechnologi
from .response import Response
from bson import ObjectId, json_util
from django.http import HttpResponse
from operator import itemgetter
# import datetime
from datetime import datetime, timedelta, timezone
from dateutil.relativedelta import relativedelta
from django.core.serializers import serialize
from django.core.files.storage import FileSystemStorage
import requests
from rest_framework.response import Response as RFResponse
from rest_framework import status
from django.core.mail import EmailMultiAlternatives
from django.template.loader import get_template
from django.core import serializers
from .serializer import *
from geojson import Feature, Point
from turfpy.measurement import distance, rhumb_distance, boolean_point_in_polygon
from turfpy.transformation import circle
from itertools import groupby
from userinfo.utils.notification import Notification
import calendar
from math import radians, cos, sin, asin, sqrt

from email.mime.image import MIMEImage
from django.template.loader import get_template
from django.template import Context

from django.db.models import Avg, Max, Min, Sum
from operator import itemgetter
import openpyxl

from notification.utils.CustomNotification import CustomNotification
from publicservice.utils import send_mail
from pathlib import Path
from mongoengine.queryset.visitor import Q


def getLaporan(request):

    if request.method == "POST":

        bulan = ["Januari", "Februari", "Maret", "April", "Mei", "Juni",
                 "Juli", "Agustus", "September", "Oktober", "November", "Desember"]

        try:
            req = request.body.decode("utf-8")
            data = json.loads(req)
            month = data['month']
            year = data['year']

            pastMonth = data['month'] - 1
            pastYear = year

            if pastMonth == 0:
                pastMonth = 12
                pastYear = year-1

            ai = '5f16b4ba149882a98fc6655e'
            bts = '5f1521524f9c6764c713d73c'

            if month is None or year is None:
                return Response.badRequest(message='Need Json Body "month" & "year"')

            def pipelineDate(month, year, jenis):

                lastDay = calendar.monthrange(year=year, month=month)[1]
                return [
                    {
                        '$addFields': {
                            'last': {
                                '$arrayElemAt': [
                                    '$status', -1
                                ]
                            }
                        }
                    }, {
                        '$match': {
                            'jenissurvey': ObjectId(jenis),
                            'last.status': 'finished',
                            'last.date': {
                                '$gte': datetime(year, month, 1, 00, 00, 00, tzinfo=timezone.utc),
                                '$lte': datetime(year, month, lastDay, 23, 59, 59, tzinfo=timezone.utc)
                            }
                        }
                    }, {
                        '$count': 'count'
                    }
                ]

            def pipelineTotal(jenis, status):
                return [
                    {
                        '$match': {
                            'jenissurvey': ObjectId(jenis),
                            'status.status': status,
                        }
                    }, {
                        '$count': 'count'
                    }
                ]

            penugasanAITotal = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=ai, status='assigned')))[0]['count']
            penugasanBTSTotal = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=bts, status='assigned')))[0]['count']

            penugasanAITotalFinish = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=ai, status='finished')))[0]['count']
            penugasanBTSTotalFinish = list(
                Penugasan.objects.aggregate(pipelineTotal(jenis=bts, status='finished')))[0]['count']

            penugasanPersentaseAI = (
                penugasanAITotalFinish/penugasanAITotal) * 100
            penugasanPersentaseBTS = (
                penugasanBTSTotalFinish / penugasanBTSTotal) * 100

            penugasanAISekarangList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=ai, month=month, year=year)))
            penugasanBTSSekarangList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=bts, month=month, year=year)))

            penugasanAISebelumnyaList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=ai, month=pastMonth, year=pastYear)))
            penugasanBTSSebelumnyaList = list(
                Penugasan.objects.aggregate(pipelineDate(jenis=bts, month=pastMonth, year=pastYear)))

            penugasanAISekarang = 0 if len(
                penugasanAISekarangList) == 0 else penugasanAISekarangList[0]['count']
            penugasanBTSSekarang = 0 if len(
                penugasanBTSSekarangList) == 0 else penugasanBTSSekarangList[0]['count']

            penugasanAISebelumnya = 0 if len(
                penugasanAISebelumnyaList) == 0 else penugasanAISebelumnyaList[0]['count']
            penugasanBTSSebelumnya = 0 if len(
                penugasanBTSSebelumnyaList) == 0 else penugasanBTSSebelumnyaList[0]['count']

            persentasiKenaikanAI = ((
                abs(penugasanAISekarang - penugasanAISebelumnya)) / penugasanAISebelumnya) * 100 if penugasanAISebelumnya > 0 else 0
            persentasiKenaikanBTS = ((
                abs(penugasanBTSSekarang - penugasanBTSSebelumnya)) / penugasanBTSSebelumnya) * 100 if penugasanAISebelumnya > 0 else 0

            try:
                subject = f'Monthly Report SMASLAB {bulan[month-1]} {year}'
                text_content = ''
                htmly = get_template('email/executive/executive.html')
                d = {
                    'bulan': f'{bulan[month-1]} {year}',

                    'ai_finish': penugasanAITotalFinish,
                    'ai_total': penugasanAITotal,

                    'bts_finish': penugasanAITotalFinish,
                    'bts_total': penugasanAITotal,

                    'ai_persentase': round(penugasanPersentaseAI, 2),
                    'ai_penambahan': persentasiKenaikanAI,
                    'ai_bulan': penugasanAISekarang,

                    'bts_persentase': round(penugasanPersentaseBTS, 2),
                    'bts_penambahan': persentasiKenaikanBTS,
                    'bts_bulan': penugasanBTSSekarang,

                    'ai_perubahan': 'Kenaikan' if penugasanAISekarang > penugasanAISebelumnya else 'Penurunan' if penugasanAISekarang < penugasanAISebelumnya else 'Tidak ada Perubahan',
                    'ai_persentase_penambahan': f'{round(persentasiKenaikanAI, 2)}%' if penugasanAISekarang != penugasanAISebelumnya else '',

                    'bts_perubahan': 'Kenaikan' if penugasanBTSSekarang > penugasanBTSSebelumnya else 'Penurunan' if penugasanBTSSekarang < penugasanBTSSebelumnya else 'Tidak ada Perubahan',
                    'bts_persentase_penambahan': f'{round(persentasiKenaikanBTS, 2)}%' if penugasanBTSSekarang != penugasanBTSSebelumnya else '',

                    'media_url': settings.URL_MEDIA,
                    'survej_url': settings.URL_LOGIN,

                    'message_bottom': 'silahkan login akun SMASLAB untuk melihat detil laporan melalui aplikasi mobile ataupun website '+settings.URL_LOGIN

                }

                executive = [x.email for x in UserInfo.objects.filter(
                    role=ObjectId('5f27f7a6ea250a01d2b99d7a'))]

                html_content = htmly.render(d)
                sender = settings.EMAIL_ADMIN

                msg = EmailMultiAlternatives(
                    subject, text_content, sender, executive)
                msg.attach_alternative(html_content, "text/html")
                response = msg.send()
                # print('asdad', response)
            except:
                pass

        except Exception as e:
            print(e)
        # return HttpResponse(f' & ')
        return Response.ok(
            values=[],
            message='Berhasil'
        )
    # return HttpResponse(f'asd')
    return Response.badRequest(
        values=[],
        message='Gagal'
    )


def haversine(lon1, lat1, lon2, lat2):
    """
    Calculate the great circle distance between two points
    on the earth (specified in decimal degrees)
    """
    # convert decimal degrees to radians
    lon1, lat1, lon2, lat2 = map(radians, [lon1, lat1, lon2, lat2])

    # haversine formula
    dlon = lon2 - lon1
    dlat = lat2 - lat1
    a = sin(dlat/2)**2 + cos(lat1) * cos(lat2) * sin(dlon/2)**2
    c = 2 * asin(sqrt(a))
    r = 6371  # Radius of earth in kilometers. Use 3956 for miles
    return c * r


def uploadsite(request):
    if request.method == 'POST':
        lokasi_gagal = ''

        file = request.FILES['doc']
        if not file:
            return Response.badRequest(message='Doc tidak boleh kosong')
        fs = FileSystemStorage(
            location=f'{settings.MEDIA_ROOT}/site/rfi/',
            base_url=f'{settings.MEDIA_URL}/site/rfi/'
        )
        body_data = request.POST.dict()

        judul = body_data.get('judul')
        tanggal_mulai_undangan = datetime.strptime(
            body_data.get('tanggal_mulai_undangan'), '%Y-%m-%d 00:00:00')
        tanggal_selesai_undangan = datetime.strptime(
            body_data.get('tanggal_selesai_undangan'), '%Y-%m-%d 23:59:59')
        if tanggal_selesai_undangan < tanggal_mulai_undangan:
            return Response.badRequest(
                values='null',
                message='tanggal_selesai_undangan harus lebih besar dari tanggal_mulai_undangan'
            )

        tanggal_mulai_kerja = datetime.strptime(
            body_data.get('tanggal_mulai_kerja'), '%Y-%m-%d 00:00:00')
        if tanggal_mulai_kerja < tanggal_selesai_undangan:
            return Response.badRequest(
                values='null',
                message='tanggal_mulai_kerja harus lebih besar dari tanggal_selesai_undangan'
            )
        tanggal_selesai_kerja = datetime.strptime(
            body_data.get('tanggal_selesai_kerja'), '%Y-%m-%d 23:59:59')
        if tanggal_selesai_kerja < tanggal_mulai_kerja:
            return Response.badRequest(
                values='null',
                message='tanggal_selesai_kerja harus lebih besar dari tanggal_mulai_kerja'
            )
        no_doc_permohonan_rfi_ = body_data.get('no_doc_permohonan_rfi')
        type = body_data.get('type')
        creator = body_data.get('creator')
        penyedia_undang = body_data.get('penyedia_undang')

        status_ = {'status': 'Dibuka', 'tanggal_pembuatan': datetime.utcnow(
        ) + timedelta(hours=7)}

        # try:
        #    # data_nomor_batch = batch.objects.latest('nomor')
        #    data_nomor_batch = batch.objects.order_by('-nomor').first()
        #    nomor_batch = int(data_nomor_batch.nomor) + 1
        #    nomor_batch = str(nomor_batch).zfill(5)
        # except:
        # except Exception as e:
        #    print(e)
        #    nomor_batch = '1'.zfill(5)

        vendor_list = penyedia_undang.split(",")
        data_batch = batch(
            judul=judul,
            type=type,
            sites=[],
            creator=creator,
            no_doc_permohonan_rfi=no_doc_permohonan_rfi_,
            tanggal_mulai_undangan=tanggal_mulai_undangan,
            tanggal_selesai_undangan=tanggal_selesai_undangan,
            tanggal_mulai_kerja=tanggal_mulai_kerja,
            tanggal_selesai_kerja=tanggal_selesai_kerja,
            penyedia_undang=vendor_list,
            # created_at = datetime.utcnow() + timedelta(hours=7),
            # updated_at = datetime.utcnow() + timedelta(hours=7)
        )
        data_batch.status.append(status_)
        data_batch.save()

        filename = fs.save(file.name, file)
        file_path = fs.url(filename)
        doc = doc_permohonan_rfi(
            name=file.name,
            path=file_path
        )
        doc.save()

        data_batch.doc_permohonan_rfi = ObjectId(doc.id)
        data_batch.save()

        req_fields = ['latitude', 'longitude', 'kecamatan']
        data_site_lok = site.objects.all().only(*req_fields)
        radius = 1.00  # in kilometer

        new_data_site_lok = []
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["LIST SITE"]

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        id_gagal = []
        for row in worksheet.iter_rows():
            try:
                lanjut = True
                if str(row[1].value) == 'None':
                    break
                if str(row[0].value) == 'NO':
                    continue
                data_provinsi = provinsi.objects.filter(
                    name=str(row[2].value).upper()).first()
                if data_provinsi is None:
                    # data_provinsi = provinsi(
                    #    name=str(row[2].value).upper()
                    # )
                    # data_provinsi.save()
                    json_dict = {}
                    json_dict["no"] = str(row[0].value).strip()
                    json_dict["provinsi"] = str(row[2].value).strip()
                    id_gagal.append(json_dict)
                    continue

                if str(row[3].value)[0:3].upper() == 'KAB':
                    kabupaten_ = kabupaten.objects.filter(
                        name=str(row[3].value).upper()).first()
                    if kabupaten_ is None:
                        # kabupaten_ = kabupaten(
                        #    name=str(row[3].value).upper(),
                        #    provinsi=ObjectId(data_provinsi.id)
                        # )
                        # kabupaten_.save()
                        json_dict = {}
                        json_dict["no"] = str(row[0].value).strip()
                        json_dict["provinsi_id"] = data_provinsi.id
                        json_dict["provinsi"] = data_provinsi.name
                        json_dict["kabupaten"] = str(row[3].value).strip()
                        id_gagal.append(json_dict)
                        continue
                    kab_kot_name = kabupaten_.name
                else:
                    kota_ = kota.objects.filter(
                        name=str(row[3].value).upper()).first()
                    if kota_ is None:
                        # kota_ = kota(
                        #    name=str(row[3].value).upper(),
                        #    provinsi=ObjectId(data_provinsi.id)
                        # )
                        # kota_.save()
                        json_dict = {}
                        json_dict["no"] = str(row[0].value).strip()
                        json_dict["provinsi_id"] = data_provinsi.id
                        json_dict["provinsi"] = data_provinsi.name
                        json_dict["kota"] = str(row[3].value).strip()
                        id_gagal.append(json_dict)
                        continue
                    kab_kot_name = kota_.name
                data_kecamatan = kecamatan.objects.filter(
                    name=str(row[4].value).upper()).first()
                if data_kecamatan is None:
                    # try:
                    #    data_kecamatan = kecamatan(
                    #        name=str(row[4].value).upper(),
                    #        kabupaten=ObjectId(kabupaten_.id)
                    #    )
                    #    data_kecamatan.save()
                    # except:
                    #    data_kecamatan = kecamatan(
                    #        name=str(row[4].value).upper(),
                    #        kota=ObjectId(kota_.id)
                    #    )
                    #    data_kecamatan.save()
                    if str(row[3].value)[0:3].upper() == 'KAB':
                        json_dict = {}
                        json_dict["no"] = str(row[0].value).strip()
                        json_dict["provinsi"] = data_provinsi.name
                        json_dict["kab_kota_id"] = kabupaten_.id
                        json_dict["kab_kota"] = kabupaten_.name
                        json_dict["kecamatan"] = str(row[4].value).strip()
                        id_gagal.append(json_dict)
                        continue
                    else:
                        json_dict = {}
                        json_dict["no"] = str(row[0].value).strip()
                        json_dict["provinsi"] = data_provinsi.name
                        json_dict["kab_kota_id"] = kota_.id
                        json_dict["kab_kota"] = kota_.name
                        json_dict["kecamatan"] = str(row[4].value).strip()
                        id_gagal.append(json_dict)
                        continue

                data_desa = desa.objects.filter(
                    name=str(row[5].value).upper()).first()
                if data_desa is None:
                    # data_desa = desa(
                    #    name=str(row[5].value).upper(),
                    #    kecamatan=ObjectId(data_kecamatan.id)
                    # )
                    # data_desa.save()
                    json_dict = {}
                    json_dict["no"] = str(row[0].value).strip()
                    json_dict["provinsi"] = data_provinsi.name
                    json_dict["kab_kota"] = kab_kot_name
                    json_dict["kecamatan_id"] = data_kecamatan.id
                    json_dict["kecamatan"] = data_kecamatan.name
                    json_dict["desa"] = str(row[5].value).strip()
                    id_gagal.append(json_dict)
                    continue
                if str(row[1].value).upper() == 'PERMOHONAN AKSES INTERNET':
                    jns = 'AI'
                else:
                    jns = 'AI'

                for dat in data_site_lok:
                    if dat.kecamatan.id != data_kecamatan.id:
                        continue

                    a = haversine(float(dat.longitude), float(dat.latitude), float(
                        str(row[7].value)), float(str(row[6].value)))
                    # print(a)
                    if a <= radius:
                        lanjut = False
                        break
                for dt in new_data_site_lok:
                    a = haversine(float(dt['longitude']), float(dt['latitude']), float(
                        str(row[7].value)), float(str(row[6].value)))
                    # print(a)
                    if a <= radius:
                        lanjut = False
                        break
                if lanjut:
                    # try:
                    #    data_nomor_site = site.objects.order_by('-unik_id').first()
                    #    nomor_site = data_nomor_site.unik_id + 1
                    #    # nomor_site = str(nomor_site).zfill(5)
                    # except Exception as e:
                    #    print(e)
                    #    # nomor_site = '1'.zfill(5)
                    #    nomor_site = 1

                    rekomentek = getRecommendTechnologi(
                        str(row[7].value), str(row[6].value))
                    data_site = site(
                        unik_id=str(row[10].value),
                        latitude=str(row[6].value),
                        longitude=str(row[7].value),
                        longlat=[float(str(row[7].value)),
                                 float(str(row[6].value))],
                        rekomendasi_teknologi=rekomentek,
                        nama=str(row[8].value),
                        desa_kelurahan=ObjectId(data_desa.id),
                        kecamatan=ObjectId(data_kecamatan.id),
                        provinsi=ObjectId(data_provinsi.id),
                        kode_pos=str(row[9].value),
                    )
                    if str(row[3].value)[0:3].upper() == 'KAB':
                        data_site.kabupaten = kabupaten_.id
                    else:
                        data_site.kota = kota_.id
                    data_site.save()

                    data_site_matchmaking = site_matchmaking(
                        siteid=data_site.id,
                        batchid=ObjectId(data_batch.id)
                    )
                    data_site_matchmaking.save()

                    data_site.site_matchmaking.append(data_site_matchmaking.id)
                    data_site.save()

                    data_batch.sites.append(ObjectId(data_site_matchmaking.id))
                    data_batch.save()

                    longlat_ = {'longitude': str(
                        row[7].value), 'latitude': str(row[6].value)}
                    new_data_site_lok.append(longlat_)
                else:
                    lokasi_gagal += '{' + \
                        str(row[6].value)+', '+str(row[7].value)+'}, '
            except:
                # lokasi_gagal += '{' + \
                #    str(row[6].value)+', '+str(row[7].value)+'}, '
                json_dict = {}
                json_dict["no"] = str(row[0].value).strip()
                json_dict["latitude"] = str(row[6].value)
                json_dict["longitude"] = str(row[7].value)
                id_gagal.append(json_dict)
                continue
        # return Response.ok(
        #    values=[],
        #    message=lokasi_gagal
        # )
        return Response.ok(
            values=json.loads(json.dumps(id_gagal, default=str)),
            message="OK"
        )


def addbatch(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    if request.method == "POST":  # Add
        try:
            file = request.FILES['doc']
            if not file:
                return Response.badRequest(message='Doc tidak boleh kosong')
            fs = FileSystemStorage(
                location=f'{settings.MEDIA_ROOT}/site/rfi/',
                base_url=f'{settings.MEDIA_URL}/site/rfi/'
            )

            body_data = request.POST.dict()

            judul = body_data.get('judul')
            tanggal_mulai_undangan = body_data.get('tanggal_mulai_undangan')
            tanggal_selesai_undangan = body_data.get(
                'tanggal_selesai_undangan')
            if tanggal_selesai_undangan < tanggal_mulai_undangan:
                return Response.badRequest(
                    values='null',
                    message='tanggal_selesai_undangan harus lebih besar dari tanggal_mulai_undangan'
                )

            tanggal_mulai_kerja = body_data.get('tanggal_mulai_kerja')
            if tanggal_mulai_kerja < tanggal_selesai_undangan:
                return Response.badRequest(
                    values='null',
                    message='tanggal_mulai_kerja harus lebih besar dari tanggal_selesai_undangan'
                )
            tanggal_selesai_kerja = body_data.get('tanggal_selesai_kerja')
            if tanggal_selesai_kerja < tanggal_mulai_kerja:
                return Response.badRequest(
                    values='null',
                    message='tanggal_selesai_kerja harus lebih besar dari tanggal_mulai_kerja'
                )
            no_doc_permohonan_rfi_ = body_data.get('no_doc_permohonan_rfi')
            type = body_data.get('type')
            creator = body_data.get('creator')
            penyedia_undang = body_data.get('penyedia_undang')
            price = body_data.get('price', 0)
            tech_type = body_data.get('tech_type', '')
            buying_type = body_data.get('buying_type', 0)

            status_ = {'status': 'Dibuka', 'tanggal_pembuatan': datetime.utcnow(
            ) + timedelta(hours=7)}

            # try:
            #    # data_nomor_batch = batch.objects.latest('nomor')
            #    data_nomor_batch = batch.objects.order_by('-nomor').first()
            #    nomor_batch = int(data_nomor_batch.nomor) + 1
            #    nomor_batch = str(nomor_batch).zfill(5)
            # except:
            # except Exception as e:
            #    print(e)
            #    nomor_batch = '1'.zfill(5)

            vendor_list = penyedia_undang.split(",")
            data_batch = batch(
                judul=judul,
                type=type,
                sites=[],
                creator=creator,
                no_doc_permohonan_rfi=no_doc_permohonan_rfi_,
                tanggal_mulai_undangan=tanggal_mulai_undangan,
                tanggal_selesai_undangan=tanggal_selesai_undangan,
                tanggal_mulai_kerja=tanggal_mulai_kerja,
                tanggal_selesai_kerja=tanggal_selesai_kerja,
                penyedia_undang=vendor_list,
                buying_type=buying_type,
                price=price
                # created_at = datetime.utcnow() + timedelta(hours=7),
                # updated_at = datetime.utcnow() + timedelta(hours=7)
            )
            if len(tech_type) > 0:
                data_batch.tech_type = tech_type
            data_batch.status.append(status_)
            data_batch.save()

            filename = fs.save(file.name, file)
            file_path = fs.url(filename)
            doc = doc_permohonan_rfi(
                name=file.name,
                path=file_path
                # create_date=datetime.utcnow() + timedelta(hours=7),
                # update_date=datetime.utcnow() + timedelta(hours=7)
            )
            doc.save()

            data_batch.doc_permohonan_rfi = ObjectId(doc.id)
            data_batch.save()

            # for vn in vendor_list:
            #    try:
            #        comp = vendor.objects.get(id=ObjectId(vn))
            #    except vendor.DoesNotExist:
            #        return Response.ok(
            #            values=[],
            #            message='Penyedia tidak ada'
            #        )
            #    data_batch_vendor = batch_vendor(
            #        vendor = comp.id,
            #        batch_id = data_batch.id,
            #        created_at = datetime.utcnow() + timedelta(hours=7),
            #        updated_at = datetime.utcnow() + timedelta(hours=7)
            #    )
            #    data_batch_vendor.status.append(status_)
            #    data_batch_vendor.save()
            result = batch.objects.get(id=ObjectId(data_batch.id)).serialize()
            # serializer = BatchSerializer(data_batch)
            # result = serializer.data

            return Response.ok(
                values=result,
                message='Berhasil'
            )
        except Exception as e:
            return Response.badRequest(message=str(e))

    else:
        return Response.badRequest(message='Hanya POST')


def addsitebyid(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    if request.method == "POST":  # Add
        try:
            body_data = json.loads(request.body)

            batch_id = body_data.get('batch')
            nama = body_data.get('nama')
            provinsi = body_data.get('provinsi')
            kab_kota = body_data.get('kab_kota')
            kecamatan = body_data.get('kecamatan')
            desa = body_data.get('desa')
            longitude = body_data.get('longitude')
            latitude = body_data.get('latitude')
            kode_pos = body_data.get('kode_pos')
            nomor_site = body_data.get('unik_id')

            req_fields = ['latitude', 'longitude', 'kecamatan']
            data_site_lok = site.objects.all().only(*req_fields)
            radius = 1.00  # in kilometer

            for dat in data_site_lok:
                if dat.kecamatan.id != ObjectId(kecamatan):
                    continue
                a = haversine(float(dat.longitude), float(
                    dat.latitude), float(longitude), float(latitude))
                # print(a)
                if a <= radius:
                    # return Response.ok(
                    #    values=[],
                    #    message='Data sudah ada'
                    # )
                    return Response().base(
                        success=False,
                        message='Data sudah ada',
                        status=409
                    )

            try:
                data_kabupaten = kabupaten.objects.get(id=kab_kota)
                data_kab_kota = 'kab'
            except kabupaten.DoesNotExist:
                data_kabupaten = kota.objects.get(id=kab_kota)
                data_kab_kota = 'kota'

            # try:
            #    data_nomor_site = site.objects.order_by('-unik_id').first()
            #    nomor_site = data_nomor_site.unik_id + 1
            #    # nomor_site = str(nomor_site).zfill(5)
            # except Exception as e:
            #    print(e)
            #    # nomor_site = '1'.zfill(5)
            #    nomor_site = 1

            rekomentek = getRecommendTechnologi(longitude, latitude)
            data_site = site(
                unik_id=nomor_site,
                latitude=latitude,
                longitude=longitude,
                longlat=[float(longitude), float(latitude)],
                rekomendasi_teknologi=rekomentek,
                nama=nama,
                desa_kelurahan=ObjectId(desa),
                kecamatan=ObjectId(kecamatan),
                provinsi=ObjectId(provinsi),
                kode_pos=kode_pos,
                # created_at = datetime.utcnow() + timedelta(hours=7),
                # updated_at = datetime.utcnow() + timedelta(hours=7)
            )
            if data_kab_kota == 'kab':
                data_site.kabupaten = data_kabupaten.id
            else:
                data_site.kota = data_kabupaten.id
            data_site.save()

            data_site_matchmaking = site_matchmaking(
                siteid=data_site.id,
                batchid=ObjectId(batch_id)
            )
            data_site_matchmaking.save()

            data_site.site_matchmaking.append(data_site_matchmaking.id)
            data_site.save()
            try:
                data_batch = batch.objects.get(id=ObjectId(batch_id))
                data_batch.sites.append(ObjectId(data_site_matchmaking.id))
                data_batch.save()
            except batch.DoesNotExist:
                # return Response.badRequest(message='Batch tidak ada')
                return Response().base(
                    success=False,
                    message='Batch tidak ada',
                    status=404
                )

            results = site.objects.get(id=ObjectId(data_site.id))

            result = results.serialize()

            return Response.ok(
                values=result,
                message='Berhasil'
            )
        except Exception as e:
            return Response.badRequest(message=str(e))

    else:
        return Response.badRequest(message='Hanya POST')


def addsite(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    if request.method == "POST":  # Add
        try:
            body_data = json.loads(request.body)

            batch_id = body_data.get('batch')
            nama = body_data.get('nama')

            provinsiName = body_data.get('provinsi')

            # try:
            province = provinsi.objects.filter(name=provinsiName).first()

            kab_kotaName = body_data.get('kab_kota')

            kab_kota = kabupaten.objects.filter(
                name=kab_kotaName, provinsi=province.id).first()
            if not kab_kota:
                kab_kota = kota.objects.filter(
                    name=kab_kotaName, provinsi=province.id).first()
                data_kab_kota = 'kota'
            else:
                data_kab_kota = 'kab'
            kecamatanName = body_data.get('kecamatan')

            kecamatans = kecamatan.objects.filter(
                name=kecamatanName, kabupaten=kab_kota.id).first()
            if not kecamatans:
                kecamatans = kecamatan.objects.filter(
                    name=kecamatanName, kota=kab_kota.id).first()

            desaName = body_data.get('desa')

            print(desaName)
            print(kecamatans.id)
            desas = desa.objects.filter(
                name=desaName, kecamatan=kecamatans.id).first()
            print(desas)
            if not desas:
                return Response().base(
                    success=False,
                    message='Desa tidak ada',
                    status=404
                )

            longitude = body_data.get('longitude')
            latitude = body_data.get('latitude')
            kode_pos = body_data.get('kode_pos')
            nomor_site = body_data.get('unik_id')

            req_fields = ['latitude', 'longitude',
                          'kecamatan', 'site_matchmaking']
            data_site_lok = site.objects.all().only(*req_fields)
            radius = 1.00  # in kilometer

            for dat in data_site_lok:
                curSmm = site_matchmaking.objects.filter(
                    id=dat.site_matchmaking[0]).first()
                if dat.kecamatan.id != kecamatans.id:
                    continue
                if str(batch_id) == str(curSmm.batchid.id):
                    continue
                a = haversine(float(dat.longitude), float(
                    dat.latitude), float(longitude), float(latitude))
                if a <= radius:
                    return Response().base(
                        success=True,
                        message='Data sudah ada',
                        status=200
                    )

            # try:
            #    data_kabupaten = kabupaten.objects.get(id=kab_kota.id)
            #    data_kab_kota = 'kab'
            # except kabupaten.DoesNotExist:
            #    data_kabupaten = kota.objects.get(id=kab_kota.id)
            #    data_kab_kota = 'kota'

            rekomentek = getRecommendTechnologi(longitude, latitude)
            data_site = site(
                unik_id=nomor_site,
                latitude=latitude,
                longitude=longitude,
                longlat=[float(longitude), float(latitude)],
                rekomendasi_teknologi=rekomentek,
                nama=nama,
                desa_kelurahan=desas['id'],
                kecamatan=kecamatans['id'],
                provinsi=province['id'],
                kode_pos=kode_pos,
                # created_at = datetime.utcnow() + timedelta(hours=7),
                # updated_at = datetime.utcnow() + timedelta(hours=7)
            )
            if data_kab_kota == 'kab':
                data_site.kabupaten = kab_kota.id
            else:
                data_site.kota = kab_kota.id
            data_site.save()

            data_site_matchmaking = site_matchmaking(
                siteid=data_site.id,
                batchid=ObjectId(batch_id)
            )
            data_site_matchmaking.save()

            data_site.site_matchmaking.append(data_site_matchmaking.id)
            data_site.save()
            try:
                data_batch = batch.objects.get(id=ObjectId(batch_id))
                data_batch.sites.append(ObjectId(data_site_matchmaking.id))
                data_batch.save()
            except batch.DoesNotExist:
                # return Response.badRequest(message='Batch tidak ada')
                return Response().base(
                    success=False,
                    message='Batch tidak ada',
                    status=404
                )

            results = site.objects.get(id=ObjectId(data_site.id))

            result = results.serialize()

            return Response.ok(
                values=result,
                message='Berhasil'
            )
        except provinsi.DoesNotExist:
            return Response.badRequest(message='Province Not Found')
        except kota.DoesNotExist:
            return Response.badRequest(message='Kab_kota Not Found')
        except kecamatan.DoesNotExist:
            return Response.badRequest(message='kecamatan Not Found')
        except desa.DoesNotExist:
            return Response.badRequest(message='desa Not Found')
        except Exception as e:
            print(e)
            return Response.badRequest(message=str(e))
    else:
        return Response.badRequest(message='Hanya POST')


"""
def addsite(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    if request.method == "POST":  # Add
        try:
            body_data = json.loads(request.body)

            batch_id = body_data.get('batch')
            nama = body_data.get('nama')

            provinsiName = body_data.get('provinsi')

            # try:
            province = provinsi.objects.get(name=provinsiName)

            kab_kotaName = body_data.get('kab_kota')

            try:
                kab_kota = kabupaten.objects.get(
                    name=kab_kotaName, provinsi=province['id'])
            except kabupaten.DoesNotExist:
                kab_kota = kota.objects.get(
                    name=kab_kotaName, provinsi=province['id'])

            kecamatanName = body_data.get('kecamatan')

            try:
                kecamatans = kecamatan.objects.get(
                    name=kecamatanName, kabupaten=kab_kota['id'])
            except kecamatan.DoesNotExist:
                kecamatans = kecamatan.objects.get(
                    name=kecamatanName, kota=kab_kota['id'])

            desaName = body_data.get('desa')

            desas = desa.objects.get(
                name=desaName, kecamatan=kecamatans['id'])

            # except province.DoesNotExist:
            #     return Response.badRequest(message='Province Not Found')
            # except kota.DoesNotExist:
            #     return Response.badRequest(message='Kab_kota Not Found')
            # except kecamatan.DoesNotExist:
            #     return Response.badRequest(message='kecamatan Not Found')
            # except desa.DoesNotExist:
            #     return Response.badRequest(message='desa Not Found')

            longitude = body_data.get('longitude')
            latitude = body_data.get('latitude')
            kode_pos = body_data.get('kode_pos')
            nomor_site = body_data.get('unik_id')

            req_fields = ['latitude', 'longitude', 'kecamatan']
            data_site_lok = site.objects.all().only(*req_fields)
            radius = 1.00  # in kilometer

            for dat in data_site_lok:
                if dat.kecamatan.id != kecamatans['id']:
                    continue
                a = haversine(float(dat.longitude), float(
                    dat.latitude), float(longitude), float(latitude))
                # print(a)
                if a <= radius:
                    # return Response.ok(
                    #    values=[],
                    #    message='Data sudah ada'
                    # )
                    return Response().base(
                        success=False,
                        message='Data sudah ada',
                        status=409
                    )

            try:
                data_kabupaten = kabupaten.objects.get(id=kab_kota['id'])
                data_kab_kota = 'kab'
            except kabupaten.DoesNotExist:
                data_kabupaten = kota.objects.get(id=kab_kota['id'])
                data_kab_kota = 'kota'

            # try:
            #    data_nomor_site = site.objects.order_by('-unik_id').first()
            #    nomor_site = data_nomor_site.unik_id + 1
            #    # nomor_site = str(nomor_site).zfill(5)
            # except Exception as e:
            #    print(e)
            #    # nomor_site = '1'.zfill(5)
            #    nomor_site = 1

            rekomentek = getRecommendTechnologi(longitude, latitude)
            data_site = site(
                unik_id=nomor_site,
                latitude=latitude,
                longitude=longitude,
                longlat=[float(longitude), float(latitude)],
                rekomendasi_teknologi=rekomentek,
                nama=nama,
                desa_kelurahan=desas['id'],
                kecamatan=kecamatans['id'],
                provinsi=province['id'],
                kode_pos=kode_pos,
                # created_at = datetime.utcnow() + timedelta(hours=7),
                # updated_at = datetime.utcnow() + timedelta(hours=7)
            )
            if data_kab_kota == 'kab':
                data_site.kabupaten = data_kabupaten.id
            else:
                data_site.kota = data_kabupaten.id
            data_site.save()

            data_site_matchmaking = site_matchmaking(
                siteid=data_site.id,
                batchid=ObjectId(batch_id)
            )
            data_site_matchmaking.save()

            data_site.site_matchmaking.append(data_site_matchmaking.id)
            data_site.save()
            try:
                data_batch = batch.objects.get(id=ObjectId(batch_id))
                data_batch.sites.append(ObjectId(data_site_matchmaking.id))
                data_batch.save()
            except batch.DoesNotExist:
                # return Response.badRequest(message='Batch tidak ada')
                return Response().base(
                    success=False,
                    message='Batch tidak ada',
                    status=404
                )

            results = site.objects.get(id=ObjectId(data_site.id))

            result = results.serialize()

            return Response.ok(
                values=result,
                message='Berhasil'
            )
        except provinsi.DoesNotExist:
            return Response.badRequest(message='Province Not Found')
        except kota.DoesNotExist:
            return Response.badRequest(message='Kab_kota Not Found')
        except kecamatan.DoesNotExist:
            return Response.badRequest(message='kecamatan Not Found')
        except desa.DoesNotExist:
            return Response.badRequest(message='desa Not Found')
        except Exception as e:
            print(e)
            return Response.badRequest(message=str(e))
    else:
        return Response.badRequest(message='Hanya POST')
"""

# def addsite(request):
#     # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
#     # ret, user = authenticate_credentials(token)
#     # if False == ret or None == user:
#     #    return JsonResponse({"state": "fail"})
#     if request.method == "POST":  # Add
#         try:
#             body_data = json.loads(request.body)

#             batch_id = body_data.get('batch')
#             nama = body_data.get('nama')

#             provinsi = body_data.get('provinsi')

#             kab_kota = body_data.get('kab_kota')
#             kecamatan = body_data.get('kecamatan')
#             desa = body_data.get('desa')

#             longitude = body_data.get('longitude')
#             latitude = body_data.get('latitude')
#             kode_pos = body_data.get('kode_pos')
#             nomor_site = body_data.get('unik_id')

#             req_fields = ['latitude', 'longitude', 'kecamatan']
#             data_site_lok = site.objects.all().only(*req_fields)
#             radius = 1.00  # in kilometer

#             for dat in data_site_lok:
#                 if dat.kecamatan.id != ObjectId(kecamatan):
#                     continue
#                 a = haversine(float(dat.longitude), float(
#                     dat.latitude), float(longitude), float(latitude))
#                 # print(a)
#                 if a <= radius:
#                     #return Response.ok(
#                     #    values=[],
#                     #    message='Data sudah ada'
#                     #)
#                     return Response().base(
#                         success=False,
#                         message='Data sudah ada',
#                         status=409
#                     )

#             try:
#                 data_kabupaten = kabupaten.objects.get(id=kab_kota)
#                 data_kab_kota = 'kab'
#             except kabupaten.DoesNotExist:
#                 data_kabupaten = kota.objects.get(id=kab_kota)
#                 data_kab_kota = 'kota'

#             #try:
#             #    data_nomor_site = site.objects.order_by('-unik_id').first()
#             #    nomor_site = data_nomor_site.unik_id + 1
#             #    # nomor_site = str(nomor_site).zfill(5)
#             #except Exception as e:
#             #    print(e)
#             #    # nomor_site = '1'.zfill(5)
#             #    nomor_site = 1

#             rekomentek = getRecommendTechnologi(longitude, latitude)
#             data_site = site(
#                 unik_id=nomor_site,
#                 latitude=latitude,
#                 longitude=longitude,
#                 longlat=[float(longitude), float(latitude)],
#                 rekomendasi_teknologi=rekomentek,
#                 nama=nama,
#                 desa_kelurahan=ObjectId(desa),
#                 kecamatan=ObjectId(kecamatan),
#                 provinsi=ObjectId(provinsi),
#                 kode_pos=kode_pos,
#                 # created_at = datetime.utcnow() + timedelta(hours=7),
#                 # updated_at = datetime.utcnow() + timedelta(hours=7)
#             )
#             if data_kab_kota == 'kab':
#                 data_site.kabupaten = data_kabupaten.id
#             else:
#                 data_site.kota = data_kabupaten.id
#             data_site.save()

#             data_site_matchmaking = site_matchmaking(
#                 siteid=data_site.id,
#                 batchid=ObjectId(batch_id)
#             )
#             data_site_matchmaking.save()

#             data_site.site_matchmaking.append(data_site_matchmaking.id)
#             data_site.save()
#             try:
#                 data_batch = batch.objects.get(id=ObjectId(batch_id))
#                 data_batch.sites.append(ObjectId(data_site_matchmaking.id))
#                 data_batch.save()
#             except batch.DoesNotExist:
#                 #return Response.badRequest(message='Batch tidak ada')
#                 return Response().base(
#                     success=False,
#                     message='Batch tidak ada',
#                     status=404
#                 )

#             results = site.objects.get(id=ObjectId(data_site.id))

#             result = results.serialize()

#             return Response.ok(
#                 values=result,
#                 message='Berhasil'
#             )
#         except Exception as e:
#             return Response.badRequest(message=str(e))

#     else:
#         return Response.badRequest(message='Hanya POST')


def editbatch(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    if request.method == "POST":  # Add

        # try:
        file = request.FILES['doc']
        if not file:
            return Response.badRequest(message='Doc tidak boleh kosong')
        fs = FileSystemStorage(
            location=f'{settings.MEDIA_ROOT}/site/rfi/',
            base_url=f'{settings.MEDIA_URL}/site/rfi/'
        )

        body_data = request.POST.dict()

        batch_id = body_data.get('batch')
        id = body_data.get('id')
        tanggal_mulai_undangan = body_data.get('tanggal_mulai_undangan')
        tanggal_selesai_undangan = body_data.get('tanggal_selesai_undangan')
        status_ = body_data.get('status')

        # if status_ == 'Selesai':
        #    return Response.ok(
        #        values=[],
        #        message='Status sudah selesai'
        #    )

        try:
            data_batch = batch.objects.get(id=ObjectId(batch_id))
            listVendor = list(
                map(itemgetter('id'), data_batch.penyedia_undang))
            listUser = UserInfo.objects(
                company__in=listVendor).only('id').only('email')
        except batch.DoesNotExist:
            # return Response.ok(message='Batch tidak ada')
            return Response().base(
                success=False,
                message='Batch tidak ada',
                status=404
            )

        users = list(map(itemgetter('id'), listUser))
        emails = list(map(itemgetter('email'), listUser))

        print(users, emails)

        data_batch.tanggal_mulai_undangan = tanggal_mulai_undangan
        data_batch.tanggal_selesai_undangan = tanggal_selesai_undangan
        type_ = ''
        message = ''
        title = ''

        if status_ == 'Dibuka':
            status__ = {'status': status_, 'tanggal_pembuatan': datetime.utcnow(
            ) + timedelta(hours=7)}
            data_batch.status.clear()
            data_batch.status.append(status__)
            type_ = 'edit batch opened'
            message = f'Undangan tawaran telah {status_.lower()}'
            title = f'Tawaran telah {status_.lower()}'
        else:
            status_buka = {'status': 'Dibuka', 'tanggal_pembuatan': datetime.utcnow(
            ) + timedelta(hours=7)}

            data_batch.status.clear()
            data_batch.status.append(status_buka)
            status_tunda = {'status': status_, 'tanggal_pembuatan': datetime.utcnow(
            ) + timedelta(hours=7)}
            data_batch.status.append(status_tunda)
            type_ = 'edit batch hold' if status_.lower() == 'ditunda' else 'edit batch closed'
            message = f'Undangan tawaran telah {status_.lower() if status_.lower() == "ditunda" else "ditutup"}'
            title = f'Tawaran {"sedang ditunda" if status_.lower() == "ditunda" else "telah ditutup"}'

        data_batch.save()

        filename = fs.save(file.name, file)
        file_path = fs.url(filename)
        doc = doc_permohonan_rfi(
            name=file.name,
            path=file_path,
            create_date=datetime.utcnow() + timedelta(hours=7),
            update_date=datetime.utcnow() + timedelta(hours=7)
        )
        doc.save()

        data_batch.rfi_doc = ObjectId(doc.id)
        data_batch.updated_at = datetime.utcnow() + timedelta(hours=7)
        data_batch.save()

        notif = CustomNotification()
        notif.create(ObjectId(id), users, type_,
                     title, message, 'Ada Pesan Baru')
        print(message)

        template = ''
        if status_.lower() == 'dibuka':
            template = 'email/webtampilandibuka.html'
        elif status_.lower() == 'ditunda':
            #template = 'email/webtampilantawaranproses.html'
            template = 'email/webtampilantawaranditunda.html'
        elif status_.lower() == 'selesai':
            template = 'email/webtampilantawaranselesai.html'

        # try:
        #     subject = f'Batch {data_batch.judul} {status_.lower()}'
        #     htmly = get_template(template)
        #     # html_content = htmly.render(d)
        #     msg = EmailMultiAlternatives(
        #         subject=subject, body='', from_email=settings.EMAIL_ADMIN, to=emails)

        #     msg.attach_alternative(htmly, "text/html")
        #     msg.content_subtype = 'html'
        #     msg.mixed_subtype = 'related'

        #     for x in [img_path_icon, img_path_logo]:
        #         with open(x, mode='rb') as f:
        #             name = Path(x).name
        #             image = MIMEImage(
        #                 f.read(), _subtype=name.split('.')[1])
        #             image.add_header('Content-ID', f"<{str(name)}>")
        #             msg.attach(image)
        #     print('aaaa')
        #     msg.send()
        # except Exception as e:
        #     print(e)

        d = {'media_url': settings.URL_MEDIA,
             'url_login': settings.URL_LOGIN}

        send_mail(f'Batch {data_batch.judul} {status_.lower()}', '', template,
                  d, settings.EMAIL_ADMIN, emails)

        result = batch.objects.get(id=ObjectId(data_batch.id)).serialize()
        # result = data_batch.serialize()

        # serializer = BatchSerializer(result,many=True)
        # result = serializer.data

        return Response.ok(
            values=result,
            message='Berhasil'
        )
        # except Exception as e:
        #    return Response.badRequest(message=str(e))

    else:
        return Response.badRequest(message='Hanya POST')


def getbatch(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    body_data = json.loads(request.body)
    batch_id = body_data.get('batch')

    # page = int(body_data.get('page', 0)) - 1
    # skip = []
    # if page >= 0:
    #     skip = [{'$skip': 20 * page},
    #             {'$limit': 20}]

    # pipeline = [
    #     {
    #         '$match': {
    #             '_id': ObjectId(batch_id)
    #         }
    #     },
    #     {
    #         '$lookup': {
    #             'from': 'site_location',
    #             'localField': 'sites',
    #             'foreignField': '_id',
    #             'as': 'sitelist'
    #         }
    #     }  # , {
    #     #    '$unwind': {
    #     #        'path': '$sitelist',
    #     #        'preserveNullAndEmptyArrays': True
    #     #    }
    #     # }
    # ]

    # pipe = pipeline + skip
    # agg_cursor = batch.objects.aggregate(*pipe)

    # batch_list = list(agg_cursor)

    try:
        data = batch.objects.get(id=batch_id)
        serializer = BatchSerializer(data)
        return Response.ok(
            values=[json.loads(json.dumps(serializer.data, default=str))],
            message=f'{len(serializer.data)} Data'
        )
    except batch.DoesNotExist:
        # return Response.ok(
        #    values=[],
        #    message='Data tidak ada'
        # )
        return Response().base(
            success=False,
            message='Data tidak ada',
            status=404
        )


def getallbatch(request):
    try:
        data = batch.objects.all()
        serializer = BatchSerializer(data, many=True)
        if len(serializer.data) > 0:
            return Response.ok(
                values=json.loads(json.dumps(serializer.data, default=str)),
                message=f'{len(serializer.data)} Data'
            )
        else:
            # return Response.ok(
            #    values=[],
            #    message='Data tidak ada'
            # )
            return Response().base(
                success=False,
                message='Data tidak ada',
                status=404
            )
    except Exception as e:
        print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def addodp(request):
    # token = request.META.get("HTTP_AUTHORIZATION").replace(" ", "")[6:]
    # ret, user = authenticate_credentials(token)
    # if False == ret or None == user:
    #    return JsonResponse({"state": "fail"})
    if request.method == "POST":  # Add
        try:
            body_data = json.loads(request.body)

            longitude = body_data.get('longitude')
            latitude = body_data.get('latitude')
            teknologi = body_data.get('teknologi')
            vendorid = body_data.get('vendorid')

            data_odp = Odp(
                latitude=latitude,
                longitude=longitude,
                longlat=[float(longitude), float(latitude)],
                teknologi=teknologi,
                vendor=vendorid
            )

            data_odp.save()
            result = []
            # results = site.objects.get(id=ObjectId(data_site.id))

            # result = results.serialize()

            return Response.ok(
                values=result,
                message='Berhasil'
            )
        except Exception as e:
            return Response.badRequest(message=str(e))

    else:
        return Response.badRequest(message='Hanya POST')


def getDashboard(request):
    try:
        vendorCount = vendor.objects.all().count()
        activeUserCount = UserInfo.objects(status='Aktif').count()
        requestedUserCount = UserInfo.objects(
            status='Belum Terverifikasi').count()
        batchCount = batch.objects.all().count()
        siteCount = site_matchmaking.objects(batchid__exists=True).count()
        rfiCount = vendor_application.objects.all().count()
        siteNonBatchCount = 0
        vendorListQuery = vendor.objects.all()
        vendorList = VendorScoreSerializer(vendorListQuery, many=True)

        #aiCount = Odp.objects.filter(teknologi__in=['VSAT','FIBER OPTIK','RADIO LINK']).count()
        aiCount = Odp.objects.filter(
            teknologi__in=['VSAT', 'FO', 'RL']).count()
        aiTech = Odp.objects.only('teknologi').distinct('teknologi')
        aiOperational = {
            "count": aiCount,
            "FO": 0,
            "VSAT": 0,
            "RL": 0
        }
        # aiOperational = {
        #    "count": aiCount,
        #    "FIBER OPTIK": 0,
        #    "VSAT": 0,
        #    "RADIO LINK": 0
        # }
        for x in list(aiTech):
            xx = x
            # if x=='FIBER OPTIK':
            #    xx='FO'
            #    aiOperational.pop('FIBER OPTIK', None)
            # if x=='RADIO LINK':
            #    xx='RL'
            #    aiOperational.pop('RADIO LINK', None)
            aiOperational.update({
                xx: Odp.objects(teknologi=x).count()
            })

        recommendTech = rekomendasi_teknologi.objects.only(
            'teknologi').distinct('teknologi')
        siteAICount = site.objects.all().count()
        aiNew = {
            "count": siteAICount,
            "FO": 0,
            "VSAT": 0,
            "RL": 0
        }
        for x in list(recommendTech):
            query = rekomendasi_teknologi.objects(teknologi=x).scalar('id')
            aiNew.update({
                x: site.objects(rekomendasi_teknologi__in=query).count()
            })

        date = datetime.now()
        listMonth = calendar.month_abbr[1:13]
        reportSite = {}
        reportRFi = {}
        for x in range(11, -1, -1):
            dateReport = date - relativedelta(months=x)
            year = dateReport.year
            month = dateReport.month
            lastDate = calendar.monthrange(year=year, month=month)[1]
            gte = datetime(year, month, 1, 00, 00, 00)
            lte = datetime(year, month, lastDate, 23, 59, 59)
            reportSite.update({
                f'{listMonth[month-1]} {year}': batch.objects(created_at__gte=gte, created_at__lte=lte).count()
            })
            reportRFi.update({
                f'{listMonth[month-1]} {year}': vendor_application.objects(created_at__gte=gte, created_at__lte=lte).count()
            })

        result = {
            "vendor": vendorCount,
            "active_user": activeUserCount,
            "requested_user": requestedUserCount,
            "batch": batchCount,
            "site": siteCount,
            "rfi": rfiCount,
            "site_not_batch": siteNonBatchCount,
            "vendor_list": json.loads(json.dumps(vendorList.data, default=str)),
            "running_ai": aiOperational,
            "new_ai": aiNew,
            "report": {
                "site": reportSite,
                "rfi": reportRFi
            }
        }

        return Response.ok(
            values=result
        )

    except Exception as e:
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def uploadsiteoffair(request):
    if request.method == 'POST':

        lokasi_gagal = ''

        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        worksheet = wb["Sheet1"]

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row

        for row in worksheet.iter_rows():
            lanjut = True
            if str(row[1].value) == 'None':
                break
            if str(row[0].value) == 'unik_id':
                continue
            data_site = site_offair.objects.filter(latitude=str(row[17].value).replace(',', '.'),
                                                   longitude=str(row[18].value).replace(',', '.'))
            if data_site:
                continue

            data_provinsi = provinsi.objects.filter(
                name__iexact=str(row[10].value).upper())
            if data_provinsi is None:
                lokasi_gagal += '{ ' + \
                    str(row[0].value)+' }, '
                continue
            prov_list = []
            for prov in data_provinsi:
                prov_list.append(prov.id)

            if str(row[11].value)[0:3].upper() == 'KAB':
                kabupaten_ = kabupaten.objects.filter(
                    name__iexact=str(row[11].value).upper(), provinsi__in=prov_list)
                if kabupaten_ is None:
                    lokasi_gagal += '{ ' + \
                        str(row[0].value)+' }, '
                    continue
                kab_list = []
                for kab in kabupaten_:
                    kab_list.append(kab.id)
            else:
                data_kota = kota.objects.filter(
                    name__iexact=str(row[11].value).upper(), provinsi__in=prov_list)
                if data_kota is None:
                    lokasi_gagal += '{ ' + \
                        str(row[0].value)+' }, '
                    continue
                kota_list = []
                for _kota in data_kota:
                    kota_list.append(_kota.id)
            if str(row[11].value)[0:3].upper() == 'KAB':
                data_kecamatan = kecamatan.objects.filter(
                    name__iexact=str(row[12].value).upper(), kabupaten__in=kab_list)
            else:
                data_kecamatan = kecamatan.objects.filter(
                    name__iexact=str(row[12].value).upper(), kota__in=kota_list)
            if data_kecamatan is None:
                lokasi_gagal += '{ ' + \
                    str(row[0].value)+' }, '
                continue
            kec_list = []
            for kec in data_kecamatan:
                kec_list.append(kec.id)

            data_desa = desa.objects.filter(
                name__iexact=str(row[13].value).upper(), kecamatan__in=kec_list)
            if data_desa is None:
                lokasi_gagal += '{ ' + \
                    str(row[0].value)+' }, '
                continue

            try:
                data_siteoffair = site_offair(
                    unik_id=str(row[0].value),
                    latitude=str(row[17].value).replace(',', '.'),
                    longitude=str(row[18].value).replace(',', '.'),
                    longlat=[float(str(row[18].value).replace(',', '.')), float(
                        str(row[17].value).replace(',', '.'))],
                    # teknologi=tekno,
                    nama=str(row[14].value),
                    desa_kelurahan=ObjectId(data_desa[0].id),
                    kecamatan=ObjectId(data_kecamatan[0].id),
                    provinsi=ObjectId(data_provinsi[0].id),
                    # status=data_vendor.id,
                )

                status_ = {'status': 'buka', 'tanggal_pembuatan': datetime.utcnow(
                ) + timedelta(hours=7)}
                data_siteoffair.status.append(status_)

                if str(row[11].value)[0:3].upper() == 'KAB':
                    data_siteoffair.kabupaten = kabupaten_[0].id
                else:
                    data_siteoffair.kota = data_kota[0].id

                data_siteoffair.save()
            except:
                lokasi_gagal += '{ ' + \
                    str(row[0].value)+' }, '
                continue

        return Response.ok(
            values=[],
            message=lokasi_gagal
        )


def getsiteoffair(request):
    try:
        #req_fields = ['latitude', 'longitude']
        try:
            start = int(request.GET.get('start')) - 1
            end = int(request.GET.get('end'))

            if start < 0:
                start = 0

            #data = site_offair.objects.all()[start:end]
            data = site_offair_norel.objects.all()[start:end]
        except:
            #data = site_offair.objects.all()
            data = site_offair_norel.objects.all()

        #serializer = siteoffairSerializer(data, many=True)
        serializer = siteoffairSerializer_norel(data, many=True)
        if len(serializer.data) > 0:
            return Response.ok(
                values=json.loads(json.dumps(serializer.data, default=str)),
                message=f'{len(serializer.data)} Data'
            )
        else:
            # return Response.ok(
            #    values=[],
            #    message='Data tidak ada'
            # )
            return Response().base(
                success=False,
                message='Data tidak ada',
                status=404
            )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def getoffairid(request):
    try:
        tech_type = str(request.GET.get('tech_type'))

        req_fields = ['unik_id', 'nama']

        if len(tech_type) > 0:
            data = site_offair.objects.get(tech_type=tech_type).only(*req_fields)
        else:
            data = site_offair.objects.all().only(*req_fields)

        result = []
        for dt in data:
            json_dict = {}
            json_dict["unik_id"] = dt.unik_id
            json_dict["nama"] = dt.nama
            result.append(json_dict)
        if len(result) > 0:
            return Response.ok(
                values=result,
                message=f'{len(result)} Data'
            )
        else:
            # return Response.ok(
            #    values=[],
            #    message='Data tidak ada'
            # )
            return Response().base(
                success=False,
                message='Data tidak ada',
                status=404
            )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def getoffairbyid(request):
    try:
        body_data = json.loads(request.body)
        unikid = body_data.get('unik_id')

        try:
            data = site_offair.objects.get(unik_id=unikid)
        except site_offair.DoesNotExist:
            # return Response.ok(
            #    values=[],
            #    message='data tidak ada'
            # )
            return Response().base(
                success=False,
                message='Data tidak ada',
                status=404
            )

        serializer = siteoffairSerializer(data)  # , many=True
        return Response.ok(
            values=json.loads(json.dumps(serializer.data, default=str)),
            message=f'{len(serializer.data)} Data'
        )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def getoffairprovinsi(request):
    try:
        provinsi = request.GET.get('provinsi')
        try:
            start = int(request.GET.get('start')) - 1
            end = int(request.GET.get('end'))

            if start < 0:
                start = 0

            data = site_offair.objects.filter(provinsi=provinsi)[start:end]
        except:
            data = site_offair.objects.filter(provinsi=provinsi)

        serializer = siteoffairSerializer(data, many=True)
        if len(serializer.data) > 0:
            return Response.ok(
                values=json.loads(json.dumps(serializer.data, default=str)),
                message=f'{len(serializer.data)} Data'
            )
        else:
            # return Response.ok(
            #    values=[],
            #    message='Data tidak ada'
            # )
            return Response().base(
                success=False,
                message='Data tidak ada',
                status=404
            )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def calculatevendorscore(request):
    # try:
    data_batch = batch.objects.filter(status__status__ne='Selesai',
                                      tanggal_selesai_undangan__gt=datetime.utcnow() + timedelta(hours=7))
    #data_batch = batch.objects.all()
    if len(data_batch) == 0:
        return Response.ok(
            values=[],
            message='Data tidak ada'
        )
    for dt_batch in data_batch:
        #data_smm = site_matchmaking.objects.all()
        data_smm = site_matchmaking.objects.filter(batchid=dt_batch.id)
        # for dt in data_smm:
        #     status = 'buka'
        #     for stt in dt.batchid.status:
        #         if 'Selesai' in stt['status']:
        #             status = 'selesai'
        #     if status=='buka':

        rw = 0
        for dt_smm in data_smm:
            # print(rw)
            smm_tek = dt_smm.siteid.rekomendasi_teknologi.teknologi
            list_days_work = []
            list_harga = []
            max_days_admin = 0
            min_days_admin = 0
            for dt_rfi in dt_smm.rfi_score:
                tgl_start = dt_rfi.vendor_app.tanggal_mulai_sla
                tgl_end = dt_rfi.integration + \
                    timedelta(dt_rfi.days_on_integration)
                days_work = int((tgl_end.date() - tgl_start.date()).days)
                list_days_work.append(days_work)
                list_harga.append(dt_rfi.biaya)

            if len(list_days_work) > 0:
                list_days_work = sorted(list_days_work)
                max_days_admin = list_days_work[-1]
                min_days_admin = list_days_work[0]
            if len(list_harga) > 0:
                list_harga = sorted(list_harga)
                max_harga = list_harga[-1]
                min_harga = list_harga[0]

            for dt_rfi in dt_smm.rfi_score:
                # if dt_rfi.total_calc:
                #    #print('continue')
                #    continue
                vendor_tek = dt_rfi.rekomendasi_teknologi
                tek_score = 0
                if smm_tek == "FO":
                    tek_score = 1
                elif smm_tek == vendor_tek:
                    tek_score = 1/2
                tgl_start = dt_rfi.vendor_app.tanggal_mulai_sla
                tgl_end = dt_rfi.integration + \
                    timedelta(dt_rfi.days_on_integration)
                days_work = int((tgl_end.date() - tgl_start.date()).days)
                if max_days_admin-min_days_admin == 0:
                    days_work = 1
                else:
                    days_work = 1-((days_work-min_days_admin) /
                                   (max_days_admin-min_days_admin))
                if max_harga-min_harga == 0:
                    nilai_harga = 1
                else:
                    nilai_harga = 1-((dt_rfi.biaya-min_harga) /
                                     (max_harga-min_harga))

                """
                vpscore_kecepatan = (
                    dt_rfi.vendor_app.vp_score_id.kecepatan-1)/(5-1)
                vpscore_ketepatan = (
                    dt_rfi.vendor_app.vp_score_id.ketepatan-1)/(5-1)
                vpscore_kualitas = (dt_rfi.vendor_app.vp_score_id.kualitas-1)/(5-1)
                """
                data_vpscore = VPScore.objects.filter(
                    vendor=dt_rfi.vendor_app.vendorid.id, kecepatan__gt=0)
                total_row = len(data_vpscore)
                vpscore_kecepatan = 0
                vpscore_ketepatan = 0
                vpscore_kualitas = 0
                for dt in data_vpscore:
                    vpscore_kecepatan += dt.kecepatan
                    vpscore_ketepatan += dt.ketepatan
                    vpscore_kualitas += dt.kualitas

                if total_row > 0:
                    vpscore_kecepatan = vpscore_kecepatan/total_row
                    vpscore_ketepatan = vpscore_ketepatan/total_row
                    vpscore_kualitas = vpscore_kualitas/total_row

                av_vp = (vpscore_kecepatan +
                         vpscore_ketepatan+vpscore_kualitas)/3

                if not dt_rfi.total_calc:
                    data_total_calc = total_calc(
                        rfi=days_work,
                        teknologi=tek_score,
                        vp=av_vp,
                        harga=nilai_harga
                    )
                    data_total_calc.save()

                    dt_rfi.total_calc = data_total_calc.id
                else:
                    data_total_calc = total_calc.objects.get(
                        id=dt_rfi.total_calc.id)

                    if not data_total_calc:
                        dt_rfi.total_calc = None
                    else:
                        data_total_calc.rfi = days_work
                        data_total_calc.teknologi = tek_score
                        data_total_calc.vp = av_vp
                        data_total_calc.harga = nilai_harga
                    data_total_calc.save()

                dt_rfi.save()
            rw += 1
    return Response.ok(
        values=[],
        message='Hitung total scoring berhasil'
    )
    """
    serializer = siteoffairSerializer(data, many=True)
    if len(serializer.data) > 0:
        return Response.ok(
            values=json.loads(json.dumps(serializer.data, default=str)),
            message=f'{len(serializer.data)} Data'
        )
    else:
        return Response.ok(
            values=[],
            message='Data tidak ada'
        )
    """
    # except Exception as e:
    #    #print(e)
    #    return Response.badRequest(
    #        values=[],
    #        message=str(e)
    #    )


def clonesiteoffair(request):
    try:
        #req_fields = ['latitude', 'longitude']
        data = site_offair.objects.all()

        for dt in data:
            data_prov = provinsi.objects.get(
                id=dt.provinsi.id)
            try:
                data_kab = kabupaten.objects.get(
                    id=dt.kabupaten.id, provinsi=data_prov.id)
            except kabupaten.DoesNotExist:
                data_kota = kota.objects.get(
                    id=dt.kota.id, provinsi=data_prov.id)
            if data_kab:
                data_kec = kecamatan.objects.get(
                    id=dt.kecamatan.id, kabupaten=data_kab.id)
            else:
                data_kec = kecamatan.objects.get(
                    id=dt.kecamatan.id, kota=data_kota.id)
            data_desa = desa.objects.get(
                id=dt.desa_kelurahan.id, kecamatan=data_kec.id)

            data_siteoffair = site_offair_norel(
                unik_id=dt.unik_id,
                latitude=dt.latitude,
                longitude=dt.longitude,
                longlat=[float(dt.longitude), float(dt.latitude)],
                nama=dt.nama,
                desa_kelurahan=data_desa.name,
                kecamatan=data_kec.name,
                provinsi=data_prov.name,
            )

            status_ = {'status': 'buka', 'tanggal_pembuatan': datetime.utcnow(
            ) + timedelta(hours=7)}
            data_siteoffair.status.append(status_)

            if data_kab:
                data_siteoffair.kabupaten = data_kab.name
            else:
                data_siteoffair.kota = data_kota.name

            data_siteoffair.save()

        return Response.ok(
            values=[],
            message='Berhasil'
        )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def sendinvitation(request):
    try:
        body_data = json.loads(request.body)

        batchid = body_data.get('batchid')
        vendors = []
        try:
            data_batch = batch.objects.get(id=ObjectId(batchid))
            for pu in data_batch.penyedia_undang:
                vendors.append(pu.id)
        except batch.DoesNotExist:
            return Response().base(
                success=False,
                message='Batch Does Not Exist',
                status=404
            )

        from_ = body_data.get('from')
        to_ = body_data.get('to').split(',')

        for vn in to_:
            if not ObjectId(vn) in vendors:
                data_batch.penyedia_undang.append(ObjectId(vn))
                data_batch.save()
            req_fields = ['id']
            vendor_users = UserInfo.objects.filter(
                company=ObjectId(vn), status='Aktif').only(*req_fields)
            if vendor_users:
                list_vendor_users = []
                for usr in vendor_users:
                    list_vendor_users.append(usr.id)
                notif = CustomNotification()
                title_ = 'Undangan batch berhasil diterima'
                if data_batch.type == 'VIP':
                    title_ = 'Invitation personal'

                notif.create(to=list_vendor_users, from_=ObjectId(from_), type='batch sent',
                             title=title_, message='batch '+data_batch.judul+' telah diterima', push_message='Ada pesan baru')

        return Response.ok(
            values=[],
            message='Berhasil'
        )
    except Exception as e:
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def checknearsiteoffair(request):
    try:
        if not request.body:
            return Response.badRequest(
                values=[],
                message="Need Json Body coordinates"
            )
        body = json.loads(request.body)
        coordinates = body.get('coordinates', None)
        retData = []

        for coord in coordinates:
            coordinate = [float(coord[0]), float(coord[1])]
            unik_id = coord[2]

            countEx = site.objects(
                longlat__geo_within_sphere=[coordinate, (1 / 6378.1)]).count()

            if countEx > 0:
                retData.append(unik_id)

        return Response.ok(
            values=json.loads(json.dumps(retData, default=str)),
            message=f'{len(retData)} Data'
        )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def getoffaircluster(request):
    try:
        reqKecamatan = request.GET.get('kecamatan')
        reqKabupaten = request.GET.get('kabupaten')
        reqKota = request.GET.get('kota')
        reqTech = request.GET.get('tech_type', '')

        data_kec = kecamatan.objects.get(
            id=reqKecamatan) if reqKecamatan is not None else None
        data_kota = kota.objects.get(
            id=reqKota) if reqKota is not None else None
        data_kab = kabupaten.objects.get(
            id=reqKabupaten) if reqKabupaten is not None else None
            
        if len(reqTech) > 0:
            data = site_offair_norel.objects.filter(tech_type=reqTech)
        else:
            data = site_offair_norel.objects.all()

        try:
            start = int(request.GET.get('start')) - 1
            end = int(request.GET.get('end'))

            if start < 0:
                start = 0

            if reqKecamatan and data_kec:
                data = data.filter(
                    kecamatan=data_kec.name)[start:end]
            elif reqKota and data_kota:
                data = data.filter(
                    kota=data_kota.name)[start:end]
            elif reqKabupaten and data_kab:
                data = data.filter(
                    kabupaten=data_kab.name)[start:end]
            else:
                return Response().base(
                    success=False,
                    message='Data tidak ada',
                    status=404
                )
        except:
            if reqKecamatan and data_kec:
                data = data.filter(
                    kecamatan=data_kec.name)
            elif reqKota and data_kota:
                data = data.filter(kota=data_kota.name)
            elif reqKabupaten and data_kab:
                data = data.filter(
                    kabupaten=data_kab.name)
            else:
                return Response().base(
                    success=False,
                    message='Data tidak ada',
                    status=404
                )

        serializer = siteoffairSerializer_norel(data, many=True)
        if len(serializer.data) > 0:
            return Response.ok(
                values=json.loads(json.dumps(serializer.data, default=str)),
                message=f'{len(serializer.data)} Data'
            )
        else:
            return Response().base(
                success=False,
                message='Data tidak ada',
                status=404
            )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def getvendorcluster(request):
    try:
        reqKecamatan = request.GET.get('kecamatan')
        reqKabupaten = request.GET.get('kabupaten')
        reqKota = request.GET.get('kota')

        match = {}
        minKec = 5
        minKab = 10
        if reqKecamatan:
            curKec = kecamatan.objects.filter(
                id=ObjectId(reqKecamatan)).first()
            if "kabupaten" in curKec:
                reqKabupaten = curKec["kabupaten"]["id"]
            else:
                reqKota = curKec["kota"]["id"]

            listOdpVenKec = list(Odp.objects.aggregate([
                {"$match": {"kecamatan": ObjectId(reqKecamatan)}},
                {"$group":
                    {
                        "_id": {
                            "vendor": "$vendor",
                            "teknologi": "$teknologi"
                        },
                        "count": {"$sum": 1}
                    }
                 },
                {"$sort":
                    {
                        "count": -1
                    }
                 }
            ]))
        else:
            listOdpVenKec = []

        if reqKota:
            match["kota"] = ObjectId(reqKota)
        elif reqKabupaten:
            match["kabupaten"] = ObjectId(reqKabupaten)
        else:
            return Response.badRequest(
                values=[],
                message="Need at least 1 location param"
            )

        listOdpVendor = list(Odp.objects.aggregate([
            {"$match": match},
            {"$group":
                {
                    "_id": {
                        "vendor": "$vendor",
                        "teknologi": "$teknologi"
                    },
                    "count": {"$sum": 1}
                }
             },
            {"$sort":
                {
                    "count": -1
                }
             }
        ]))

        respData = {
            "recommendations": [],
            "message": ""
        }

        if len(listOdpVendor) > 0 or len(listOdpVenKec) > 0:
            message = "Rekomendasi yang diberikan merupakan hasil analisa sistem."

            if len(listOdpVenKec) > 0:
                for curVendor in listOdpVenKec:
                    if curVendor["count"] >= minKec:
                        data = vendor.objects.get(
                            id=curVendor["_id"]["vendor"])
                        respData["recommendations"].append({
                            "vendor": data.name,
                            "teknologi": curVendor["_id"]["teknologi"],
                            "sla_daily": data.sla_avg if hasattr(data, "sla_avg") else 0,
                            "sla_monthly": data.sla_avg if hasattr(data, "sla_avg") else 0
                        })

                        message += " Penyedia " + data.name + \
                            " direkomendasikan dikarenakan mempunyai jumlah titik on air dengan total " + \
                            str(curVendor["count"]) + " titik di kecamatan ini."

            if len(respData["recommendations"]) == 0:
                for curVendor in listOdpVendor:
                    if curVendor["count"] >= minKab and curVendor["_id"]["vendor"] != "-":
                        data = vendor.objects.get(
                            id=curVendor["_id"]["vendor"])
                        respData["recommendations"].append({
                            "vendor": data.name,
                            "teknologi": curVendor["_id"]["teknologi"],
                            "sla_daily": data.sla_avg if hasattr(data, "sla_avg") else 0,
                            "sla_monthly": data.sla_avg if hasattr(data, "sla_avg") else 0
                        })

                        message += " Penyedia " + data.name + \
                            " direkomendasikan dikarenakan mempunyai jumlah titik on air dengan total " + \
                            str(curVendor["count"]) + " titik di kabupaten ini."

            if len(respData["recommendations"]) == 0:
                message = "Tidak ada rekomendasi vendor, anda bisa mengundang vendor siapa saja"
        else:
            message = "Tidak ada rekomendasi vendor, anda bisa mengundang vendor siapa saja"

        respData["message"] = message

        if len(respData) > 0:
            return Response.ok(
                values=json.loads(json.dumps(respData, default=str)),
                message=str(len(respData["recommendations"])) + ' Data'
            )
        else:
            return Response().base(
                success=False,
                message='Data tidak ada',
                status=404
            )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def syncsiteoffair(request):
    try:
        dateNow = datetime.now()

        # api-endpoint
        # listUrl = "http://localhost:3000/api/v1.0/request/offair-sis"
        listUrl = "https://pastiapi.datatsintesa.id/api/v1.0/request/offair-sis"

        # params here
        params = {
            "limit": 5000,
            "offset": 0,
            "date": dateNow - timedelta(days=1)
        }

        # sending get request and saving the response as response object
        r = requests.get(listUrl, params=params, headers={
                         "api-key": "S5QUo34beZvhEMifLU8D7GRRi9wSkrjp"})

        # extracting data in json format
        data = r.json()
        count = 1
        kecCount = 0
        nonCount = 0
        arr = []
        kecArr = []
        nonArr = []

        if data["data"]:
            for row in data["data"]:
                count += 1
                curReq = site_offair.objects.filter(
                    Q(unik_id=row["unik_id"]) |
                    Q(longitude=row["location"]["longitude"],
                      latitude=row["location"]["latitude"])
                ).first()

                if curReq:
                    arr.append(row["unik_id"])
                else:
                    if "kabupaten" in row and row["kabupaten"] is not None:
                        kabupatenCur = kabupaten.objects.filter(
                            id=ObjectId(name=row["kabupaten"]["name"])).first()
                        if "kecamatan" in row and row["kecamatan"] is not None:
                            kecamatanCur = kecamatan.objects.filter(
                                name=row["kecamatan"]["name"], kabupaten=ObjectId(kabupatenCur["id"])).first()
                    elif "kota" in row and row["kota"] is not None:
                        kotaCur = kota.objects.filter(
                            id=ObjectId(name=row["kota"]["name"])).first()
                        if "kecamatan" in row and row["kecamatan"] is not None:
                            kecamatanCur = kecamatan.objects.filter(
                                name=row["kecamatan"]["name"], kota=ObjectId(kotaCur["id"])).first()
                    else:
                        if "kecamatan" in row and row["kecamatan"] is not None:
                            kecamatanCur = kecamatan.objects.filter(
                                name=row["kecamatan"]["name"]).first()

                    if kecamatanCur:
                        curDesa = 0
                        curKota = 0
                        curKab = 0
                        curProv = 0
                        if "desa" in row and row["desa"] is not None:
                            desaCur = desa.objects.filter(
                                name=row["desa"]["name"]).first()
                            if desaCur:
                                curDesa = desaCur["id"]
                            else:
                                desaCur = desa.objects.filter(
                                    kecamatan=ObjectId(kecamatanCur["id"])).first()
                                if desaCur:
                                    curDesa = desaCur["id"]
                        else:
                            desaCur = desa.objects.filter(
                                kecamatan=ObjectId(kecamatanCur["id"])).first()
                            if desaCur:
                                curDesa = desaCur["id"]

                        if "kota" in kecamatanCur:
                            kotaCur = kota.objects.filter(
                                id=kecamatanCur["kota"]["id"]).first()
                            if kotaCur:
                                curKota = kotaCur["id"]
                                curProv = ObjectId(
                                    kotaCur["provinsi"]["id"])
                        elif "kabupaten" in kecamatanCur:
                            kabupatenCur = kabupaten.objects.filter(
                                id=ObjectId(kecamatanCur["kabupaten"]["id"])).first()
                            if kabupatenCur:
                                curKab = kabupatenCur["id"]
                                curProv = ObjectId(
                                    kabupatenCur["provinsi"]["id"])

                        data_site = site_offair(
                            unik_id=row["unik_id"],
                            latitude=row["location"]["latitude"],
                            longitude=row["location"]["longitude"],
                            nama=row["location"]["name"],
                            desa_kelurahan=curDesa,
                            kecamatan=kecamatanCur["id"],
                            provinsi=curProv,
                            kode_pos=row["location"]["kodepos"],
                            status=[
                                {
                                    "status": "buka",
                                    "tanggal_pembuatan": dateNow
                                }
                            ]
                        )

                        if curKab != 0:
                            data_site.kabupaten = curKab
                        elif curKota:
                            data_site.kota = curKota

                        data_site.save()
                    else:
                        kecCount += 1
                        kecArr.append(row["kecamatan"]["name"])

        return Response.ok(
            values=json.loads(json.dumps({
                "existing": arr,
                "kecProb": kecArr,
                "incomplete": nonArr
            }, default=str)),
            message='Berhasil'
        )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )


def validatebatchsites(request):
    try:
        if not request.body:
            return Response.badRequest(
                values=[],
                message="Need Json Body sites"
            )
        body = json.loads(request.body)
        sites = body.get('sites', None)
        retData = []

        for curSite in sites:
            curErr = []
            try:
                coordinate = [float(curSite["longitude"]),
                              float(curSite["latitude"])]
                countCoord = site.objects(
                    longlat__geo_within_sphere=[coordinate, (1 / 6378.1)]).count()

                if countCoord > 0:
                    curErr.append("Sudah terdapat titik lain di sekitar ini")
            except:
                curErr.append("Koordinat tidak valid")

            unik_id = curSite["unik_id"]
            countUnik = site.objects(unik_id=unik_id).count()
            countOff = site_offair.objects(unik_id=unik_id).count()

            if countOff == 0:
                curErr.append("Unik id tidak valid")

            if countUnik > 0:
                curErr.append("Titik sudah ada")

            nama_lokasi = curSite["nama_lokasi"]
            countNam = site.objects(nama=nama_lokasi).count()

            if countNam > 0:
                curErr.append("Nama Lokasi sudah ada")

            curProv = curSite["provinsi"]
            countProv = provinsi.objects(name=curProv.upper()).count()

            if countProv == 0:
                curErr.append("Provinsi tidak valid")

            curKab = curSite["kab_kota"]
            countKab = kabupaten.objects(name=curKab.upper()).count()

            if countKab == 0:
                countKab = kota.objects(name=curKab.upper()).count()
                if countKab == 0:
                    curErr.append("Kab/Kota tidak valid")

            curKec = curSite["kecamatan"]
            countKec = kecamatan.objects(name=curKec.upper()).count()

            if countKec == 0:
                curErr.append("Kecamatan tidak valid")

            curDesa = curSite["desa"]
            countDesa = desa.objects(name=curDesa.upper()).count()

            if countDesa == 0:
                curErr.append("Desa tidak valid")

            retData.append(curErr)

        return Response.ok(
            values=json.loads(json.dumps(retData, default=str)),
            message="Berhasil"
        )
    except Exception as e:
        # print(e)
        return Response.badRequest(
            values=[],
            message=str(e)
        )
